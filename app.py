import os
import json
import io
import logging
import psycopg2
import psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta, datetime
from functools import wraps
from flask import (Flask, render_template, request, redirect,
                   url_for, jsonify, g, Response, session, flash)
import bcrypt
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

app = Flask(__name__)

# ── Secret key — required at startup ─────────────────────────────────────────
_SECRET_KEY      = os.environ.get('SECRET_KEY', '').strip()
_KNOWN_WEAK_KEYS = {
    'alsondos-secret-change-in-production-2024',
    'secret', 'dev', 'development', 'test', 'changeme', '',
}
if not _SECRET_KEY or _SECRET_KEY in _KNOWN_WEAK_KEYS:
    raise RuntimeError(
        "FATAL: SECRET_KEY environment variable is not set or uses an insecure default.\n"
        "Generate a strong key and add it to /etc/sales.env:\n"
        "  python -c \"import secrets; print(secrets.token_hex(32))\""
    )
app.secret_key = _SECRET_KEY

# ── Session security ──────────────────────────────────────────────────────────
app.permanent_session_lifetime    = timedelta(hours=8)
app.config['SESSION_COOKIE_SECURE']   = True   # HTTPS only
app.config['SESSION_COOKIE_HTTPONLY'] = True   # no JS access to cookie
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF mitigation layer 2

# ── CSRF protection (covers all POST forms + AJAX via X-CSRFToken header) ─────
csrf = CSRFProtect(app)

# ── Rate limiter (brute-force protection) ─────────────────────────────────────
limiter = Limiter(key_func=get_remote_address, default_limits=[])
limiter.init_app(app)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

PER_PAGE = 50  # rows per page

# ── Database helpers ──────────────────────────────────────────────────────────
def get_db():
    """
    Get a per-request database connection.
    Auto-heals only when the connection is in INERROR state
    (transaction aborted due to a failed query).
    Does NOT roll back STATUS_IN_TRANSACTION — that is the normal
    idle state between queries and rolling it back would destroy
    valid uncommitted work from earlier in the same request.
    """
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = psycopg2.connect(
            os.environ.get("DATABASE_URL"),
            cursor_factory=psycopg2.extras.RealDictCursor
        )
        db.autocommit = False
    else:
        try:
            txn_status = db.get_transaction_status()
            if txn_status == psycopg2.extensions.TRANSACTION_STATUS_INERROR:
                # Previous query failed — rollback to recover the connection
                db.rollback()
        except Exception:
            # Connection is broken entirely — create a fresh one
            try: db.close()
            except: pass
            db = g._database = psycopg2.connect(
                os.environ.get("DATABASE_URL"),
                cursor_factory=psycopg2.extras.RealDictCursor
            )
            db.autocommit = False
    return db

@app.teardown_appcontext
def close_connection(exception):
    """
    Teardown: rollback any uncommitted work then close.
    NEVER commits here — execute_db() is the only place that commits.
    Double-committing causes 'no transaction in progress' errors on the
    next request that reuses the connection.
    """
    db = getattr(g, '_database', None)
    if db is not None:
        try:
            db.rollback()   # clear any uncommitted state (safe even if nothing pending)
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass
        g._database = None

def query_db(query, args=(), one=False):
    db  = None
    try:
        db  = get_db()
        cur = db.cursor()
        cur.execute(query, args)
        rv  = cur.fetchall()
        return (rv[0] if rv else None) if one else rv
    except Exception as e:
        if db:
            try: db.rollback()
            except: pass
        logger.error(f"query_db error: {e} | query: {query[:120]}")
        raise

def execute_db(query, args=()):
    """Execute INSERT/UPDATE/DELETE. Returns RETURNING value when present."""
    db  = None
    cur = None
    try:
        db  = get_db()
        cur = db.cursor()
        cur.execute(query, args)
        # CRITICAL: fetch RETURNING value BEFORE commit
        # psycopg2 closes the cursor result set after commit()
        result = None
        if 'RETURNING' in query.upper():
            row = cur.fetchone()
            if row:
                result = row[0] if not hasattr(row, 'keys') else (
                    row.get('id') or row[list(row.keys())[0]])
        db.commit()
        return result
    except Exception as e:
        # Use captured db ref — never call get_db() in except (may re-enter error)
        if db:
            try: db.rollback()
            except: pass
        logger.error(f"execute_db error: {e} | query: {query[:120]}")
        raise

def paginate(query, params, page, per_page=PER_PAGE):
    """Returns (rows, total_count, total_pages)"""
    count_q = f"SELECT COUNT(*) as cnt FROM ({query}) sub"
    total = query_db(count_q, params, one=True)['cnt']
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    paginated_q = query + f" LIMIT {per_page} OFFSET {(page-1)*per_page}"
    rows = query_db(paginated_q, params)
    return rows, total, total_pages

# ── Dashboard date-range filter helper ────────────────────────────────────────
_DASH_QUICK_FILTERS = ['today', 'yesterday', 'this_week', 'this_month',
                       'last_month', 'this_year', 'custom']

def _month_bounds(d):
    """Return (first_day, last_day) of the calendar month containing d."""
    start = d.replace(day=1)
    if d.month == 12:
        end = d.replace(month=12, day=31)
    else:
        end = (d.replace(month=d.month + 1, day=1)) - timedelta(days=1)
    return start, end

def resolve_dashboard_dates(args):
    """
    Resolve the dashboard's date_from/date_to/quick filter from request args.
    Precedence: explicit quick=... shortcut > explicit date_from/date_to >
    default (current month). Returns (date_from, date_to, quick) as strings.
    """
    today = date.today()
    quick     = (args.get('quick', '') or '').strip().lower()
    raw_from  = (args.get('date_from', '') or '').strip()
    raw_to    = (args.get('date_to', '') or '').strip()

    if quick and quick in _DASH_QUICK_FILTERS and quick != 'custom':
        if quick == 'today':
            start = end = today
        elif quick == 'yesterday':
            start = end = today - timedelta(days=1)
        elif quick == 'this_week':
            start = today - timedelta(days=today.weekday())  # Monday start
            end = today
        elif quick == 'this_month':
            start, end = _month_bounds(today)
        elif quick == 'last_month':
            last_month_day = today.replace(day=1) - timedelta(days=1)
            start, end = _month_bounds(last_month_day)
        elif quick == 'this_year':
            start, end = today.replace(month=1, day=1), today.replace(month=12, day=31)
        else:
            start, end = _month_bounds(today)
        return str(start), str(end), quick

    if raw_from or raw_to:
        # Custom range — fill any missing side with a sane bound rather than
        # leaving it open-ended, so every query gets two concrete dates.
        start = raw_from or str(today.replace(day=1))
        end   = raw_to or str(today)
        return start, end, 'custom'

    # Default view: current month
    start, end = _month_bounds(today)
    return str(start), str(end), 'this_month'


def init_db():
    db = psycopg2.connect(os.environ.get("DATABASE_URL"))
    db.autocommit = True
    cur = db.cursor()

    cur.execute('''
        CREATE TABLE IF NOT EXISTS sales (
            id          SERIAL PRIMARY KEY,
            from_loc    TEXT NOT NULL,
            to_loc      TEXT NOT NULL,
            via         TEXT DEFAULT '',
            trip_type   TEXT DEFAULT '',
            buy_from    TEXT DEFAULT '',
            company     TEXT NOT NULL,
            tickets     INTEGER DEFAULT 1,
            customer    TEXT NOT NULL,
            sale_date   TEXT NOT NULL,
            travel_date TEXT DEFAULT '',
            net         REAL NOT NULL DEFAULT 0,
            sell        REAL NOT NULL DEFAULT 0,
            profit      REAL NOT NULL DEFAULT 0,
            status      TEXT DEFAULT 'STILL',
            remarks     TEXT DEFAULT '',
            deleted     BOOLEAN DEFAULT FALSE,
            created_at  TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS payments (
            id          SERIAL PRIMARY KEY,
            company     TEXT NOT NULL,
            amount      REAL NOT NULL,
            pay_date    TEXT NOT NULL,
            notes       TEXT DEFAULT '',
            deleted     BOOLEAN DEFAULT FALSE,
            created_at  TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id            SERIAL PRIMARY KEY,
            username      TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role          TEXT NOT NULL DEFAULT 'user',
            created_at    TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS audit_logs (
            id          SERIAL PRIMARY KEY,
            user_id     INTEGER REFERENCES users(id) ON DELETE SET NULL,
            username    TEXT NOT NULL,
            action      TEXT NOT NULL,
            table_name  TEXT NOT NULL,
            record_id   INTEGER,
            detail      TEXT DEFAULT '',
            created_at  TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    ''')

    # Add deleted column FIRST (before indexes) — handles old schema upgrades
    for tbl in ('sales', 'payments'):
        cur.execute(f"""
            DO $$ BEGIN
                ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS deleted BOOLEAN DEFAULT FALSE;
            EXCEPTION WHEN duplicate_column THEN NULL; END $$;
        """)
        cur.execute(f"""
            DO $$ BEGIN
                ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS is_archived BOOLEAN DEFAULT FALSE;
            EXCEPTION WHEN duplicate_column THEN NULL; END $$;
        """)

    # New ticket tracking columns
    new_cols = [
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS return_date         TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS return_supplier     TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS outbound_delivery   TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS return_delivery     TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS outbound_status     TEXT DEFAULT 'PENDING'",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS return_status       TEXT DEFAULT 'PENDING'",
    ]
    for col in new_cols:
        cur.execute(f"DO $$ BEGIN {col}; EXCEPTION WHEN duplicate_column THEN NULL; END $$;")

    db.commit()

    # Indexes for performance
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_sales_company    ON sales(company)",
        "CREATE INDEX IF NOT EXISTS idx_sales_sale_date  ON sales(sale_date)",
        "CREATE INDEX IF NOT EXISTS idx_sales_status     ON sales(status)",
        "CREATE INDEX IF NOT EXISTS idx_sales_travel     ON sales(travel_date)",
        "CREATE INDEX IF NOT EXISTS idx_sales_deleted    ON sales(deleted)",
        "CREATE INDEX IF NOT EXISTS idx_sales_archived   ON sales(is_archived)",
        "CREATE INDEX IF NOT EXISTS idx_payments_archived ON payments(is_archived)",
        "CREATE INDEX IF NOT EXISTS idx_payments_company ON payments(company)",
        "CREATE INDEX IF NOT EXISTS idx_payments_date    ON payments(pay_date)",
        "CREATE INDEX IF NOT EXISTS idx_audit_user       ON audit_logs(user_id)",
        "CREATE INDEX IF NOT EXISTS idx_audit_table      ON audit_logs(table_name)",
        "CREATE INDEX IF NOT EXISTS idx_audit_created    ON audit_logs(created_at)",
        # Employee dashboard / my_sales — filters by user
        "CREATE INDEX IF NOT EXISTS idx_sales_user       ON sales(created_by_user_id) WHERE deleted=FALSE",
        # Statement queries — always filter by company + date together
        "CREATE INDEX IF NOT EXISTS idx_sales_co_date    ON sales(company, sale_date DESC) WHERE deleted=FALSE",
        # Accounting entries — queried by ref on every edit/delete
        "CREATE INDEX IF NOT EXISTS idx_acct_ref         ON accounting_entries(ref_type, ref_id, ref_table) WHERE is_reversed=FALSE",
        # Supplier payments lookup
        "CREATE INDEX IF NOT EXISTS idx_sup_pay_supplier ON supplier_payments(supplier) WHERE deleted=FALSE",
    ]
    for idx in indexes:
        cur.execute(idx)

    db.commit()

    cur.execute('SELECT COUNT(*) FROM users')
    if cur.fetchone()[0] == 0:
        import secrets as _sec
        _tmp_pw  = _sec.token_urlsafe(14)
        pw_hash  = bcrypt.hashpw(_tmp_pw.encode(), bcrypt.gensalt()).decode()
        cur.execute("INSERT INTO users (username, password_hash, role) VALUES (%s,%s,%s)",
                    ('admin', pw_hash, 'admin'))
        db.commit()
        print(f"✅ First-run admin created. Username: admin  Password: {_tmp_pw}")
        print("⚠️  Change this password immediately after first login.")

    cur.execute('SELECT COUNT(*) FROM sales')
    if cur.fetchone()[0] == 0:
        seed_file = os.path.join(os.path.dirname(__file__), 'seed_data.json')
        if os.path.exists(seed_file):
            with open(seed_file) as f:
                rows = json.load(f)
            for row in rows:
                cur.execute('''
                    INSERT INTO sales
                    (from_loc,to_loc,via,trip_type,buy_from,company,tickets,
                     customer,sale_date,travel_date,net,sell,profit,status,remarks)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ''', (row['from_loc'],row['to_loc'],row['via'],row['trip_type'],
                      row['buy_from'],row['company'],row['tickets'],row['customer'],
                      row['sale_date'],row['travel_date'],row['net'],row['sell'],
                      row['profit'],row['status'],row['remarks']))
            db.commit()
            print(f"✅ Seeded {len(rows)} records")

    cur.close()
    db.close()

try:
    init_db()
    logger.info("DB initialised OK")
except Exception as _db_err:
    logger.error(f"init_db error: {_db_err}")

# ── Audit log helper ──────────────────────────────────────────────────────────
def _get_client_ip():
    """
    Return the real client IP.
    Behind Nginx the actual IP arrives in X-Forwarded-For; remote_addr is 127.0.0.1.
    We take only the first (leftmost) entry to avoid spoofing via chained proxies.
    """
    try:
        xff = request.headers.get('X-Forwarded-For', '')
        if xff:
            return xff.split(',')[0].strip()
        return request.remote_addr or ''
    except Exception:
        return ''

def log_action(action, table_name, record_id=None, detail=''):
    """
    Write an audit log entry.
    SEC-07: Captures client IP address for every action.
    detail should be a human-readable summary; structured data can be embedded as JSON.
    """
    try:
        uid   = session.get('user_id')
        uname = session.get('username', 'system')
        ip    = _get_client_ip()
        execute_db('''
            INSERT INTO audit_logs (user_id, username, action, table_name, record_id, detail, ip_address)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        ''', (uid, uname, action, table_name, record_id, detail, ip))
    except Exception as e:
        logger.error(f"Audit log failed: {e}")

def post_ledger(ref_type, ref_id, ref_table, company, description, debit, credit, entry_date=None):
    """
    Post an immutable accounting entry to the shadow ledger.
    Never updates or deletes — use is_reversed=TRUE for corrections.
    Called after every sale, payment, supplier payment, and refund.
    Returns the new entry id, or None on failure.
    """
    try:
        entry_id = execute_db("""
            INSERT INTO accounting_entries
            (entry_date, ref_type, ref_id, ref_table, company, description, debit, credit, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            entry_date or str(date.today()),
            ref_type,
            ref_id,
            ref_table,
            company.upper().strip() if company else '',
            description or '',
            round(float(debit or 0), 3),
            round(float(credit or 0), 3),
            session.get('user_id') if 'user_id' in session else None,
        ))
        return entry_id
    except Exception as _le:
        logger.error(f"Ledger post failed ({ref_type}/{ref_id}): {_le}")
        # Never crash the main flow — ledger is supplementary
        return None


def post_sale_items(sale_id, service_type, supplier, description, net_cost, sell_price, quantity=1):
    """Write a sale_item row. Called after every new sale."""
    try:
        execute_db("""
            INSERT INTO sale_items (sale_id, service_type, supplier, description, net_cost, sell_price, quantity)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            sale_id,
            (service_type or 'FLIGHT').upper(),
            (supplier or '').upper().strip(),
            description or '',
            round(float(net_cost or 0), 3),
            round(float(sell_price or 0), 3),
            int(quantity or 1),
        ))
    except Exception as _si:
        logger.error(f"sale_items post failed (sale {sale_id}): {_si}")


def validate_ledger_balance(company):
    """
    Compare shadow ledger balance vs dynamic calculation.
    Returns (ledger_balance, dynamic_balance, is_ok, detail_dict).

    Dynamic balance formula:
      total_sold_to_company          (sales.sell   where company=X)
      - total_received_from_company  (payments.amount where company=X)
      + total_purchased_from_supplier(outbound_cost/net where buy_from=X)
      - total_paid_to_supplier       (supplier_payments.amount where supplier=X)

    Positive = they owe us (net receivable).
    Negative = we owe them (net payable).

    ACC-02 FIX: previous version had identical supplier_payments sub-queries on
    both + and - sides, causing them to cancel to zero (always wrong).
    """
    try:
        ledger_row = query_db("""
            SELECT
                COALESCE(SUM(debit),  0) AS total_debit,
                COALESCE(SUM(credit), 0) AS total_credit,
                COALESCE(SUM(debit - credit), 0) AS bal
            FROM accounting_entries
            WHERE UPPER(company)=UPPER(%s) AND is_reversed=FALSE
        """, [company], one=True)
        ledger_bal   = float(ledger_row['bal']          if ledger_row else 0)
        ledger_dr    = float(ledger_row['total_debit']  if ledger_row else 0)
        ledger_cr    = float(ledger_row['total_credit'] if ledger_row else 0)

        # Dynamic components — each calculated independently.
        # ACC-03: exclude archived records to stay consistent with what statement() shows.
        sold_row = query_db(
            "SELECT COALESCE(SUM(sell),0) AS v FROM sales WHERE UPPER(company)=UPPER(%s) AND deleted=FALSE AND is_archived=FALSE",
            [company], one=True)
        rcvd_row = query_db(
            "SELECT COALESCE(SUM(amount),0) AS v FROM payments WHERE UPPER(company)=UPPER(%s) AND deleted=FALSE AND is_archived=FALSE",
            [company], one=True)
        # Outbound cost — what we paid this supplier for the outbound leg
        purch_row = query_db(
            """SELECT COALESCE(SUM(
                   CASE WHEN outbound_cost > 0 THEN outbound_cost ELSE COALESCE(net,0) END
               ),0) AS v
               FROM sales WHERE UPPER(buy_from)=UPPER(%s) AND deleted=FALSE AND is_archived=FALSE""",
            [company], one=True)
        # Return cost — what we paid this supplier for the return leg when they differ from outbound supplier
        # Include same-supplier round trips too (return_cost on same buy_from) — no exclusion needed since
        # outbound_cost and return_cost are separate fields with no double-counting risk.
        ret_purch_row = query_db(
            """SELECT COALESCE(SUM(return_cost), 0) AS v
               FROM sales
               WHERE UPPER(return_supplier)=UPPER(%s)
                 AND UPPER(COALESCE(buy_from,'')) <> UPPER(%s)
                 AND return_cost > 0
                 AND deleted=FALSE AND is_archived=FALSE""",
            [company, company], one=True)
        paid_row = query_db(
            "SELECT COALESCE(SUM(amount),0) AS v FROM supplier_payments WHERE UPPER(supplier)=UPPER(%s) AND deleted=FALSE AND is_archived=FALSE",
            [company], one=True)

        total_sold     = float(sold_row['v']      if sold_row      else 0)
        total_received = float(rcvd_row['v']      if rcvd_row      else 0)
        total_purchased= float(purch_row['v']     if purch_row     else 0)
        total_purchased+= float(ret_purch_row['v'] if ret_purch_row else 0)
        total_paid_out = float(paid_row['v']      if paid_row      else 0)

        dynamic_bal = total_sold - total_received + total_purchased - total_paid_out
        dynamic_bal = round(dynamic_bal, 3)

        is_ok = abs(ledger_bal - dynamic_bal) < 0.01
        if not is_ok:
            logger.warning(
                f"Balance mismatch [{company}]: ledger={ledger_bal:.3f} dynamic={dynamic_bal:.3f} "
                f"(sold={total_sold:.3f} rcvd={total_received:.3f} "
                f"purch={total_purchased:.3f} paid_out={total_paid_out:.3f})"
            )
        detail = {
            'total_sold':      total_sold,
            'total_received':  total_received,
            'total_purchased': total_purchased,
            'total_paid_out':  total_paid_out,
            'ledger_debit':    ledger_dr,
            'ledger_credit':   ledger_cr,
        }
        return ledger_bal, dynamic_bal, is_ok, detail
    except Exception as _e:
        logger.error(f"validate_ledger_balance failed: {_e}")
        return 0, 0, False, {}


def compute_ticket_status(outbound_delivery, return_delivery):
    """Auto-compute outbound/return status based on today's date."""
    today_str = str(date.today())
    outbound_status = 'DONE' if outbound_delivery and today_str >= outbound_delivery else 'PENDING'
    return_status   = 'DONE' if return_delivery  and today_str >= return_delivery  else 'PENDING'
    # Overall status: DONE only when all applicable sectors done
    if return_delivery:
        overall = 'DONE' if outbound_status == 'DONE' and return_status == 'DONE' else 'STILL'
    else:
        overall = 'DONE' if outbound_status == 'DONE' else 'STILL'
    return outbound_status, return_status, overall

# ── Input validation ──────────────────────────────────────────────────────────
def validate_sale_form(form):
    errors = []
    svc = form.get('service_type', 'FLIGHT').upper().strip()
    # from/to only required for flight-based types
    if svc in ('FLIGHT', 'PACKAGE'):
        if not form.get('from_loc','').strip():
            errors.append('From location is required for flights.')
        if not form.get('to_loc','').strip():
            errors.append('To location is required for flights.')
    _ctype = form.get('customer_type', 'COMPANY').upper().strip()
    if _ctype != 'DIRECT' and not form.get('company','').strip():
        errors.append('Company is required.')
    if not form.get('customer','').strip():
        errors.append('Customer name is required.')
    if not form.get('sale_date','').strip():
        errors.append('Sale date is required.')
    try:
        net  = float(form.get('net', 0) or 0)
        sell = float(form.get('sell', 0) or 0)
        if net < 0:  errors.append('Net cost cannot be negative.')
        if sell < 0: errors.append('Sell price cannot be negative.')
    except ValueError:
        errors.append('Net and Sell must be valid numbers.')
    try:
        tickets = int(form.get('tickets', 1) or 1)
        if tickets < 1: errors.append('Tickets must be at least 1.')
    except ValueError:
        errors.append('Tickets must be a valid number.')
    return errors

# ── Auth helpers ──────────────────────────────────────────────────────────────
def get_current_user():
    if 'user_id' not in session:
        return None
    try:
        return query_db('SELECT * FROM users WHERE id=%s', [session['user_id']], one=True)
    except Exception:
        return None

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in.', 'warning')
            return redirect(url_for('login'))
        if session.get('user_role') != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('employee_dashboard'))
        return f(*args, **kwargs)
    return decorated

def finance_required(f):
    """Allow admin and manager roles; block plain sales agents."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in.', 'warning')
            return redirect(url_for('login'))
        if session.get('user_role') not in ('admin', 'manager'):
            flash('This page is restricted to management.', 'danger')
            return redirect(url_for('employee_dashboard'))
        return f(*args, **kwargs)
    return decorated

@app.before_request
def _enforce_force_password_change():
    """
    Gate page navigation (GET only) when an admin has flagged this account
    with "force password change on next login". Scoped to GET requests only
    so it never interferes with existing POST/AJAX actions across the app.
    """
    if not session.get('must_change_password'):
        return None
    if request.endpoint in ('force_change_password', 'logout', 'login', 'static', None):
        return None
    if request.method != 'GET':
        return None
    return redirect(url_for('force_change_password'))

@app.context_processor
def inject_user():
    try:
        cur_user = get_current_user()
    except Exception:
        cur_user = None
    return {
        'current_user': cur_user,
        'is_admin': session.get('user_role') == 'admin',
        'logged_in': 'user_id' in session
    }

@app.template_filter('from_json')
def from_json_filter(value):
    import json
    try:
        return json.loads(value or '[]')
    except Exception:
        return []

@app.template_filter('svc_icon')
def svc_icon_filter(svc):
    icons = {'FLIGHT':'fa-plane','HOTEL':'fa-hotel','PACKAGE':'fa-box-open',
             'VISA':'fa-passport','INSURANCE':'fa-shield-alt',
             'TRANSFER':'fa-car','TOUR':'fa-map-marked-alt'}
    return icons.get((svc or 'FLIGHT').upper(), 'fa-exchange-alt')

@app.template_filter('svc_color')
def svc_color_filter(svc):
    colors = {'FLIGHT':'#1B3A6B','HOTEL':'#C8A84B','PACKAGE':'#8b1a1a',
              'VISA':'#6D28D9','INSURANCE':'#1E7B34',
              'TRANSFER':'#D97706','TOUR':'#2980B9'}
    return colors.get((svc or 'FLIGHT').upper(), '#6B7A99')

@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404,
        title='Page Not Found',
        msg='The page you are looking for does not exist or has been moved.'), 404

@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', code=403,
        title='Access Denied',
        msg='You do not have permission to access this page. Contact your administrator.'), 403

@app.errorhandler(500)
def server_error(e):
    logger.error(f"500 error: {e}", exc_info=True)
    return render_template('error.html', code=500,
        title='System Error',
        msg='Something went wrong. The issue has been logged. Please try again or contact your administrator.'), 500

from flask_wtf.csrf import CSRFError
@app.errorhandler(CSRFError)
def csrf_error(e):
    flash('Your session has expired or the form timed out. Please try again.', 'warning')
    return redirect(request.referrer or url_for('login'))

# ── Auth Routes ───────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute", error_message="Too many login attempts. Please wait a minute and try again.")
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '').encode()
        if not username or not password:
            flash('Username and password are required.', 'danger')
            return render_template('login.html', next=request.args.get('next',''))
        user = query_db('SELECT * FROM users WHERE username=%s', [username], one=True)
        if user and bcrypt.checkpw(password, user['password_hash'].encode()):
            # SEC-01: reject disabled accounts BEFORE granting session
            if not user.get('is_active', True):
                log_action('LOGIN_DENIED', 'users', user['id'],
                           f"Login blocked — account disabled: {username}")
                flash('Your account has been disabled. Please contact your administrator.', 'danger')
                return render_template('login.html', next=request.args.get('next', ''))
            session.clear()
            session['user_id']   = user['id']
            session['username']  = user['username']
            session['user_role'] = user['role']
            session['must_change_password'] = bool(user.get('must_change_password'))
            session.permanent    = True   # uses app.permanent_session_lifetime = 8h
            execute_db("UPDATE users SET last_login=%s WHERE id=%s",
                      (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user['id']))
            log_action('LOGIN', 'users', user['id'], f"User {username} logged in")
            if session['must_change_password']:
                return redirect(url_for('force_change_password'))
            # SEC-06: validate next= is a safe relative URL (prevent open redirect)
            _next = request.form.get('next', '').strip()
            if _next and _next.startswith('/') and not _next.startswith('//'):
                next_page = _next
            elif user['role'] not in ('admin', 'manager'):
                next_page = url_for('employee_dashboard')
            else:
                next_page = url_for('index')
            return redirect(next_page)
        flash('Invalid username or password.', 'danger')
    return render_template('login.html', next=request.args.get('next', ''))

@app.route('/logout')
def logout():
    log_action('LOGOUT', 'users', session.get('user_id'), f"User {session.get('username')} logged out")
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# ── Employee Management (extends the existing Users module — NOT a separate
#    HR system; `users` remains the single source of truth for login) ────────
_EMP_STATUSES    = ['Active', 'On Leave', 'Suspended', 'Resigned']
_EMP_POSITIONS   = ['General Manager', 'Operation Manager', 'Sales Executive',
                    'Accountant', 'Operations Officer', 'Administrator']
_EMP_DEPARTMENTS = ['Management', 'Operations', 'Sales', 'Finance', 'Administration']


@app.route('/users')
@admin_required
def manage_users():
    q_search = request.args.get('q', '').strip()
    q_role   = request.args.get('role', '').strip()
    q_dept   = request.args.get('department', '').strip()
    q_pos    = request.args.get('position', '').strip()
    q_status = request.args.get('status', '').strip()

    where, params = ['1=1'], []
    if q_search:
        where.append("(LOWER(u.username) LIKE %s OR LOWER(COALESCE(u.full_name,'')) LIKE %s)")
        params.extend([f'%{q_search.lower()}%', f'%{q_search.lower()}%'])
    if q_role:
        where.append("u.role=%s"); params.append(q_role)
    if q_dept:
        where.append("u.department=%s"); params.append(q_dept)
    if q_pos:
        where.append("u.position=%s"); params.append(q_pos)
    if q_status:
        where.append("COALESCE(u.employment_status,'Active')=%s"); params.append(q_status)

    users = query_db(f"""
        SELECT u.id, u.username, u.role, u.created_at,
               COALESCE(u.full_name,'') as full_name,
               COALESCE(u.is_active, TRUE) as is_active,
               COALESCE(u.commission_rate, 20.0) as commission_rate,
               COALESCE(u.position,'') as position,
               COALESCE(u.department,'') as department,
               COALESCE(u.employment_status,'Active') as employment_status,
               COALESCE(u.join_date,'') as join_date,
               COALESCE(u.last_login,'') as last_login,
               COALESCE(u.must_change_password, FALSE) as must_change_password,
               m.username as manager_username,
               COALESCE(m.full_name, m.username) as manager_name,
               COALESCE(s.txn_count, 0)     as txn_count,
               COALESCE(s.total_sell, 0)    as total_sell,
               COALESCE(s.total_profit, 0)  as total_profit
        FROM users u
        LEFT JOIN users m ON m.id = u.manager_id
        LEFT JOIN (
            SELECT created_by_user_id,
                   COUNT(*) as txn_count,
                   SUM(sell) as total_sell,
                   SUM(profit) as total_profit
            FROM sales WHERE deleted=FALSE
            GROUP BY created_by_user_id
        ) s ON u.id = s.created_by_user_id
        WHERE {' AND '.join(where)}
        ORDER BY u.id
    """, params)
    users = [dict(u) for u in users]
    for u in users:
        r = float(u.get('commission_rate') or 20)
        u['commission'] = round(float(u.get('total_profit') or 0) * r / 100, 2)

    return render_template('users.html',
        users=users, statuses=_EMP_STATUSES,
        positions=_EMP_POSITIONS, departments=_EMP_DEPARTMENTS,
        filters=dict(q=q_search, role=q_role, department=q_dept,
                     position=q_pos, status=q_status),
    )


@app.route('/users/add', methods=['POST'])
@admin_required
def add_user():
    username = request.form.get('username', '').strip().lower()
    password = request.form.get('password', '')
    role     = request.form.get('role', 'user')
    if not username or not password:
        flash('Username and password are required.', 'danger')
        return redirect(url_for('manage_users'))
    if len(password) < 6:
        flash('Password must be at least 6 characters.', 'danger')
        return redirect(url_for('manage_users'))
    if query_db('SELECT id FROM users WHERE username=%s', [username], one=True):
        flash(f'Username "{username}" already exists.', 'danger')
        return redirect(url_for('manage_users'))

    full_name  = request.form.get('full_name', '').strip()
    position   = request.form.get('position', '').strip()
    department = request.form.get('department', '').strip()
    manager_id = request.form.get('manager_id') or None

    pw_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    uid = execute_db("""
        INSERT INTO users (username, password_hash, role, full_name, position, department,
                            manager_id, join_date, employment_status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'Active')
    """, (username, pw_hash, role, full_name, position, department, manager_id, str(date.today())))
    log_action('CREATE', 'users', uid, f"Created user {username} with role {role}")
    flash(f'User "{username}" created successfully.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    u = query_db('SELECT * FROM users WHERE id=%s', [user_id], one=True)
    if not u:
        flash('User not found.', 'danger')
        return redirect(url_for('manage_users'))

    if request.method == 'POST':
        f = request.form
        new_username = f.get('username', '').strip().lower()
        new_role     = f.get('role', u['role'])
        manager_id   = f.get('manager_id') or None

        if not new_username:
            flash('Username is required.', 'danger')
            return redirect(url_for('edit_user', user_id=user_id))
        if manager_id and int(manager_id) == user_id:
            flash('A user cannot report to themselves.', 'danger')
            return redirect(url_for('edit_user', user_id=user_id))
        if new_username != u['username']:
            dup = query_db('SELECT id FROM users WHERE username=%s AND id!=%s', [new_username, user_id], one=True)
            if dup:
                flash(f'Username "{new_username}" is already taken.', 'danger')
                return redirect(url_for('edit_user', user_id=user_id))

        execute_db("""
            UPDATE users SET
              username=%s, full_name=%s, email=%s, phone=%s,
              position=%s, department=%s, manager_id=%s, join_date=%s,
              employment_status=%s, notes=%s, role=%s, commission_rate=%s
            WHERE id=%s
        """, (
            new_username,
            f.get('full_name', '').strip(),
            f.get('email', '').strip(),
            f.get('phone', '').strip(),
            f.get('position', '').strip(),
            f.get('department', '').strip(),
            manager_id,
            f.get('join_date', '').strip(),
            f.get('employment_status', 'Active'),
            f.get('notes', '').strip(),
            new_role,
            float(f.get('commission_rate', 20) or 20),
            user_id,
        ))

        # Distinct audit entries for sensitive changes (per spec: track role changes,
        # username changes — separate from the general "profile updated" entry).
        if new_username != u['username']:
            log_action('UPDATE', 'users', user_id, f"Username changed: {u['username']} → {new_username}")
        if new_role != u['role']:
            log_action('ROLE_CHANGE', 'users', user_id, f"Role changed for {new_username}: {u['role']} → {new_role}")
        log_action('UPDATE', 'users', user_id, f"Profile updated for {new_username}")

        flash('User updated successfully.', 'success')
        return redirect(url_for('user_profile', user_id=user_id))

    managers = query_db(
        "SELECT id, username, COALESCE(full_name, username) as full_name FROM users WHERE id!=%s ORDER BY full_name",
        [user_id]
    )
    return render_template('user_edit.html',
        u=u, managers=managers, statuses=_EMP_STATUSES,
        positions=_EMP_POSITIONS, departments=_EMP_DEPARTMENTS,
    )


@app.route('/users/<int:user_id>/profile')
@admin_required
def user_profile(user_id):
    u = query_db("""
        SELECT us.*, m.username as manager_username,
               COALESCE(m.full_name, m.username) as manager_name
        FROM users us
        LEFT JOIN users m ON m.id = us.manager_id
        WHERE us.id=%s
    """, [user_id], one=True)
    if not u:
        flash('User not found.', 'danger')
        return redirect(url_for('manage_users'))

    stats = query_db("""
        SELECT COUNT(*) as txn_count,
               COALESCE(SUM(sell),0) as total_sell,
               COALESCE(SUM(profit),0) as total_profit
        FROM sales WHERE deleted=FALSE AND created_by_user_id=%s
    """, [user_id], one=True) or {}
    rate = float(u.get('commission_rate') or 20)
    commission = round(float(stats.get('total_profit') or 0) * rate / 100, 2)

    direct_reports = query_db(
        "SELECT id, username, COALESCE(full_name, username) as full_name, role "
        "FROM users WHERE manager_id=%s ORDER BY full_name",
        [user_id]
    )

    login_history = query_db("""
        SELECT action, detail, created_at FROM audit_logs
        WHERE user_id=%s AND action IN ('LOGIN','LOGIN_DENIED','LOGOUT')
        ORDER BY created_at DESC LIMIT 25
    """, [user_id])

    activity_log = query_db("""
        SELECT action, table_name, detail, created_at FROM audit_logs
        WHERE user_id=%s
        ORDER BY created_at DESC LIMIT 25
    """, [user_id])

    return render_template('user_profile.html',
        u=u, stats=stats, commission=commission,
        direct_reports=direct_reports,
        login_history=login_history, activity_log=activity_log,
    )


@app.route('/users/delete/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    """
    H-05 FIX: Soft-delete only — sets is_active=FALSE so the user cannot log in
    and their audit history and transaction attribution remain intact.
    Hard DELETE is never done; it would null out audit_logs.user_id entries.
    """
    if user_id == session.get('user_id'):
        flash('You cannot deactivate your own account.', 'danger')
        return redirect(url_for('manage_users'))
    u = query_db('SELECT username, is_active FROM users WHERE id=%s', [user_id], one=True)
    if not u:
        flash('User not found.', 'danger')
        return redirect(url_for('manage_users'))
    # Soft-delete: disable the account permanently (same as toggle but labelled as delete)
    execute_db("UPDATE users SET is_active=FALSE, employment_status='Resigned' WHERE id=%s", [user_id])
    log_action('DELETE', 'users', user_id,
               f"User {u['username']} deactivated (soft-delete). "
               f"Audit history and transactions preserved.")
    flash(f'User "{u["username"]}" has been deactivated. Their transaction history is preserved.', 'success')
    return redirect(url_for('manage_users'))

@app.route('/users/change-password', methods=['POST'])
@login_required
def change_password():
    current = request.form.get('current_password', '').encode()
    new_pw  = request.form.get('new_password', '')
    confirm = request.form.get('confirm_password', '')
    user = query_db('SELECT * FROM users WHERE id=%s', [session['user_id']], one=True)
    if not bcrypt.checkpw(current, user['password_hash'].encode()):
        flash('Current password is incorrect.', 'danger')
    elif new_pw != confirm:
        flash('New passwords do not match.', 'danger')
    elif len(new_pw) < 6:
        flash('Password must be at least 6 characters.', 'danger')
    else:
        pw_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
        execute_db("UPDATE users SET password_hash=%s, must_change_password=FALSE WHERE id=%s",
                   (pw_hash, session['user_id']))
        session['must_change_password'] = False
        log_action('UPDATE', 'users', session['user_id'], "Password changed")
        flash('Password changed successfully.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/users/force-change-password', methods=['GET', 'POST'])
@login_required
def force_change_password():
    """Standalone forced password-change flow for accounts flagged by an admin."""
    if not session.get('must_change_password'):
        return redirect(url_for('index'))
    if request.method == 'POST':
        new_pw  = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        if new_pw != confirm:
            flash('Passwords do not match.', 'danger')
        elif len(new_pw) < 8 or not any(c.isdigit() for c in new_pw):
            flash('Password must be at least 8 characters and contain at least one number.', 'danger')
        else:
            pw_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
            execute_db("UPDATE users SET password_hash=%s, must_change_password=FALSE WHERE id=%s",
                       (pw_hash, session['user_id']))
            session['must_change_password'] = False
            log_action('UPDATE', 'users', session['user_id'], "Password changed (forced)")
            flash('Password updated. You can now continue.', 'success')
            return redirect(url_for('index'))
    return render_template('force_change_password.html')


@app.route('/users/reset-pw/<int:uid>', methods=['POST'])
@admin_required
def admin_reset_pw(uid):
    new_pw = request.form.get('new_password','')
    force_change = request.form.get('force_change') == '1'
    if len(new_pw) < 8 or not any(c.isdigit() for c in new_pw):
        flash('Password must be at least 8 characters and contain at least one number.', 'danger')
        return redirect(url_for('manage_users'))
    pw_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
    execute_db("UPDATE users SET password_hash=%s, must_change_password=%s WHERE id=%s",
              (pw_hash, force_change, uid))
    u = query_db('SELECT username FROM users WHERE id=%s', [uid], one=True)
    suffix = ' (must change on next login)' if force_change else ''
    log_action('UPDATE', 'users', uid, f"Password reset for {u['username'] if u else uid}{suffix}")
    flash(f'Password reset successfully.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/users/<int:uid>/force-change-toggle', methods=['POST'])
@admin_required
def toggle_force_password(uid):
    """Flag/unflag an account to require a password change on its next login."""
    flag = request.form.get('flag') == '1'
    execute_db("UPDATE users SET must_change_password=%s WHERE id=%s", (flag, uid))
    u = query_db('SELECT username FROM users WHERE id=%s', [uid], one=True)
    msg = 'will be required to change their password on next login' if flag else 'no longer required to change their password'
    log_action('UPDATE', 'users', uid, f"{u['username'] if u else uid} {msg}")
    flash('Updated.', 'success')
    return redirect(request.referrer or url_for('manage_users'))


@app.route('/users/toggle/<int:uid>', methods=['POST'])
@admin_required
def toggle_user(uid):
    if uid == session.get('user_id'):
        flash('You cannot disable your own account.', 'danger')
        return redirect(url_for('manage_users'))
    action = request.form.get('action', 'disable')
    active = (action == 'enable')
    execute_db('UPDATE users SET is_active=%s WHERE id=%s', (active, uid))
    u = query_db('SELECT username FROM users WHERE id=%s', [uid], one=True)
    msg = 'enabled' if active else 'disabled'
    log_action('UPDATE', 'users', uid, f"Account {msg}: {u['username'] if u else uid}")
    flash(f'Account {msg} successfully.', 'success')
    return redirect(url_for('manage_users'))

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route('/')
@login_required
def index():
    # Sales agents get their own scoped dashboard — only admins/managers see
    # the company-wide view rendered below.
    if session.get('user_role') not in ('admin', 'manager'):
        return redirect(url_for('employee_dashboard'))

    date_from, date_to, quick = resolve_dashboard_dates(request.args)
    drange = [date_from, date_to]

    stats = query_db('''
        SELECT COUNT(*) as total_transactions,
               COALESCE(SUM(sell),0) as total_sell,
               COALESCE(SUM(net),0) as total_net,
               COALESCE(SUM(profit),0) as total_profit
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND sale_date BETWEEN %s AND %s
    ''', drange, one=True)
    total_paid = query_db(
        'SELECT COALESCE(SUM(amount),0) as paid FROM payments '
        'WHERE deleted=FALSE AND is_archived=FALSE AND pay_date BETWEEN %s AND %s',
        drange, one=True
    )['paid']
    balance = (stats['total_sell'] or 0) - total_paid

    monthly = query_db('''
        SELECT to_char(to_date(sale_date,'YYYY-MM-DD'),'MM') as month,
               COALESCE(SUM(sell),0) as total_sell,
               COALESCE(SUM(profit),0) as total_profit,
               COUNT(*) as count
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND sale_date BETWEEN %s AND %s
        GROUP BY month ORDER BY month
    ''', drange)

    top_companies = query_db('''
        SELECT company, COALESCE(SUM(sell),0) as total, COUNT(*) as cnt
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND sale_date BETWEEN %s AND %s
        GROUP BY company ORDER BY total DESC LIMIT 10
    ''', drange)

    tomorrow_date = (date.today() + timedelta(days=1)).strftime('%Y-%m-%d')
    # Outbound tickets: travel_date or outbound_delivery = tomorrow
    tomorrow = query_db('''
        SELECT company, customer, from_loc, to_loc, travel_date,
               outbound_delivery, return_date, tickets, status, outbound_status
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND (
              (outbound_delivery != '' AND outbound_delivery = %s)
              OR (COALESCE(outbound_delivery,'') = '' AND travel_date = %s)
          )
          AND COALESCE(outbound_status,'PENDING') != 'DONE'
        ORDER BY company
    ''', [tomorrow_date, tomorrow_date])
    # Return tickets: return_date or return_delivery = tomorrow
    tomorrow_returns = query_db('''
        SELECT company, customer, to_loc, from_loc, return_date,
               return_delivery, return_supplier, tickets, status, return_status
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND trip_type IN ('RETURN','ROUNDTRIP','ROUND')
          AND (
              (return_delivery != '' AND return_delivery = %s)
              OR (COALESCE(return_delivery,'') = '' AND return_date = %s)
          )
          AND return_date != ''
          AND COALESCE(return_status,'PENDING') != 'DONE'
        ORDER BY company
    ''', [tomorrow_date, tomorrow_date]) or []

    # Recent activity from audit log
    recent_logs = query_db('''
        SELECT * FROM audit_logs ORDER BY created_at DESC LIMIT 8
    ''')

    companies = get_companies_list()

    # Status distribution for pie chart
    status_counts = query_db("""
        SELECT
            SUM(CASE WHEN status='DONE'   THEN 1 ELSE 0 END) as done_count,
            SUM(CASE WHEN status='STILL'  THEN 1 ELSE 0 END) as still_count,
            SUM(CASE WHEN status NOT IN ('DONE','STILL') THEN 1 ELSE 0 END) as other_count
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND sale_date BETWEEN %s AND %s
    """, drange, one=True)
    if status_counts:
        stats = dict(stats)
        stats['done_count']  = status_counts['done_count']  or 0
        stats['still_count'] = status_counts['still_count'] or 0
        stats['other_count'] = status_counts['other_count'] or 0

    # Service type breakdown for charts
    svc_breakdown = query_db("""
        SELECT COALESCE(service_type,'FLIGHT') as svc,
               COUNT(*) as cnt,
               COALESCE(SUM(sell),0) as total_sell,
               COALESCE(SUM(profit),0) as total_profit
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND sale_date BETWEEN %s AND %s
        GROUP BY svc ORDER BY total_sell DESC
    """, drange) or []
    svc_labels  = [r['svc']            for r in svc_breakdown]
    svc_counts  = [int(r['cnt'])        for r in svc_breakdown]
    svc_sells   = [float(r['total_sell'])   for r in svc_breakdown]
    svc_profits = [float(r['total_profit']) for r in svc_breakdown]

    # Outstanding balances by company - correct SQL with HAVING
    outstanding_by_company = query_db("""
        SELECT s.company,
               s.sell      AS total_sell,
               COALESCE(p.paid, 0) AS total_paid,
               s.sell - COALESCE(p.paid, 0) AS balance
        FROM (
            SELECT company, SUM(sell) AS sell
            FROM sales WHERE deleted=FALSE AND is_archived=FALSE
            GROUP BY company
        ) s
        LEFT JOIN (
            SELECT company, SUM(amount) AS paid
            FROM payments WHERE deleted=FALSE AND is_archived=FALSE
            GROUP BY company
        ) p ON s.company = p.company
        WHERE s.sell - COALESCE(p.paid, 0) > 0.01
        ORDER BY balance DESC
        LIMIT 10
    """) or []

    # Pre-process chart data as plain Python lists (avoids Decimal/RealDictRow JSON issues)
    chart_month_labels  = [str(r['month']) for r in monthly]
    chart_month_sells   = [float(r['total_sell'] or 0) for r in monthly]
    chart_month_profits = [float(r['total_profit'] or 0) for r in monthly]
    chart_company_names = [str(r['company']) for r in top_companies]
    chart_company_totals= [float(r['total'] or 0) for r in top_companies]
    has_monthly_data    = any(v > 0 for v in chart_month_sells)

    # Top 10 companies by total tickets purchased
    top_debtors = query_db("""
        SELECT company,
               COUNT(*)          AS txn_count,
               COALESCE(SUM(tickets),0) AS total_tickets,
               COALESCE(SUM(sell),0)    AS total_sell
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND sale_date BETWEEN %s AND %s
          AND TRIM(COALESCE(company,'')) <> ''
        GROUP BY company
        ORDER BY total_tickets DESC
        LIMIT 10
    """, drange) or []

    # Top 10 suppliers by purchase volume
    top_suppliers = query_db("""
        SELECT supplier,
               SUM(total_cost) AS total_cost,
               SUM(txn_count)  AS txn_count
        FROM (
            SELECT UPPER(TRIM(buy_from)) AS supplier,
                   COALESCE(SUM(CASE WHEN outbound_cost>0 THEN outbound_cost ELSE net END),0) AS total_cost,
                   COUNT(*) AS txn_count
            FROM sales
            WHERE deleted=FALSE AND is_archived=FALSE
              AND sale_date BETWEEN %s AND %s
              AND TRIM(COALESCE(buy_from,'')) <> ''
            GROUP BY UPPER(TRIM(buy_from))
            UNION ALL
            SELECT UPPER(TRIM(hotel_supplier)),
                   COALESCE(SUM(hotel_net),0),
                   COUNT(*)
            FROM sales
            WHERE deleted=FALSE AND is_archived=FALSE
              AND sale_date BETWEEN %s AND %s
              AND TRIM(COALESCE(hotel_supplier,'')) <> ''
            GROUP BY UPPER(TRIM(hotel_supplier))
        ) t
        GROUP BY supplier
        ORDER BY SUM(total_cost) DESC
        LIMIT 10
    """, drange + drange) or []

    # Recent transactions with agent names
    recent_txns = query_db("""
        SELECT s.*,
               COALESCE(u.full_name, s.created_by_username, '—') AS agent_display,
               COALESCE(s.created_by_username, '—') AS agent_name
        FROM sales s
        LEFT JOIN users u ON s.created_by_user_id = u.id
        WHERE s.deleted=FALSE AND s.is_archived=FALSE
          AND s.sale_date BETWEEN %s AND %s
        ORDER BY s.created_at DESC, s.id DESC
        LIMIT 12
    """, drange) or []

    return render_template('index.html',
        stats=stats, total_paid=total_paid, balance=balance,
        monthly=monthly, top_companies=top_companies,
        tomorrow=tomorrow, tomorrow_returns=tomorrow_returns, companies=companies,
        recent_logs=recent_logs,
        recent_txns=recent_txns,
        outstanding_by_company=outstanding_by_company,
        chart_month_labels=chart_month_labels,
        chart_month_sells=chart_month_sells,
        chart_month_profits=chart_month_profits,
        chart_company_names=chart_company_names,
        chart_company_totals=chart_company_totals,
        has_monthly_data=has_monthly_data,
        svc_labels=svc_labels,
        svc_counts=svc_counts,
        svc_sells=svc_sells,
        svc_profits=svc_profits,
        top_debtors=top_debtors,
        top_suppliers=top_suppliers,
        today=date.today().strftime('%d %B %Y'),
        date_from=date_from, date_to=date_to, quick=quick,
    )

# ── Sales ─────────────────────────────────────────────────────────────────────
# Phase 4: GET /add redirects to the active /add-v2 route.
# POST /add is kept for backward-compat (direct form submissions from older bookmarks).
@app.route('/add', methods=['GET', 'POST'])
@login_required
def add_sale():
    if request.method == 'GET':
        return redirect(url_for('add_sale_v2'), 301)
    try:
        companies = get_companies_list()
    except Exception:
        companies = []
    if request.method == 'POST':
        errors = validate_sale_form(request.form)
        if errors:
            for e in errors: flash(e, 'danger')
            return render_template('add.html', companies=companies,
                                   today=str(date.today()), form=request.form)
        net  = float(request.form.get('net', 0))
        sell = float(request.form.get('sell', 0))

        outbound_delivery = request.form.get('outbound_delivery','').strip()
        return_delivery   = request.form.get('return_delivery','').strip()
        return_date       = request.form.get('return_date','').strip()
        return_supplier   = request.form.get('return_supplier','').upper().strip()

        outbound_status, return_status, overall = compute_ticket_status(
            outbound_delivery, return_delivery
        )


        # ── Service-type specific fields ─────────────────────────────────────
        service_type = request.form.get('service_type','FLIGHT').upper().strip()
        # Tours: collect dynamic rows from form
        tour_names    = request.form.getlist('tour_name[]')
        tour_dates    = request.form.getlist('tour_date[]')
        tour_pickups  = request.form.getlist('tour_pickup[]')
        tour_statuses = request.form.getlist('tour_status[]')
        tour_suppliers= request.form.getlist('tour_supplier[]')
        tour_costs    = request.form.getlist('tour_cost[]')
        tour_notes    = request.form.getlist('tour_notes[]')
        tours_list = []
        for i in range(len(tour_names)):
            if tour_names[i].strip():
                tours_list.append({
                    'name':     tour_names[i].strip(),
                    'date':     tour_dates[i].strip()     if i < len(tour_dates)    else '',
                    'pickup':   tour_pickups[i].strip()   if i < len(tour_pickups)  else '',
                    'status':   tour_statuses[i].strip()  if i < len(tour_statuses) else 'INCLUDED',
                    'supplier': tour_suppliers[i].strip() if i < len(tour_suppliers)else '',
                    'cost':     float(tour_costs[i])      if i < len(tour_costs) and tour_costs[i] else 0,
                    'notes':    tour_notes[i].strip()     if i < len(tour_notes)    else '',
                })
        # Build passengers list from form
        _pax_n  = request.form.getlist('pax_name[]')
        _pax_p  = request.form.getlist('pax_passport[]')
        _pax_nt = request.form.getlist('pax_nationality[]')
        _pax_d  = request.form.getlist('pax_dob[]')
        passengers_list = [
            {'name': _pax_n[i].strip(),
             'passport':    _pax_p[i].strip()  if i < len(_pax_p)  else '',
             'nationality': _pax_nt[i].strip() if i < len(_pax_nt) else '',
             'dob':         _pax_d[i].strip()  if i < len(_pax_d)  else ''}
            for i in range(len(_pax_n)) if _pax_n[i].strip()
        ]

        extra_vals = dict(
            service_type      = service_type,
            hotel_supplier    = request.form.get('hotel_supplier','').strip(),
            hotel_name        = request.form.get('hotel_name','').strip(),
            hotel_room        = request.form.get('hotel_room','').strip(),
            hotel_meal        = request.form.get('hotel_meal','').strip(),
            hotel_checkin     = request.form.get('hotel_checkin','').strip(),
            hotel_checkout    = request.form.get('hotel_checkout','').strip(),
            hotel_nights      = int(request.form.get('hotel_nights',0) or 0),
            hotel_net         = float(request.form.get('hotel_net',0) or 0),
            transfer_supplier = request.form.get('transfer_supplier','').strip(),
            transfer_type     = request.form.get('transfer_type','').strip(),
            transfer_pickup   = request.form.get('transfer_pickup','').strip(),
            transfer_vehicle  = request.form.get('transfer_vehicle','').strip(),
            transfer_net      = float(request.form.get('transfer_net',0) or 0),
            tours_json        = json.dumps(tours_list),
            visa_supplier     = request.form.get('visa_supplier','').strip(),
            visa_type         = request.form.get('visa_type','').strip(),
            passport_number   = request.form.get('passport_number','').upper().strip(),
            visa_status       = request.form.get('visa_status','').strip(),
            insurance_supplier= request.form.get('insurance_supplier','').strip(),
            insurance_type    = request.form.get('insurance_type','').strip(),
            airline           = request.form.get('airline','').upper().strip(),
            pnr               = request.form.get('pnr','').upper().strip(),
            baggage           = request.form.get('baggage','').strip(),
            passengers_json   = json.dumps(passengers_list),
        )
        new_id = execute_db('''
            INSERT INTO sales
            (from_loc,to_loc,via,trip_type,buy_from,company,tickets,
             customer,sale_date,travel_date,return_date,return_supplier,
             outbound_delivery,return_delivery,outbound_status,return_status,
             net,sell,profit,status,remarks,
             created_by_user_id,created_by_username,
             service_type,hotel_supplier,hotel_name,hotel_room,hotel_meal,
             hotel_checkin,hotel_checkout,hotel_nights,hotel_net,
             transfer_supplier,transfer_type,transfer_pickup,transfer_vehicle,transfer_net,
             tours_json,visa_supplier,visa_type,passport_number,visa_status,
             insurance_supplier,insurance_type,airline,pnr,baggage,passengers_json,
             outbound_cost,return_cost)
            VALUES (
                %s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,%s
            ) RETURNING id
        ''', (
            (request.form.get('from_loc','').upper().strip() or '-'),
            (request.form.get('to_loc','').upper().strip() or '-'),
            request.form.get('via','').upper().strip(),
            request.form.get('trip_type',''),
            request.form.get('buy_from','').upper().strip(),
            request.form.get('company','').upper().strip(),
            int(request.form.get('tickets', 1)),
            request.form.get('customer','').upper().strip(),
            request.form.get('sale_date', str(date.today())),
            request.form.get('travel_date','').strip(),
            return_date, return_supplier,
            outbound_delivery, return_delivery,
            outbound_status, return_status,
            net, sell, sell - net, overall,
            request.form.get('remarks','').strip(),
            session.get('user_id'), session.get('username',''),
            extra_vals['service_type'],extra_vals['hotel_supplier'],extra_vals['hotel_name'],
            extra_vals['hotel_room'],extra_vals['hotel_meal'],extra_vals['hotel_checkin'],
            extra_vals['hotel_checkout'],extra_vals['hotel_nights'],extra_vals['hotel_net'],
            extra_vals['transfer_supplier'],extra_vals['transfer_type'],extra_vals['transfer_pickup'],
            extra_vals['transfer_vehicle'],extra_vals['transfer_net'],extra_vals['tours_json'],
            extra_vals['visa_supplier'],extra_vals['visa_type'],extra_vals['passport_number'],
            extra_vals['visa_status'],extra_vals['insurance_supplier'],extra_vals['insurance_type'],
            extra_vals['airline'],extra_vals['pnr'],extra_vals['baggage'],
            extra_vals['passengers_json'],
            float(request.form.get('outbound_cost',0) or 0),
            float(request.form.get('return_cost',0) or 0)
        ))
        # Assign transaction number
        if new_id:
            _txn = next_txn_number()
            try:
                execute_db('UPDATE sales SET txn_number=%s WHERE id=%s', (_txn, new_id))
            except Exception:
                pass
        try:
            log_action('CREATE','sales',new_id,
                       f"{request.form.get('customer','').upper()} | {extra_vals['service_type']} | Sell:{sell}")
        except Exception:
            pass
        # ── Shadow ledger entry ───────────────────────────────────────────────
        if new_id:
            _co  = request.form.get('company','').upper().strip()
            _svc = extra_vals.get('service_type','FLIGHT')
            _fl  = request.form.get('from_loc','') or ''
            _tl  = request.form.get('to_loc','')   or ''
            _cust= request.form.get('customer','')
            _dt  = request.form.get('sale_date', str(date.today()))
            post_ledger('SALE', new_id, 'sales', _co,
                        f"{_svc} {_fl}{'→'+_tl if _tl else ''} | {_cust}",
                        debit=sell, credit=0, entry_date=_dt)
            post_sale_items(new_id, _svc,
                            extra_vals.get('buy_from') or request.form.get('buy_from',''),
                            f"{_svc} {_fl}{'→'+_tl if _tl else ''}",
                            extra_vals.get('outbound_cost') or net,
                            sell,
                            request.form.get('tickets',1))
        # Sync PACKAGE transactions to the packages table
        if new_id and extra_vals.get('service_type', 'FLIGHT') == 'PACKAGE':
            try:
                sync_package_from_sale(new_id)
            except Exception:
                pass
        flash('Transaction saved successfully.', 'success')
        return redirect(url_for('sales_report'))  # POST→REDIRECT→GET
    return render_template('add.html', companies=companies, today=str(date.today()), form={})


def safe_sale(sale):
    """Normalize DB row with default values for all expected columns."""
    if not sale: return sale
    d = dict(sale)
    defaults = {
        'service_type':'FLIGHT','airline':'','pnr':'','baggage':'',
        'trip_type':'RETURN','buy_from':'','return_supplier':'',
        'outbound_cost':0,'return_cost':0,
        'hotel_supplier':'','hotel_name':'','hotel_room':'',
        'hotel_meal':'','hotel_checkin':'','hotel_checkout':'',
        'hotel_nights':0,'hotel_net':0,
        'transfer_supplier':'','transfer_type':'','transfer_pickup':'',
        'transfer_vehicle':'','transfer_net':0,
        'tours_json':'[]','passengers_json':'[]',
        'visa_supplier':'','visa_type':'','passport_number':'',
        'visa_status':'','insurance_supplier':'','insurance_type':'',
        'travel_date':'','return_date':'','via':'','remarks':'',
        'from_loc':'','to_loc':'','outbound_delivery':'',
        'return_delivery':'','outbound_status':'','return_status':'',
        # Phase 3 — ticket numbers and return PNR
        'ticket_number':'','return_ticket_number':'','return_pnr':'',
    }
    for k,v in defaults.items():
        if k not in d or d[k] is None: d[k] = v
    return d

@app.route('/edit/<int:sale_id>', methods=['GET', 'POST'])
@admin_required
def edit_sale(sale_id):
    # ── Fetch & normalise ─────────────────────────────────────────────────────
    sale_row = query_db('SELECT * FROM sales WHERE id=%s AND deleted=FALSE', [sale_id], one=True)
    if not sale_row:
        flash('Transaction not found.', 'danger')
        return redirect(url_for('sales_report'))
    sale = safe_sale(sale_row)

    try:
        companies = get_companies_list()
    except Exception:
        companies = []
    # ── POST — save changes ───────────────────────────────────────────────────
    if request.method == 'POST':
        errors = validate_sale_form(request.form)
        if errors:
            for e in errors: flash(e, 'danger')
            return render_template('add.html', sale=sale, companies=companies,
                                   edit=True, today=str(date.today()), form=request.form)

        net  = float(request.form.get('net',  0) or 0)
        sell = float(request.form.get('sell', 0) or 0)

        outbound_delivery = request.form.get('outbound_delivery', '').strip()
        return_delivery   = request.form.get('return_delivery',   '').strip()
        return_date       = request.form.get('return_date',       '').strip()
        return_supplier   = request.form.get('return_supplier',   '').upper().strip()
        outbound_status, return_status, overall = compute_ticket_status(
            outbound_delivery, return_delivery)

        service_type  = request.form.get('service_type', 'FLIGHT').upper().strip()

        # Tours
        tour_names     = request.form.getlist('tour_name[]')
        tour_dates     = request.form.getlist('tour_date[]')
        tour_pickups   = request.form.getlist('tour_pickup[]')
        tour_statuses  = request.form.getlist('tour_status[]')
        tour_suppliers = request.form.getlist('tour_supplier[]')
        tour_costs     = request.form.getlist('tour_cost[]')
        tour_notes     = request.form.getlist('tour_notes[]')
        tours_list = []
        for i in range(len(tour_names)):
            if tour_names[i].strip():
                tours_list.append({
                    'name':     tour_names[i].strip(),
                    'date':     tour_dates[i].strip()     if i < len(tour_dates)     else '',
                    'pickup':   tour_pickups[i].strip()   if i < len(tour_pickups)   else '',
                    'status':   tour_statuses[i].strip()  if i < len(tour_statuses)  else 'INCLUDED',
                    'supplier': tour_suppliers[i].strip() if i < len(tour_suppliers) else '',
                    'cost':     float(tour_costs[i]) if i < len(tour_costs) and tour_costs[i] else 0,
                    'notes':    tour_notes[i].strip()     if i < len(tour_notes)     else '',
                })

        # Passengers
        _pax_n  = request.form.getlist('pax_name[]')
        _pax_p  = request.form.getlist('pax_passport[]')
        _pax_nt = request.form.getlist('pax_nationality[]')
        _pax_d  = request.form.getlist('pax_dob[]')
        passengers_list = [
            {'name': _pax_n[i].strip(),
             'passport':    _pax_p[i].strip()  if i < len(_pax_p)  else '',
             'nationality': _pax_nt[i].strip() if i < len(_pax_nt) else '',
             'dob':         _pax_d[i].strip()  if i < len(_pax_d)  else ''}
            for i in range(len(_pax_n)) if _pax_n[i].strip()
        ]

        ev = dict(
            service_type=service_type,
            hotel_supplier=request.form.get('hotel_supplier','').strip(),
            hotel_name=request.form.get('hotel_name','').strip(),
            hotel_room=request.form.get('hotel_room','').strip(),
            hotel_meal=request.form.get('hotel_meal','').strip(),
            hotel_checkin=request.form.get('hotel_checkin','').strip(),
            hotel_checkout=request.form.get('hotel_checkout','').strip(),
            hotel_nights=int(request.form.get('hotel_nights', 0) or 0),
            hotel_net=float(request.form.get('hotel_net', 0) or 0),
            transfer_supplier=request.form.get('transfer_supplier','').strip(),
            transfer_type=request.form.get('transfer_type','').strip(),
            transfer_pickup=request.form.get('transfer_pickup','').strip(),
            transfer_vehicle=request.form.get('transfer_vehicle','').strip(),
            transfer_net=float(request.form.get('transfer_net', 0) or 0),
            tours_json=json.dumps(tours_list),
            visa_supplier=request.form.get('visa_supplier','').strip(),
            visa_type=request.form.get('visa_type','').strip(),
            passport_number=request.form.get('passport_number','').upper().strip(),
            visa_status=request.form.get('visa_status','').strip(),
            insurance_supplier=request.form.get('insurance_supplier','').strip(),
            insurance_type=request.form.get('insurance_type','').strip(),
            airline=request.form.get('airline','').upper().strip(),
            pnr=request.form.get('pnr','').upper().strip(),
            baggage=request.form.get('baggage','').strip(),
            passengers_json=json.dumps(passengers_list),
            outbound_cost=float(request.form.get('outbound_cost', 0) or 0),
            return_cost=float(request.form.get('return_cost', 0) or 0),
            # Phase 3 — ticket numbers
            ticket_number=request.form.get('ticket_number','').strip(),
            return_ticket_number=request.form.get('return_ticket_number','').strip(),
            # Return PNR — separate PNR for return leg
            return_pnr=request.form.get('return_pnr','').upper().strip(),
        )

        try:
            # ACC-01 / ACC-04 — Snapshot old values BEFORE update for ledger reversal
            _old = query_db(
                'SELECT sell, company, sale_date FROM sales WHERE id=%s AND deleted=FALSE',
                [sale_id], one=True
            )
            _old_sell    = float(_old['sell']    if _old else 0)
            _old_company = str(_old['company']   if _old else '').upper().strip()
            _old_date    = str(_old['sale_date'] if _old else str(date.today()))

            _ctype_edit = request.form.get('customer_type', 'COMPANY').upper().strip() or 'COMPANY'
            execute_db("""
                UPDATE sales SET
                    from_loc=%s, to_loc=%s, via=%s, trip_type=%s, buy_from=%s,
                    company=%s, tickets=%s, customer=%s, sale_date=%s, travel_date=%s,
                    return_date=%s, return_supplier=%s,
                    outbound_delivery=%s, return_delivery=%s,
                    outbound_status=%s, return_status=%s,
                    net=%s, sell=%s, profit=%s, status=%s, remarks=%s,
                    service_type=%s, hotel_supplier=%s, hotel_name=%s, hotel_room=%s, hotel_meal=%s,
                    hotel_checkin=%s, hotel_checkout=%s, hotel_nights=%s, hotel_net=%s,
                    transfer_supplier=%s, transfer_type=%s, transfer_pickup=%s, transfer_vehicle=%s, transfer_net=%s,
                    tours_json=%s, visa_supplier=%s, visa_type=%s, passport_number=%s, visa_status=%s,
                    insurance_supplier=%s, insurance_type=%s, airline=%s, pnr=%s, baggage=%s,
                    passengers_json=%s, outbound_cost=%s, return_cost=%s,
                    ticket_number=%s, return_ticket_number=%s, return_pnr=%s,
                    customer_type=%s, direct_phone=%s, direct_nationality=%s, direct_email=%s
                WHERE id=%s
            """, (
                (request.form.get('from_loc', '').upper().strip() or '-'),
                (request.form.get('to_loc',   '').upper().strip() or '-'),
                request.form.get('via', '').upper().strip(),
                request.form.get('trip_type', ''),
                request.form.get('buy_from', '').upper().strip(),
                request.form.get('company', '').upper().strip(),
                int(request.form.get('tickets', 1) or 1),
                request.form.get('customer', '').upper().strip(),
                request.form.get('sale_date', str(date.today())),
                request.form.get('travel_date', '').strip(),
                return_date, return_supplier,
                outbound_delivery, return_delivery,
                outbound_status, return_status,
                net, sell, round(sell - net, 3), overall,
                request.form.get('remarks', '').strip(),
                ev['service_type'], ev['hotel_supplier'], ev['hotel_name'], ev['hotel_room'], ev['hotel_meal'],
                ev['hotel_checkin'], ev['hotel_checkout'], ev['hotel_nights'], ev['hotel_net'],
                ev['transfer_supplier'], ev['transfer_type'], ev['transfer_pickup'], ev['transfer_vehicle'], ev['transfer_net'],
                ev['tours_json'], ev['visa_supplier'], ev['visa_type'], ev['passport_number'], ev['visa_status'],
                ev['insurance_supplier'], ev['insurance_type'], ev['airline'], ev['pnr'], ev['baggage'],
                ev['passengers_json'], ev['outbound_cost'], ev['return_cost'],
                ev['ticket_number'], ev['return_ticket_number'], ev['return_pnr'],
                _ctype_edit,
                request.form.get('direct_phone', '').strip(),
                request.form.get('direct_nationality', '').strip(),
                request.form.get('direct_email', '').strip(),
                sale_id,
            ))

            # ACC-01 / ACC-04 — Reverse original ledger entry, post corrective entry
            _orig_entry = query_db(
                """SELECT id FROM accounting_entries
                   WHERE ref_type='SALE' AND ref_id=%s AND ref_table='sales'
                     AND is_reversed=FALSE
                   ORDER BY id DESC LIMIT 1""",
                [sale_id], one=True
            )
            _new_company  = request.form.get('company', _old_company).upper().strip()
            _new_sale_date= request.form.get('sale_date', _old_date)
            _new_customer = request.form.get('customer', '').upper().strip()
            _svc_label    = ev['service_type']
            _fl           = request.form.get('from_loc', '') or ''
            _tl           = request.form.get('to_loc',   '') or ''
            _ledger_ent_edit = _new_customer if _ctype_edit == 'DIRECT' else _new_company
            if _orig_entry:
                execute_db(
                    "UPDATE accounting_entries SET is_reversed=TRUE WHERE id=%s",
                    [_orig_entry['id']]
                )
                post_ledger(
                    'SALE_REVERSAL', sale_id, 'sales', _old_company or _new_customer,
                    f"Reversal of SALE-{sale_id:04d} (edit by {session.get('username','')})",
                    debit=0, credit=_old_sell, entry_date=_old_date
                )
            post_ledger(
                'SALE', sale_id, 'sales', _ledger_ent_edit,
                f"{_svc_label} {_fl}{'→'+_tl if _tl else ''} | {_new_customer} (corrected)",
                debit=sell, credit=0, entry_date=_new_sale_date
            )

            log_action('UPDATE', 'sales', sale_id,
                       f"{_new_customer} | {_svc_label} | OldSell:{_old_sell} NewSell:{sell}")
            # Sync PACKAGE transactions to the packages table
            if service_type == 'PACKAGE':
                try:
                    sync_package_from_sale(sale_id)
                except Exception:
                    pass
            flash('Transaction updated successfully.', 'success')
            return redirect(url_for('sales_report'))

        except Exception as _ex:
            logger.error(f"edit_sale UPDATE error: {_ex}")
            try: get_db().rollback()
            except: pass
            flash('Error saving changes. Please try again.', 'danger')

    # ── GET (or POST error) — render edit form ────────────────────────────────
    return render_template('add.html', sale=sale, companies=companies,
                           edit=True, today=str(date.today()), form={})


@app.route('/delete/<int:sale_id>', methods=['POST'])
@admin_required
def delete_sale(sale_id):
    sale = query_db('SELECT customer, company FROM sales WHERE id=%s', [sale_id], one=True)
    # Soft delete
    execute_db('UPDATE sales SET deleted=TRUE WHERE id=%s', [sale_id])
    log_action('DELETE', 'sales', sale_id,
               f"{sale['customer'] if sale else ''} | {sale['company'] if sale else ''}")
    flash('Sale deleted.', 'success')
    return redirect(url_for('sales_report'))

# ── Sales Report (paginated) ──────────────────────────────────────────────────
@app.route('/report')
@login_required
def sales_report():
    company   = request.args.get('company', '')
    status    = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    agent_id  = request.args.get('agent_id', '')
    svc_type  = request.args.get('svc_type', '')
    customer  = request.args.get('customer', '').strip()
    page      = max(1, int(request.args.get('page', 1)))

    base_q  = '''
        SELECT s.*,
               COALESCE(s.created_by_username, '—') AS agent_name,
               COALESCE(u.full_name, s.created_by_username, '—') AS agent_display
        FROM sales s
        LEFT JOIN users u ON s.created_by_user_id = u.id
        WHERE s.deleted=FALSE AND s.is_archived=FALSE
    '''
    count_q = 'SELECT COALESCE(SUM(sell),0) as sell, COALESCE(SUM(net),0) as net, COALESCE(SUM(profit),0) as profit, COUNT(*) as cnt FROM sales WHERE deleted=FALSE AND is_archived=FALSE'
    params  = []
    cparams = []

    if company:
        base_q  += ' AND s.company=%s'; count_q += ' AND company=%s'
        params.append(company); cparams.append(company)
    if status:
        base_q  += ' AND s.status=%s';  count_q += ' AND status=%s'
        params.append(status); cparams.append(status)
    if date_from:
        base_q  += ' AND s.sale_date>=%s'; count_q += ' AND sale_date>=%s'
        params.append(date_from); cparams.append(date_from)
    if date_to:
        base_q  += ' AND s.sale_date<=%s'; count_q += ' AND sale_date<=%s'
        params.append(date_to); cparams.append(date_to)
    if agent_id:
        base_q  += ' AND s.created_by_user_id=%s'; count_q += ' AND created_by_user_id=%s'
        params.append(agent_id); cparams.append(agent_id)
    if svc_type:
        base_q  += ' AND UPPER(COALESCE(s.service_type,\'FLIGHT\'))=%s'
        count_q += ' AND UPPER(COALESCE(service_type,\'FLIGHT\'))=%s'
        params.append(svc_type.upper()); cparams.append(svc_type.upper())
    if customer:
        base_q  += ' AND UPPER(s.customer) LIKE %s'
        count_q += ' AND UPPER(customer) LIKE %s'
        _c = '%' + customer.upper() + '%'
        params.append(_c); cparams.append(_c)

    base_q += ' ORDER BY s.sale_date DESC, s.id DESC'

    agg = query_db(count_q, cparams, one=True)
    totals = {'sell': agg['sell'], 'net': agg['net'],
              'profit': agg['profit'], 'count': agg['cnt']}

    sales, total_rows, total_pages = paginate(base_q, params, page)

    companies = get_companies_list()
    agents = []
    if session.get('user_role') == 'admin':
        # Agent filter dropdown — one entry per active user, regardless of
        # whether they have any transactions yet (LEFT JOIN, not JOIN).
        # DISTINCT + grouping by id guards against any row fan-out from the
        # join; NULLIF/TRIM strips blank full_name values so the username
        # fallback kicks in instead of showing an empty name.
        agent_rows = query_db("""
            SELECT DISTINCT u.id, u.username,
                   NULLIF(TRIM(COALESCE(u.full_name, '')), '') AS full_name,
                   LOWER(COALESCE(NULLIF(TRIM(u.full_name), ''), u.username)) AS sort_key
            FROM users u
            LEFT JOIN sales s ON s.created_by_user_id = u.id
            WHERE COALESCE(u.is_active, TRUE) = TRUE
              AND TRIM(COALESCE(u.username, '')) != ''
            ORDER BY sort_key ASC
        """) or []
        agents = []
        seen_ids = set()
        for a in agent_rows:
            if a['id'] in seen_ids:
                continue
            seen_ids.add(a['id'])
            name = (a['full_name'] or '').strip()
            agents.append({
                'id': a['id'],
                'username': a['username'],
                'display': f"{name} (@{a['username']})" if name else f"@{a['username']}",
            })

    return render_template('report.html',
        sales=sales, totals=totals, companies=companies, svc_filter=svc_type, agents=agents,
        filters={'company':company,'status':status,'date_from':date_from,
                 'date_to':date_to, 'agent_id':agent_id, 'svc_type':svc_type,
                 'service_type':svc_type, 'customer':customer},
        page=page, total_pages=total_pages, total_rows=total_rows
    )

# ── Statement ─────────────────────────────────────────────────────────────────
@app.route('/statement')
@finance_required
def statement():
    company   = request.args.get('company','').strip()
    date_from = request.args.get('date_from','').strip()
    date_to   = request.args.get('date_to','').strip()
    svc_type  = request.args.get('svc_type','').strip().upper()

    try:    companies = get_companies_list()
    except: companies = []

    ledger_entries = []
    summary        = {}

    if company:
        # ── H-01 FIX: single UNION ALL query — one round-trip instead of 8 ─────
        # All ledger sources (sales/purchases/payments/adjustments) combined in SQL.
        # Sort order: date ASC, then type priority (sale=0, purchase=1, ...), then id.
        # Date filter applies to the natural date column of each source type.

        _co = company.upper().strip()

        # Build optional date/service-type clauses per source
        _s_date_sql, _s_params = '', []          # sales date filter
        _p_date_sql, _p_params = '', []          # payments date filter
        _sp_date_sql, _sp_params = '', []        # supplier_payments date filter
        _adj_date_sql, _adj_params = '', []      # adjustments date filter
        _svc_sql, _svc_param = '', []            # service_type filter (sales only)

        if date_from:
            _s_date_sql  += ' AND sale_date>=%s';   _s_params.append(date_from)
            _p_date_sql  += ' AND pay_date>=%s';    _p_params.append(date_from)
            _sp_date_sql += ' AND pay_date>=%s';    _sp_params.append(date_from)
            _adj_date_sql+= ' AND adj_date>=%s';   _adj_params.append(date_from)
        if date_to:
            _s_date_sql  += ' AND sale_date<=%s';   _s_params.append(date_to)
            _p_date_sql  += ' AND pay_date<=%s';    _p_params.append(date_to)
            _sp_date_sql += ' AND pay_date<=%s';    _sp_params.append(date_to)
            _adj_date_sql+= ' AND adj_date<=%s';   _adj_params.append(date_to)
        if svc_type:
            _svc_sql = " AND UPPER(TRIM(COALESCE(service_type,'FLIGHT')))=%s"
            _svc_param = [svc_type]

        try:
            raw_rows = query_db(f"""
            SELECT * FROM (
                -- 1. Sales to this company (DEBIT)
                SELECT
                    sale_date            AS entry_date,
                    'TXN-' || LPAD(id::TEXT,4,'0') AS ref,
                    COALESCE(service_type,'FLIGHT') || CASE
                        WHEN from_loc<>'' AND to_loc<>'' THEN ' '||from_loc||chr(8594)||to_loc ELSE ''
                    END || CASE WHEN pnr<>'' THEN ' ['||pnr||']' ELSE '' END  AS description,
                    customer             AS passenger,
                    COALESCE(service_type,'FLIGHT') AS svc,
                    CAST(sell AS FLOAT)  AS debit,
                    0.0                  AS credit,
                    'sale'               AS entry_type,
                    COALESCE(travel_date,'')             AS travel_date,
                    id,
                    COALESCE(ticket_number,'')           AS ticket_number,
                    COALESCE(return_ticket_number,'')    AS return_ticket_number
                FROM sales
                WHERE deleted=FALSE AND is_archived=FALSE
                  AND UPPER(TRIM(company))=%s
                  {_s_date_sql} {_svc_sql}

                UNION ALL

                -- 2. Outbound purchases from this supplier (CREDIT)
                SELECT
                    sale_date,
                    'PUR-' || LPAD(id::TEXT,4,'0'),
                    'Purchase Flight ' || from_loc || chr(8594) || to_loc,
                    customer, 'FLIGHT',
                    0.0,
                    CAST(CASE WHEN outbound_cost>0 THEN outbound_cost ELSE COALESCE(net,0) END AS FLOAT),
                    'purchase',
                    COALESCE(travel_date,''), id, '', ''
                FROM sales
                WHERE deleted=FALSE AND is_archived=FALSE
                  AND UPPER(TRIM(buy_from))=%s
                  AND (outbound_cost>0 OR net>0)
                  {_s_date_sql}

                UNION ALL

                -- 3. Return leg purchases from this supplier (CREDIT)
                SELECT
                    sale_date,
                    'RTN-' || LPAD(id::TEXT,4,'0'),
                    'Purchase Return ' || to_loc || chr(8594) || from_loc
                        || CASE WHEN return_pnr<>'' THEN ' ['||return_pnr||']' ELSE '' END,
                    customer, 'FLIGHT',
                    0.0,
                    CAST(return_cost AS FLOAT),
                    'purchase',
                    COALESCE(return_date, travel_date,''), id,
                    '', COALESCE(return_ticket_number,'')
                FROM sales
                WHERE deleted=FALSE AND is_archived=FALSE
                  AND UPPER(TRIM(return_supplier))=%s
                  AND return_cost > 0
                  {_s_date_sql}

                UNION ALL

                -- 4. Hotel purchases from this supplier (CREDIT)
                SELECT
                    sale_date,
                    'HOT-' || LPAD(id::TEXT,4,'0'),
                    'Purchase Hotel' || CASE WHEN hotel_name<>'' THEN ' - '||hotel_name ELSE '' END,
                    customer, 'HOTEL',
                    0.0, CAST(hotel_net AS FLOAT),
                    'purchase',
                    COALESCE(travel_date,''), id, '', ''
                FROM sales
                WHERE deleted=FALSE AND is_archived=FALSE
                  AND UPPER(TRIM(hotel_supplier))=%s AND hotel_net>0
                  {_s_date_sql}

                UNION ALL

                -- 5. Transfer purchases (CREDIT)
                SELECT
                    sale_date,
                    'TRN-' || LPAD(id::TEXT,4,'0'),
                    'Purchase Transfer',
                    customer, 'TRANSFER',
                    0.0, CAST(transfer_net AS FLOAT),
                    'purchase',
                    COALESCE(travel_date,''), id, '', ''
                FROM sales
                WHERE deleted=FALSE AND is_archived=FALSE
                  AND UPPER(TRIM(transfer_supplier))=%s AND transfer_net>0
                  {_s_date_sql}

                UNION ALL

                -- 6. Visa purchases (CREDIT)
                SELECT
                    sale_date,
                    'VIS-' || LPAD(id::TEXT,4,'0'),
                    'Purchase Visa' || CASE WHEN visa_type<>'' THEN ' - '||visa_type ELSE '' END,
                    customer, 'VISA',
                    0.0, CAST(net AS FLOAT),
                    'purchase',
                    COALESCE(travel_date,''), id, '', ''
                FROM sales
                WHERE deleted=FALSE AND is_archived=FALSE
                  AND UPPER(TRIM(visa_supplier))=%s AND net>0
                  {_s_date_sql}

                UNION ALL

                -- 7. Insurance purchases (CREDIT)
                SELECT
                    sale_date,
                    'INS-' || LPAD(id::TEXT,4,'0'),
                    'Purchase Insurance',
                    customer, 'INSURANCE',
                    0.0, CAST(net AS FLOAT),
                    'purchase',
                    COALESCE(travel_date,''), id, '', ''
                FROM sales
                WHERE deleted=FALSE AND is_archived=FALSE
                  AND UPPER(TRIM(insurance_supplier))=%s AND net>0
                  {_s_date_sql}

                UNION ALL

                -- 8. Payments received from company (CREDIT)
                SELECT
                    pay_date,
                    'PAY-' || LPAD(id::TEXT,4,'0'),
                    'Payment received' || CASE WHEN COALESCE(notes,'')<>'' THEN ' - '||notes ELSE '' END,
                    '', '',
                    0.0, CAST(amount AS FLOAT),
                    'payment_in',
                    '', id, '', ''
                FROM payments
                WHERE deleted=FALSE AND is_archived=FALSE
                  AND COALESCE(voided,FALSE)=FALSE
                  AND UPPER(TRIM(company))=%s
                  {_p_date_sql}

                UNION ALL

                -- 9. Payments made to supplier (DEBIT)
                SELECT
                    pay_date,
                    'SPY-' || LPAD(id::TEXT,4,'0'),
                    'Payment made' || CASE WHEN COALESCE(notes,'')<>'' THEN ' - '||notes ELSE '' END,
                    '', COALESCE(service_type,''),
                    CAST(amount AS FLOAT), 0.0,
                    'payment_out',
                    '', id, '', ''
                FROM supplier_payments
                WHERE deleted=FALSE AND is_archived=FALSE
                  AND UPPER(TRIM(supplier))=%s
                  {_sp_date_sql}

                UNION ALL

                -- 10. Balance adjustments (DEBIT or CREDIT)
                SELECT
                    adj_date,
                    'ADJ-' || LPAD(id::TEXT,4,'0'),
                    COALESCE(reason,'Balance Adjustment'),
                    COALESCE(notes,''), '',
                    CASE WHEN entry_type='DEBIT'  THEN CAST(amount AS FLOAT) ELSE 0.0 END,
                    CASE WHEN entry_type='CREDIT' THEN CAST(amount AS FLOAT) ELSE 0.0 END,
                    'adjustment',
                    '', id, '', ''
                FROM balance_adjustments
                WHERE deleted=FALSE AND account_type='COMPANY'
                  AND UPPER(TRIM(account))=%s
                  {_adj_date_sql}

            ) AS _ledger
                ORDER BY
                    entry_date ASC,
                    CASE entry_type
                        WHEN 'sale'        THEN 0
                        WHEN 'purchase'    THEN 1
                        WHEN 'payment_in'  THEN 2
                        WHEN 'payment_out' THEN 3
                        WHEN 'adjustment'  THEN 4
                        ELSE 9
                    END ASC,
                    ref ASC
            """, (
                [_co] + _s_params + _svc_param    # sales to company
                + [_co] + _s_params               # outbound purchases
                + [_co] + _s_params               # return purchases
                + [_co] + _s_params               # hotel
                + [_co] + _s_params               # transfer
                + [_co] + _s_params               # visa
                + [_co] + _s_params               # insurance
                + [_co] + _p_params               # payments received
                + [_co] + _sp_params              # supplier payments
                + [_co] + _adj_params             # adjustments
            )) or []
        except Exception as _se:
            logger.error(f"statement UNION query error: {_se}")
            try: get_db().rollback()
            except: pass
            raw_rows = []

        # Build ledger entries and compute running balance in one pass
        running = 0.0
        _pay_n  = 0
        for row in raw_rows:
            dr = float(row['debit']  or 0)
            cr = float(row['credit'] or 0)
            running += dr - cr
            entry_type = row['entry_type']
            ref = row['ref']
            if entry_type == 'payment_in':
                _pay_n += 1
                ref = f"PAY-{_pay_n:02d}"
            ledger_entries.append({
                'date':                 row['entry_date'],
                'ref':                  ref,
                'description':          row['description'] or '',
                'passenger':            row['passenger'] or '',
                'svc':                  row['svc'] or '',
                'debit':                dr,
                'credit':               cr,
                'type':                 entry_type,
                'travel_date':          row['travel_date'] or '',
                'id':                   row['id'],
                'ticket_number':        row['ticket_number'] or '',
                'return_ticket_number': row['return_ticket_number'] or '',
                'balance':              round(running, 3),
            })

        net_bal  = round(running, 3)
        sold     = sum(e['debit']  for e in ledger_entries if e['type']=='sale')
        bought   = sum(e['credit'] for e in ledger_entries if e['type']=='purchase')
        received = sum(e['credit'] for e in ledger_entries if e['type']=='payment_in')
        paid_out = sum(e['debit']  for e in ledger_entries if e['type']=='payment_out')
        summary  = {
            'total_sold':     sold,
            'total_bought':   bought,
            'total_received': received,
            'total_paid_out': paid_out,
            'net_receivable': round(sold   - received, 3),
            'net_payable':    round(bought - paid_out, 3),
            'net_balance':    net_bal,
            'total_debit':    sum(e['debit']  for e in ledger_entries),
            'total_credit':   sum(e['credit'] for e in ledger_entries),
            'entry_count':    len(ledger_entries),
            'status': 'RECEIVABLE' if net_bal >  0.005
                 else ('PAYABLE'   if net_bal < -0.005 else 'SETTLED'),
        }

    return render_template('statement.html',
        companies=companies, ledger=ledger_entries, summary=summary,
        company=company, svc_type=svc_type,
        filters={'date_from':date_from,'date_to':date_to,'svc_type':svc_type},
        today=date.today().strftime('%d %B %Y'),
        today_iso=str(date.today()),
    )


@app.route('/payments', methods=['GET', 'POST'])
@login_required
def payments():
    companies  = get_companies_list()
    today_str  = str(date.today())

    if request.method == 'POST':
        if session.get('user_role') not in ('admin', 'manager'):
            flash('Admin access required to record payments.', 'danger')
            return redirect(url_for('payments'))
        company_val    = request.form.get('company','').upper().strip()
        amount_val     = request.form.get('amount','').strip()
        pay_date       = request.form.get('pay_date', today_str)
        method_val     = request.form.get('method','').strip()
        reference_val  = request.form.get('reference_number','').strip().upper()
        notes_val      = request.form.get('notes','').strip()
        if not company_val:
            flash('Company is required.', 'danger')
            return redirect(url_for('payments'))
        try:
            amount = float(amount_val)
            if amount <= 0: raise ValueError
        except ValueError:
            flash('Amount must be a positive number.', 'danger')
            return redirect(url_for('payments'))
        # EMP-02 FIX: large payment alert — amounts above JOD 5,000 require admin role
        _LARGE_PMT_THRESHOLD = 5000.0
        if amount >= _LARGE_PMT_THRESHOLD and session.get('user_role') != 'admin':
            flash(
                f'Payments of JOD {amount:,.3f} or more require admin approval. '
                f'Please ask an administrator to record this payment.',
                'danger'
            )
            return redirect(url_for('payments'))

        # Duplicate payment guard
        _force = request.form.get('force_duplicate', '').strip() == '1'
        _dup = query_db(
            """SELECT id FROM payments
               WHERE UPPER(TRIM(company))=UPPER(%s)
                 AND ABS(amount - %s) < 0.001
                 AND pay_date = %s
                 AND deleted = FALSE AND voided = FALSE
               LIMIT 1""",
            [company_val, amount, pay_date], one=True
        )
        if _dup and not _force:
            flash(
                f'⚠️ Duplicate warning: A payment of JOD {amount:.3f} for {company_val} on {pay_date} '
                f'already exists (ID #{_dup["id"]}). Submit again to confirm.',
                'warning'
            )
            return redirect(url_for('payments',
                dup_warn='1', dup_company=company_val, dup_amount=amount_val,
                dup_date=pay_date, dup_method=method_val, dup_ref=reference_val,
                dup_notes=notes_val))
        new_id = execute_db(
            '''INSERT INTO payments
               (company, amount, pay_date, payment_method, reference_number, notes, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id''',
            (company_val, amount, pay_date, method_val, reference_val, notes_val,
             session.get('user_id'))
        )
        log_action('CREATE', 'payments', new_id,
                   f"{company_val} | JOD {amount} | {method_val or 'N/A'} | Ref:{reference_val or '-'}"
                   + (" [FORCE-DUP]" if _force else ""))
        if new_id:
            ldesc = f"Payment received — {method_val or ''}"
            if reference_val: ldesc += f" [Ref:{reference_val}]"
            if notes_val:     ldesc += f" | {notes_val}"
            post_ledger('PAYMENT', new_id, 'payments', company_val,
                        ldesc, debit=0, credit=amount, entry_date=pay_date)
        flash(f'Payment of JOD {amount:.3f} recorded for {company_val}.', 'success')
        return redirect(url_for('payments'))

    # ── GET: filters ─────────────────────────────────────────────
    f_company   = request.args.get('company','').upper().strip()
    f_method    = request.args.get('method','').strip()
    f_date_from = request.args.get('date_from','').strip()
    f_date_to   = request.args.get('date_to','').strip()
    f_ref       = request.args.get('reference','').strip().upper()
    f_amt_min   = request.args.get('amt_min','').strip()
    f_amt_max   = request.args.get('amt_max','').strip()
    f_created   = request.args.get('created_by','').strip()
    page        = max(1, int(request.args.get('page', 1)))

    where = ['p.deleted=FALSE', 'p.is_archived=FALSE', 'COALESCE(p.voided,FALSE)=FALSE']
    params = []
    if f_company:
        where.append('UPPER(p.company) LIKE %s'); params.append('%'+f_company+'%')
    if f_method:
        where.append('p.payment_method=%s'); params.append(f_method)
    if f_date_from:
        where.append('p.pay_date >= %s'); params.append(f_date_from)
    if f_date_to:
        where.append('p.pay_date <= %s'); params.append(f_date_to)
    if f_ref:
        where.append('UPPER(p.reference_number) LIKE %s'); params.append('%'+f_ref+'%')
    if f_amt_min:
        try: where.append('p.amount >= %s'); params.append(float(f_amt_min))
        except ValueError: pass
    if f_amt_max:
        try: where.append('p.amount <= %s'); params.append(float(f_amt_max))
        except ValueError: pass
    if f_created:
        where.append('CAST(p.created_by AS TEXT)=%s'); params.append(f_created)

    w_sql = ' AND '.join(where)
    base_q = f"""
        SELECT p.*,
               u.username  AS created_by_name,
               u.full_name AS created_by_full
        FROM payments p
        LEFT JOIN users u ON u.id = p.created_by
        WHERE {w_sql}
        ORDER BY p.pay_date DESC, p.id DESC
    """
    all_payments, total_rows, total_pages = paginate(base_q, params, page)

    # KPIs
    today_date = date.today()
    month_start = today_date.replace(day=1)

    kpi = query_db("""
        SELECT
          COALESCE(SUM(amount) FILTER (WHERE pay_date = %s),  0) AS today_total,
          COALESCE(SUM(amount) FILTER (WHERE pay_date >= %s), 0) AS month_total,
          COALESCE(SUM(amount) FILTER (WHERE payment_method='Cash'),           0) AS cash_total,
          COALESCE(SUM(amount) FILTER (WHERE payment_method='Bank Transfer'),  0) AS bank_total,
          COALESCE(SUM(amount) FILTER (WHERE payment_method='CliQ'),           0) AS cliq_total,
          COALESCE(SUM(amount),                                                0) AS grand_total,
          COUNT(*) AS count_all
        FROM payments
        WHERE deleted=FALSE AND is_archived=FALSE AND COALESCE(voided,FALSE)=FALSE
    """, [str(today_date), str(month_start)], one=True)

    # Outstanding receivables = total sell - total payments
    outstanding = query_db("""
        SELECT
          COALESCE((SELECT SUM(sell) FROM sales WHERE deleted=FALSE AND is_archived=FALSE),0)
          - COALESCE((SELECT SUM(amount) FROM payments WHERE deleted=FALSE AND is_archived=FALSE AND COALESCE(voided,FALSE)=FALSE),0)
          AS outstanding
    """, one=True)

    # Staff list for filter
    staff_list = query_db(
        "SELECT id, COALESCE(full_name, username) AS name FROM users ORDER BY full_name",
        []
    ) or []

    # Duplicate warn pre-fill from redirect
    dup_warn    = request.args.get('dup_warn') == '1'
    dup_company = request.args.get('dup_company','')
    dup_amount  = request.args.get('dup_amount','')
    dup_date    = request.args.get('dup_date', today_str)
    dup_method  = request.args.get('dup_method','')
    dup_ref     = request.args.get('dup_ref','')
    dup_notes   = request.args.get('dup_notes','')

    return render_template('payments.html',
        payments=all_payments, companies=companies,
        total_paid=kpi['grand_total'], today=today_str,
        page=page, total_pages=total_pages, total_rows=total_rows,
        kpi=kpi, outstanding=float(outstanding['outstanding'] or 0),
        staff_list=staff_list,
        filters=dict(company=f_company, method=f_method, date_from=f_date_from,
                     date_to=f_date_to, reference=f_ref,
                     amt_min=f_amt_min, amt_max=f_amt_max, created_by=f_created),
        dup_warn=dup_warn, dup_company=dup_company, dup_amount=dup_amount,
        dup_date=dup_date, dup_method=dup_method, dup_ref=dup_ref, dup_notes=dup_notes,
    )

@app.route('/payments/edit/<int:pay_id>', methods=['GET', 'POST'])
@admin_required
def edit_payment(pay_id):
    payment = query_db('SELECT * FROM payments WHERE id=%s AND deleted=FALSE', [pay_id], one=True)
    if not payment:
        flash('Payment not found.', 'danger')
        return redirect(url_for('payments'))
    companies = get_companies_list()
    if request.method == 'POST':
        try:
            amount = float(request.form.get('amount', 0))
            if amount <= 0: raise ValueError
        except ValueError:
            flash('Amount must be a positive number.', 'danger')
            return render_template('edit_payment.html', payment=payment,
                                   companies=companies, today=str(date.today()))
        new_company  = request.form.get('company','').upper().strip()
        new_pay_date = request.form.get('pay_date', str(date.today()))
        new_method   = request.form.get('method','').strip()
        new_ref      = request.form.get('reference_number','').strip().upper()
        new_notes    = request.form.get('notes','').strip()
        old_amount   = float(payment['amount'] or 0)
        old_company  = str(payment['company'] or '').upper().strip()

        # ACC-01 / ACC-04: reverse original ledger entry, post corrective entry
        # Step 1 — find original accounting_entries row for this payment
        _orig_entry = query_db(
            """SELECT id FROM accounting_entries
               WHERE ref_type='PAYMENT' AND ref_id=%s AND ref_table='payments'
                 AND is_reversed=FALSE
               ORDER BY id DESC LIMIT 1""",
            [pay_id], one=True
        )
        if _orig_entry:
            # Mark original as reversed
            execute_db(
                "UPDATE accounting_entries SET is_reversed=TRUE WHERE id=%s",
                [_orig_entry['id']]
            )
            # Post reversal (debit what was credited, credit what was debited)
            post_ledger('PAYMENT_REVERSAL', pay_id, 'payments', old_company,
                        f"Reversal of payment PAY-{pay_id:04d} (edit correction)",
                        debit=old_amount, credit=0, entry_date=new_pay_date)

        # Step 2 — update the payments row
        execute_db('''
            UPDATE payments SET company=%s, amount=%s, pay_date=%s,
                   payment_method=%s, reference_number=%s, notes=%s WHERE id=%s
        ''', (new_company, amount, new_pay_date, new_method, new_ref, new_notes, pay_id))

        # Step 3 — post corrective ledger entry with new values
        post_ledger('PAYMENT', pay_id, 'payments', new_company,
                    f"Payment received (corrected) — {new_notes or ''}",
                    debit=0, credit=amount, entry_date=new_pay_date)

        log_action('UPDATE', 'payments', pay_id,
                   f"Corrected: {old_company} JOD {old_amount} → {new_company} JOD {amount}")
        flash('Payment updated successfully.', 'success')
        return redirect(url_for('payments'))
    return render_template('edit_payment.html',
        payment=payment, companies=companies, today=str(date.today()))

@app.route('/payments/delete/<int:pay_id>', methods=['POST'])
@admin_required
def delete_payment_page(pay_id):
    p = query_db('SELECT company, amount, pay_date FROM payments WHERE id=%s', [pay_id], one=True)
    if p:
        # Reverse the original ledger entry before soft-deleting
        _orig = query_db(
            """SELECT id FROM accounting_entries
               WHERE ref_type='PAYMENT' AND ref_id=%s AND ref_table='payments'
                 AND is_reversed=FALSE
               ORDER BY id DESC LIMIT 1""",
            [pay_id], one=True
        )
        if _orig:
            execute_db("UPDATE accounting_entries SET is_reversed=TRUE WHERE id=%s", [_orig['id']])
        post_ledger('PAYMENT_REVERSAL', pay_id, 'payments', p['company'],
                    f"Reversal of PAY-{pay_id:04d} (deleted)",
                    debit=p['amount'], credit=0, entry_date=p['pay_date'])
    execute_db('UPDATE payments SET deleted=TRUE WHERE id=%s', [pay_id])
    log_action('DELETE', 'payments', pay_id,
               f"{p['company'] if p else ''} | Amount:{p['amount'] if p else ''}")
    flash('Payment deleted.', 'success')
    return redirect(url_for('payments'))

@app.route('/payments/void/<int:pay_id>', methods=['POST'])
@admin_required
def void_payment(pay_id):
    """Soft-void a payment: marks voided=TRUE, reverses ledger entry, never hard-deletes."""
    p = query_db('SELECT company, amount, pay_date FROM payments WHERE id=%s AND deleted=FALSE', [pay_id], one=True)
    if not p:
        flash('Payment not found.', 'danger')
        return redirect(url_for('payments'))
    _orig = query_db(
        """SELECT id FROM accounting_entries
           WHERE ref_type='PAYMENT' AND ref_id=%s AND ref_table='payments'
             AND is_reversed=FALSE
           ORDER BY id DESC LIMIT 1""",
        [pay_id], one=True
    )
    if _orig:
        execute_db("UPDATE accounting_entries SET is_reversed=TRUE WHERE id=%s", [_orig['id']])
    post_ledger('PAYMENT_REVERSAL', pay_id, 'payments', p['company'],
                f"Void of PAY-{pay_id:04d}",
                debit=float(p['amount']), credit=0, entry_date=p['pay_date'])
    execute_db('UPDATE payments SET voided=TRUE WHERE id=%s', [pay_id])
    log_action('VOID', 'payments', pay_id,
               f"{p['company']} | Amount:{p['amount']} | VOIDED")
    flash(f'Payment #{pay_id} voided. Ledger entry reversed.', 'warning')
    return redirect(url_for('payments'))

# ── Deliver Tomorrow — shared helpers ─────────────────────────────────────────

def _dt_parse_filters(args):
    """Parse all Deliver Tomorrow filter params from a request args or form dict."""
    tomorrow_date = (date.today() + timedelta(days=1)).strftime('%Y-%m-%d')
    view_date   = (args.get('view_date',   '') or '').strip() or tomorrow_date
    f_date_from = (args.get('date_from',   '') or '').strip()
    f_date_to   = (args.get('date_to',     '') or '').strip()
    f_company   = (args.get('company',     '') or '').strip().upper()
    f_supplier  = (args.get('supplier',    '') or '').strip().upper()
    f_passenger = (args.get('passenger',   '') or '').strip().upper()
    f_status    = (args.get('status',      '') or '').strip().upper()
    f_agent     = (args.get('agent_id',    '') or '').strip()

    # When date_from/to are set they define the date window; otherwise single view_date
    start_date = f_date_from or view_date
    end_date   = f_date_to   or view_date

    extra_where, extra_params = [], []
    if f_company:
        extra_where.append("AND UPPER(s.company) LIKE %s")
        extra_params.append('%' + f_company + '%')
    if f_supplier:
        extra_where.append("AND (UPPER(COALESCE(s.buy_from,'')) LIKE %s"
                           " OR UPPER(COALESCE(s.return_supplier,'')) LIKE %s)")
        extra_params.extend(['%' + f_supplier + '%', '%' + f_supplier + '%'])
    if f_passenger:
        extra_where.append("AND UPPER(s.customer) LIKE %s")
        extra_params.append('%' + f_passenger + '%')
    if f_status:
        extra_where.append("AND (COALESCE(s.outbound_status,s.status,'PENDING')=%s"
                           " OR COALESCE(s.return_status,s.status,'PENDING')=%s)")
        extra_params.extend([f_status, f_status])
    if f_agent:
        extra_where.append("AND s.created_by_user_id=%s")
        extra_params.append(f_agent)

    return dict(
        view_date=view_date, start_date=start_date, end_date=end_date,
        f_company=f_company, f_supplier=f_supplier, f_passenger=f_passenger,
        f_status=f_status, f_agent=f_agent,
        f_date_from=f_date_from, f_date_to=f_date_to,
        extra_str=' '.join(extra_where),
        extra_params=extra_params,
    )


def _dt_query_tickets(p):
    """Run outbound + return delivery queries using a parsed filter dict."""
    agent_select = "COALESCE(u.full_name, s.created_by_username, '—') AS agent_display"
    agent_join   = "LEFT JOIN users u ON s.created_by_user_id = u.id"
    es, ep       = p['extra_str'], p['extra_params']
    sd, ed       = p['start_date'], p['end_date']

    outbound = query_db(f'''
        SELECT s.*, {agent_select}, 'OUTBOUND' AS delivery_type
        FROM sales s {agent_join}
        WHERE s.deleted=FALSE
          AND (
            (s.outbound_delivery != '' AND s.outbound_delivery BETWEEN %s AND %s)
            OR
            (COALESCE(s.outbound_delivery,'')='' AND s.travel_date BETWEEN %s AND %s
             AND COALESCE(s.travel_date,'') != '')
          )
        {es}
        ORDER BY s.company, s.customer
    ''', [sd, ed, sd, ed] + ep) or []

    returns = query_db(f'''
        SELECT s.*, {agent_select}, 'RETURN' AS delivery_type
        FROM sales s {agent_join}
        WHERE s.deleted=FALSE
          AND (
            (s.return_delivery != '' AND s.return_delivery BETWEEN %s AND %s)
            OR
            (COALESCE(s.return_delivery,'')='' AND s.return_date BETWEEN %s AND %s
             AND COALESCE(s.return_date,'') != '')
          )
        {es}
        ORDER BY s.company, s.customer
    ''', [sd, ed, sd, ed] + ep) or []

    return outbound, returns


def _dt_build_excel(outbound_tickets, return_tickets, label):
    """Build Deliver Tomorrow Excel in-memory and return raw bytes."""
    _thin = Side(style='thin')
    _bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    hdr_font = Font(bold=True, color='FFFFFF')
    out_fill = PatternFill('solid', fgColor='1B3A6B')
    ret_fill = PatternFill('solid', fgColor='FFF7ED')
    center   = Alignment(horizontal='center', vertical='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Deliver Tomorrow'

    headers = ['#', 'Type', 'Customer Name', 'Company', 'Route',
               'Airline / Supplier', 'Ticket Number', 'Travel Date', 'Delivery Status']
    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(1, ci)
        c.font = hdr_font; c.fill = out_fill
        c.alignment = center; c.border = _bdr

    row_num = 1
    for t in outbound_tickets:
        row_num += 1
        status = t.get('outbound_status') or t.get('status') or 'PENDING'
        ws.append([
            row_num - 1, 'OUTBOUND',
            t.get('customer', ''), t.get('company', ''),
            f"{t.get('from_loc','') or ''} → {t.get('to_loc','') or ''}",
            t.get('buy_from') or '—',
            t.get('ticket_number') or '—',
            str(t.get('travel_date') or '—'),
            status,
        ])
        for ci in range(1, len(headers)+1):
            ws.cell(row_num, ci).border = _bdr

    for t in return_tickets:
        row_num += 1
        status = t.get('return_status') or t.get('status') or 'PENDING'
        ws.append([
            row_num - 1, 'RETURN',
            t.get('customer', ''), t.get('company', ''),
            f"{t.get('to_loc','') or ''} → {t.get('from_loc','') or ''}",
            t.get('return_supplier') or t.get('buy_from') or '—',
            t.get('return_ticket_number') or t.get('ticket_number') or '—',
            str(t.get('return_date') or '—'),
            status,
        ])
        for ci in range(1, len(headers)+1):
            c = ws.cell(row_num, ci)
            c.fill = ret_fill; c.border = _bdr

    for col, width in zip('ABCDEFGHI', [5, 10, 26, 22, 18, 24, 22, 14, 14]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _dt_send_email(outbound_tickets, return_tickets, label, sender_username='system'):
    """
    Send Deliver Tomorrow email with Excel attachment to all NOTIFY_EMAIL recipients.
    Returns (success: bool, message: str).
    NOTIFY_EMAIL supports comma-separated addresses.
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    from email.mime.base      import MIMEBase
    from email                import encoders

    smtp_host     = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    smtp_port     = int(os.environ.get('SMTP_PORT', 587))
    smtp_user     = os.environ.get('SMTP_USER', '')
    smtp_password = os.environ.get('SMTP_PASSWORD', '')
    email_to_raw  = os.environ.get('NOTIFY_EMAIL', smtp_user)

    if not smtp_user or not smtp_password:
        return False, ('Email not configured. '
                       'Set SMTP_USER, SMTP_PASSWORD, NOTIFY_EMAIL in environment variables.')

    recipients = [r.strip() for r in email_to_raw.split(',') if r.strip()]
    if not recipients:
        return False, 'No valid recipient addresses found in NOTIFY_EMAIL.'

    total    = len(outbound_tickets) + len(return_tickets)
    xl_bytes = _dt_build_excel(outbound_tickets, return_tickets, label)
    xl_name  = f"deliver_{label.replace(' ','_')}.xlsx"

    # Build HTML summary table
    rows_html = ''
    for i, t in enumerate(outbound_tickets + return_tickets, 1):
        is_ret = (t.get('delivery_type') == 'RETURN')
        status = (t.get('return_status') if is_ret else t.get('outbound_status')) \
                 or t.get('status') or 'PENDING'
        sc     = '#1E7B34' if status == 'DONE' else '#E67E22'
        route  = (f"{t.get('to_loc','')} → {t.get('from_loc','')}" if is_ret
                  else f"{t.get('from_loc','')} → {t.get('to_loc','')}")
        supp   = ((t.get('return_supplier') or t.get('buy_from') or '—') if is_ret
                  else (t.get('buy_from') or '—'))
        dt     = str((t.get('return_date') if is_ret else t.get('travel_date')) or '—')
        dtype  = 'RETURN' if is_ret else 'OUTBOUND'
        bg     = '#fff7ed' if is_ret else ('#f9fafb' if i % 2 == 0 else '#ffffff')
        rows_html += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:7px 10px;border:1px solid #ddd;text-align:center">{i}</td>'
            f'<td style="padding:7px 10px;border:1px solid #ddd;font-weight:700;'
            f'color:{"#8B6914" if is_ret else "#1B3A6B"};font-size:11px">{dtype}</td>'
            f'<td style="padding:7px 10px;border:1px solid #ddd;font-weight:700">{t.get("company","")}</td>'
            f'<td style="padding:7px 10px;border:1px solid #ddd">{t.get("customer","")}</td>'
            f'<td style="padding:7px 10px;border:1px solid #ddd;text-align:center"><strong>{route}</strong></td>'
            f'<td style="padding:7px 10px;border:1px solid #ddd;text-align:center">{supp}</td>'
            f'<td style="padding:7px 10px;border:1px solid #ddd;text-align:center">{dt}</td>'
            f'<td style="padding:7px 10px;border:1px solid #ddd;text-align:center;'
            f'color:{sc};font-weight:700">{status}</td>'
            f'</tr>'
        )

    html_body = f"""<html><body style="font-family:Arial,sans-serif;margin:0;padding:20px;background:#f4f7fc">
<div style="max-width:900px;margin:0 auto;background:#fff;border-radius:10px;overflow:hidden;
            box-shadow:0 2px 12px rgba(0,0,0,.1)">
  <div style="background:#1B3A6B;padding:24px 28px">
    <h1 style="color:#C8A84B;margin:0;font-size:20px">&#9992; ALSONDOS TRAVEL</h1>
    <h2 style="color:#fff;margin:8px 0 0;font-size:16px">Daily Deliver Tomorrow Report &mdash; {label}</h2>
    <p style="color:rgba(255,255,255,.7);margin:4px 0 0;font-size:13px">
      {len(outbound_tickets)} outbound &nbsp;&middot;&nbsp; {len(return_tickets)} return
      &nbsp;&middot;&nbsp; {total} total tickets
    </p>
  </div>
  <div style="padding:20px">
    <p style="color:#374151;font-size:13px">Please find attached today's Deliver Tomorrow report.</p>
    <table style="width:100%;border-collapse:collapse;font-size:12px">
      <thead>
        <tr style="background:#1B3A6B;color:#fff">
          <th style="padding:9px 10px;border:1px solid #1B3A6B">#</th>
          <th style="padding:9px 10px;border:1px solid #1B3A6B">Type</th>
          <th style="padding:9px 10px;border:1px solid #1B3A6B">Company</th>
          <th style="padding:9px 10px;border:1px solid #1B3A6B">Customer</th>
          <th style="padding:9px 10px;border:1px solid #1B3A6B">Route</th>
          <th style="padding:9px 10px;border:1px solid #1B3A6B">Supplier</th>
          <th style="padding:9px 10px;border:1px solid #1B3A6B">Travel Date</th>
          <th style="padding:9px 10px;border:1px solid #1B3A6B">Status</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
    <p style="color:#6B7A99;font-size:11px;margin-top:16px;text-align:center">
      Generated by ALSONDOS TRAVEL System &middot; {label} &middot; Sent by {sender_username}
    </p>
  </div>
</div>
</body></html>"""

    msg = MIMEMultipart('mixed')
    msg['Subject'] = f'Daily Deliver Tomorrow Report - {label}'
    msg['From']    = smtp_user
    msg['To']      = ', '.join(recipients)
    msg.attach(MIMEText(html_body, 'html'))

    att = MIMEBase('application',
                   'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    att.set_payload(xl_bytes)
    encoders.encode_base64(att)
    att.add_header('Content-Disposition', f'attachment; filename="{xl_name}"')
    msg.attach(att)

    try:
        import smtplib
        with smtplib.SMTP(smtp_host, smtp_port, timeout=25) as srv:
            srv.ehlo(); srv.starttls(); srv.ehlo()
            srv.login(smtp_user, smtp_password)
            srv.sendmail(smtp_user, recipients, msg.as_string())
        logger.info(f"Deliver Tomorrow email sent to {recipients} ({total} tickets)")
        return True, f'Email sent to {", ".join(recipients)} — {total} ticket(s).'
    except Exception as ex:
        logger.error(f"Deliver Tomorrow email failed: {ex}")
        return False, f'Email failed: {ex}'


# ── Deliver Tomorrow — routes ──────────────────────────────────────────────────
def _dt_auto_update_statuses():
    """
    C-04 FIX: Auto-update delivery statuses based on today's date.
    Called ONLY from the daily cron endpoint, not on every page load.
    Marks tickets as DONE when their delivery date has passed.
    """
    today_str = str(date.today())
    execute_db('''
        UPDATE sales SET outbound_status='DONE'
        WHERE outbound_delivery != '' AND outbound_delivery <= %s
          AND outbound_status='PENDING' AND deleted=FALSE AND is_archived=FALSE
    ''', [today_str])
    execute_db('''
        UPDATE sales SET return_status='DONE'
        WHERE return_delivery != '' AND return_delivery <= %s
          AND return_status='PENDING' AND deleted=FALSE AND is_archived=FALSE
    ''', [today_str])
    execute_db('''
        UPDATE sales SET status='DONE'
        WHERE outbound_status='DONE'
          AND (return_delivery='' OR return_status='DONE')
          AND status != 'DONE' AND deleted=FALSE AND is_archived=FALSE
    ''')
    logger.info("_dt_auto_update_statuses: delivery statuses refreshed")


@app.route('/deliver-tomorrow')
@login_required
def deliver_tomorrow():
    tomorrow_date = (date.today() + timedelta(days=1)).strftime('%Y-%m-%d')

    p = _dt_parse_filters(request.args)
    outbound_tickets, return_tickets = _dt_query_tickets(p)
    travel_date_tickets = []

    agents    = query_db(
        "SELECT id, COALESCE(full_name, username) AS name FROM users ORDER BY full_name, username"
    ) or []
    companies = get_companies_list()

    tomorrow_str = (date.today() + timedelta(days=1)).strftime('%d %B %Y')
    try:
        view_date_str = datetime.strptime(p['view_date'], '%Y-%m-%d').strftime('%d %B %Y')
    except Exception:
        view_date_str = tomorrow_str

    return render_template('deliver.html',
        outbound_tickets=outbound_tickets,
        return_tickets=return_tickets,
        travel_date_tickets=travel_date_tickets,
        tomorrow=view_date_str,
        tomorrow_date=tomorrow_date,
        view_date=p['view_date'],
        agents=agents,
        companies=companies,
        filters=dict(
            travel_date='',
            company=p['f_company'],
            supplier=p['f_supplier'],
            passenger=p['f_passenger'],
            status=p['f_status'],
            agent_id=p['f_agent'],
            view_date=p['view_date'],
            date_from=p['f_date_from'],
            date_to=p['f_date_to'],
        )
    )


@app.route('/deliver-tomorrow/export')
@login_required
def deliver_tomorrow_export():
    """Export filtered Deliver Tomorrow view to Excel."""
    p = _dt_parse_filters(request.args)
    outbound_tickets, return_tickets = _dt_query_tickets(p)

    if p['f_date_from'] or p['f_date_to']:
        label = f"{p['start_date']}_to_{p['end_date']}"
    else:
        label = p['view_date']

    xl_bytes = _dt_build_excel(outbound_tickets, return_tickets, label)
    return Response(
        xl_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="deliver_{label}.xlsx"'}
    )


@app.route('/deliver-tomorrow/mark/<int:sale_id>', methods=['POST'])
@login_required
def mark_delivered(sale_id):
    """
    M-06 FIX: Employees can mark any sale delivered (operations task, not ownership).
    Admins and managers always allowed. Plain users allowed only for tickets in
    the delivery dashboard — they are not editing financial data, only status.
    Audit log captures who marked it.
    """
    dtype       = request.form.get('dtype', 'OUTBOUND')
    redirect_q  = request.form.get('redirect_args', '')
    actor       = session.get('username', '?')

    # Verify the sale exists and is not deleted
    sale = query_db('SELECT id, company, customer FROM sales WHERE id=%s AND deleted=FALSE',
                    [sale_id], one=True)
    if not sale:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'error': 'Sale not found'}), 404
        flash('Sale not found.', 'danger')
        return redirect(url_for('deliver_tomorrow'))

    if dtype == 'RETURN':
        execute_db("UPDATE sales SET return_status='DONE' WHERE id=%s AND deleted=FALSE", [sale_id])
        log_action('UPDATE', 'sales', sale_id,
                   f"Return ticket marked DONE by {actor} | {sale['customer']} | {sale['company']}")
    elif dtype == 'TRAVEL':
        execute_db("UPDATE sales SET status='DONE' WHERE id=%s AND deleted=FALSE", [sale_id])
        log_action('UPDATE', 'sales', sale_id,
                   f"Travel ticket marked DONE by {actor} | {sale['customer']} | {sale['company']}")
    else:
        execute_db("UPDATE sales SET outbound_status='DONE' WHERE id=%s AND deleted=FALSE", [sale_id])
        log_action('UPDATE', 'sales', sale_id,
                   f"Outbound ticket marked DONE by {actor} | {sale['customer']} | {sale['company']}")
    # AJAX call → return JSON so the page doesn't reload
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    flash('Ticket marked as delivered.', 'success')
    dest = url_for('deliver_tomorrow')
    if redirect_q:
        dest += '?' + redirect_q
    return redirect(dest)

@app.route('/deliver-tomorrow/send-email', methods=['POST'])
@login_required
def send_deliver_email():
    """Email Team button — sends filtered Deliver Tomorrow report with Excel attachment."""
    import threading

    p = _dt_parse_filters(request.form)
    outbound_tickets, return_tickets = _dt_query_tickets(p)
    total = len(outbound_tickets) + len(return_tickets)

    if total == 0:
        flash('No tickets to send — delivery list is empty for the selected date/filters.', 'warning')
        return redirect(url_for('deliver_tomorrow', **{k: v for k, v in request.form.items()
                                                        if k != 'csrf_token' and v}))

    if p['f_date_from'] or p['f_date_to']:
        label = f"{p['start_date']} to {p['end_date']}"
    else:
        try:
            label = datetime.strptime(p['view_date'], '%Y-%m-%d').strftime('%d %B %Y')
        except Exception:
            label = p['view_date']

    sender = session.get('username', 'system')

    def _send():
        ok, msg_text = _dt_send_email(outbound_tickets, return_tickets, label, sender)
        if not ok:
            logger.error(f"send_deliver_email background error: {msg_text}")

    threading.Thread(target=_send, daemon=True).start()

    email_to_raw = os.environ.get('NOTIFY_EMAIL', os.environ.get('SMTP_USER', ''))
    if not os.environ.get('SMTP_USER') or not os.environ.get('SMTP_PASSWORD'):
        flash('Email not configured. Set SMTP_USER, SMTP_PASSWORD, NOTIFY_EMAIL in environment variables.',
              'danger')
        return redirect(url_for('deliver_tomorrow'))

    log_action('EMAIL', 'sales', None,
               f"Deliver Tomorrow report '{label}' queued to {email_to_raw} "
               f"({len(outbound_tickets)} outbound + {len(return_tickets)} return)")
    flash(f'Deliver Tomorrow report emailed successfully — {total} ticket(s) sent to {email_to_raw}.',
          'success')
    return redirect(url_for('deliver_tomorrow', view_date=p['view_date'],
                             date_from=p['f_date_from'], date_to=p['f_date_to'],
                             company=p['f_company']))


@app.route('/deliver-tomorrow/send-whatsapp', methods=['POST'])
@login_required
def send_whatsapp_deliver():
    """WhatsApp Team button — user-entered recipients, sends Excel via WhatsApp Cloud API."""
    import io, re as _re, requests as _req

    # ── API credentials from environment (never from client) ──────────────
    wa_token    = os.environ.get('WA_TOKEN', '')
    wa_phone_id = os.environ.get('WA_PHONE_ID', '')

    if not wa_token or not wa_phone_id:
        return jsonify({'ok': False,
                        'error': 'WhatsApp not configured on server. '
                                 'Set WA_TOKEN and WA_PHONE_ID in environment variables.'}), 400

    # ── Recipients from the modal form (user-entered) ─────────────────────
    raw_numbers = request.form.getlist('wa_number')      # multiple <input name="wa_number">
    wa_numbers  = [n.strip() for n in raw_numbers if n.strip()]

    if not wa_numbers:
        return jsonify({'ok': False, 'error': 'Enter at least one WhatsApp number.'}), 400

    # Validate E.164: + then 7–15 digits
    _e164 = _re.compile(r'^\+\d{7,15}$')
    bad   = [n for n in wa_numbers if not _e164.match(n)]
    if bad:
        return jsonify({'ok': False,
                        'error': f'Invalid number format (use +CountryCode…): {", ".join(bad)}'}), 400

    # Strip leading + for Meta API (it expects digits only)
    api_numbers = [n.lstrip('+') for n in wa_numbers]

    # ── Build Excel from current filters ──────────────────────────────────
    p = _dt_parse_filters(request.form)
    outbound_tickets, return_tickets = _dt_query_tickets(p)
    total = len(outbound_tickets) + len(return_tickets)

    if total == 0:
        return jsonify({'ok': False, 'error': 'No tickets for the selected date/filters.'}), 400

    if p['f_date_from'] or p['f_date_to']:
        label     = f"{p['start_date']}_to_{p['end_date']}"
        label_txt = f"{p['start_date']} to {p['end_date']}"
    else:
        label = p['view_date']
        try:
            label_txt = datetime.strptime(p['view_date'], '%Y-%m-%d').strftime('%d %B %Y')
        except Exception:
            label_txt = p['view_date']

    xl_bytes = _dt_build_excel(outbound_tickets, return_tickets, label)
    filename = f"deliver_{label}.xlsx"
    sender   = session.get('username', 'system')

    graph_url = f"https://graph.facebook.com/v19.0/{wa_phone_id}"
    auth_hdr  = {'Authorization': f'Bearer {wa_token}'}

    # ── Upload Excel once, reuse media_id for all recipients ──────────────
    try:
        up = _req.post(
            f"{graph_url}/media",
            headers=auth_hdr,
            files={
                'file': (filename, io.BytesIO(xl_bytes),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                'type': (None, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                'messaging_product': (None, 'whatsapp'),
            },
            timeout=30,
        )
        up.raise_for_status()
        media_id = up.json().get('id')
        if not media_id:
            raise ValueError(f"No media ID in response: {up.text}")
    except Exception as _ue:
        logger.error(f"send_whatsapp_deliver upload error: {_ue}")
        return jsonify({'ok': False, 'error': f'File upload to WhatsApp failed: {_ue}'}), 500

    caption = (
        f"\U0001f4cb Deliver Tomorrow — {label_txt}\n"
        f"✈ {len(outbound_tickets)} outbound  |  \U0001f6ec {len(return_tickets)} return\n"
        f"Sent by {sender}"
    )

    # ── Send document to each number ──────────────────────────────────────
    sent, errors = [], []
    for number in api_numbers:
        try:
            r = _req.post(
                f"{graph_url}/messages",
                headers={**auth_hdr, 'Content-Type': 'application/json'},
                json={
                    'messaging_product': 'whatsapp',
                    'to': number,
                    'type': 'document',
                    'document': {'id': media_id, 'filename': filename, 'caption': caption},
                },
                timeout=20,
            )
            r.raise_for_status()
            sent.append('+' + number)
        except Exception as _me:
            api_err = ''
            try: api_err = r.json().get('error', {}).get('message', '')
            except Exception: pass
            msg = api_err or str(_me)
            logger.error(f"send_whatsapp_deliver message error to +{number}: {msg}")
            errors.append(f"+{number}: {msg}")

    recipients_str = ', '.join(wa_numbers)
    log_action('WHATSAPP', 'sales', None,
               f"Deliver Tomorrow '{label_txt}' sent via WhatsApp to [{recipients_str}] "
               f"by {sender} ({len(outbound_tickets)} outbound + {len(return_tickets)} return)"
               + (f" — ERRORS: {'; '.join(errors)}" if errors else ''))

    if errors and not sent:
        return jsonify({'ok': False,
                        'error': 'Failed to send to all numbers: ' + '; '.join(errors)}), 500
    if errors:
        return jsonify({'ok': True,
                        'message': f'Sent to {len(sent)} number(s). Failed: {"; ".join(errors)}',
                        'partial': True}), 207

    return jsonify({'ok': True,
                    'message': f'Report sent to {len(sent)} WhatsApp number(s). '
                               f'{len(outbound_tickets)} outbound + {len(return_tickets)} return.',
                    'total': total})


@app.route('/deliver-tomorrow/daily-email', methods=['POST'])
@csrf.exempt
def deliver_tomorrow_daily_email():
    """
    Automated daily email endpoint — call from OS cron, not from the browser.

    Cron setup on Hostinger VPS (fires at 08:00 every day):
        0 8 * * * curl -s -X POST -H "X-Cron-Key: YOUR_CRON_KEY" \\
            http://127.0.0.1:5000/deliver-tomorrow/daily-email >> /var/log/deliver_email.log 2>&1

    Set CRON_KEY env var to a random secret to protect the endpoint.
    """
    cron_key = os.environ.get('CRON_KEY', '')
    if cron_key and request.headers.get('X-Cron-Key', '') != cron_key:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401

    # C-04 FIX: auto-update delivery statuses runs here (daily cron) not on page load
    try:
        _dt_auto_update_statuses()
    except Exception as _ue:
        logger.error(f"deliver_tomorrow_daily_email: status update failed: {_ue}")

    tomorrow_date = (date.today() + timedelta(days=1)).strftime('%Y-%m-%d')
    p = _dt_parse_filters({'view_date': tomorrow_date})
    outbound_tickets, return_tickets = _dt_query_tickets(p)
    total = len(outbound_tickets) + len(return_tickets)

    try:
        label = datetime.strptime(tomorrow_date, '%Y-%m-%d').strftime('%d %B %Y')
    except Exception:
        label = tomorrow_date

    ok, msg_text = _dt_send_email(outbound_tickets, return_tickets, label, sender_username='cron')
    log_action('EMAIL', 'sales', None,
               f"Auto daily email '{label}': {msg_text} ({total} tickets)")
    return jsonify({'ok': ok, 'message': msg_text, 'tickets': total})

# ── Admin DB viewer ───────────────────────────────────────────────────────────
@app.route('/admin')
@admin_required
def admin():
    company   = request.args.get('company', '')
    status    = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    table     = request.args.get('table', 'sales')
    page      = max(1, int(request.args.get('page', 1)))

    companies = get_companies_list()

    sales_data, payments_data = [], []
    total_pages = total_rows = total_payments = 1

    if table == 'payments':
        pq = 'SELECT * FROM payments WHERE deleted=FALSE AND is_archived=FALSE'
        pp = []
        if company:   pq += ' AND company=%s';   pp.append(company)
        if date_from: pq += ' AND pay_date>=%s'; pp.append(date_from)
        if date_to:   pq += ' AND pay_date<=%s'; pp.append(date_to)
        pq += ' ORDER BY pay_date DESC, id DESC'
        payments_data, total_rows, total_pages = paginate(pq, pp, page)
        total_payments = sum(r['amount'] for r in payments_data)
    else:
        sq = 'SELECT * FROM sales WHERE deleted=FALSE AND is_archived=FALSE'
        sp = []
        if company:   sq += ' AND company=%s';    sp.append(company)
        if status:    sq += ' AND status=%s';     sp.append(status)
        if date_from: sq += ' AND sale_date>=%s'; sp.append(date_from)
        if date_to:   sq += ' AND sale_date<=%s'; sp.append(date_to)
        sq += ' ORDER BY sale_date DESC, id DESC'
        sales_data, total_rows, total_pages = paginate(sq, sp, page)

    db_stats = query_db('''
        SELECT
            (SELECT COUNT(*) FROM sales WHERE deleted=FALSE AND is_archived=FALSE) as sales_count,
            (SELECT COUNT(*) FROM payments WHERE deleted=FALSE AND is_archived=FALSE) as payments_count,
            (SELECT COALESCE(SUM(sell),0) FROM sales WHERE deleted=FALSE AND is_archived=FALSE) as total_sell,
            (SELECT COALESCE(SUM(profit),0) FROM sales WHERE deleted=FALSE AND is_archived=FALSE) as total_profit,
            (SELECT COALESCE(SUM(amount),0) FROM payments WHERE deleted=FALSE AND is_archived=FALSE) as total_paid
    ''', one=True)

    return render_template('admin.html',
        sales=sales_data, payments=payments_data,
        companies=companies, db_stats=db_stats, table=table,
        filters={'company':company,'status':status,'date_from':date_from,'date_to':date_to},
        total_payments=total_payments, today=date.today().strftime('%d %B %Y'),
        page=page, total_pages=total_pages, total_rows=total_rows
    )

@app.route('/admin/delete-payment/<int:pay_id>', methods=['POST'])
@admin_required
def delete_payment(pay_id):
    p = query_db('SELECT company, amount, pay_date FROM payments WHERE id=%s', [pay_id], one=True)
    if p:
        _orig = query_db(
            """SELECT id FROM accounting_entries
               WHERE ref_type='PAYMENT' AND ref_id=%s AND ref_table='payments'
                 AND is_reversed=FALSE
               ORDER BY id DESC LIMIT 1""",
            [pay_id], one=True
        )
        if _orig:
            execute_db("UPDATE accounting_entries SET is_reversed=TRUE WHERE id=%s", [_orig['id']])
        post_ledger('PAYMENT_REVERSAL', pay_id, 'payments', p['company'],
                    f"Reversal of PAY-{pay_id:04d} (deleted via admin)",
                    debit=p['amount'], credit=0, entry_date=p['pay_date'])
    execute_db('UPDATE payments SET deleted=TRUE WHERE id=%s', [pay_id])
    log_action('DELETE', 'payments', pay_id,
               f"{p['company'] if p else ''} | {p['amount'] if p else ''}")
    return redirect(url_for('admin', table='payments'))

# ── Audit Log ─────────────────────────────────────────────────────────────────
@app.route('/audit')
@admin_required
def audit_log():
    action     = request.args.get('action', '')
    table_name = request.args.get('table_name', '')
    username   = request.args.get('username', '')
    date_from  = request.args.get('date_from', '')
    date_to    = request.args.get('date_to', '')
    page       = max(1, int(request.args.get('page', 1)))

    q = 'SELECT * FROM audit_logs WHERE 1=1'
    p = []
    if action:     q += ' AND action=%s';             p.append(action)
    if table_name: q += ' AND table_name=%s';         p.append(table_name)
    if username:   q += ' AND username ILIKE %s';     p.append(f'%{username}%')
    if date_from:  q += ' AND created_at>=%s';        p.append(date_from)
    if date_to:    q += ' AND created_at<=%s';        p.append(date_to + ' 23:59:59')
    q += ' ORDER BY created_at DESC'

    logs, total_rows, total_pages = paginate(q, p, page)

    return render_template('audit.html',
        logs=logs, page=page, total_pages=total_pages, total_rows=total_rows,
        filters={'action':action,'table_name':table_name,'username':username,
                 'date_from':date_from,'date_to':date_to}
    )

# ── Excel Export ──────────────────────────────────────────────────────────────
@app.route('/export/excel')
@finance_required
def export_excel():
    wb = openpyxl.Workbook()
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1B3A6B")
    gold_fill    = PatternFill("solid", fgColor="C8A84B")
    center       = Alignment(horizontal="center")
    currency_fmt = '#,##0.00'

    ws1 = wb.active
    ws1.title = "Sales"
    hdrs = ["ID","Sale Date","Company","Customer","From","To","Via",
            "Trip Type","Buy From","Tickets","Travel Date",
            "Net (JOD)","Sell (JOD)","Profit (JOD)","Status","Remarks","Sales Agent"]
    ws1.append(hdrs)
    for col in range(1, len(hdrs)+1):
        c = ws1.cell(row=1, column=col)
        c.font = header_font; c.fill = header_fill; c.alignment = center

    sales = query_db('''
        SELECT s.*, COALESCE(u.full_name, s.created_by_username, '—') AS agent_display
        FROM sales s
        LEFT JOIN users u ON s.created_by_user_id = u.id
        WHERE s.deleted=FALSE AND s.is_archived=FALSE
        ORDER BY s.sale_date DESC, s.id DESC
    ''')
    for s in sales:
        ws1.append([s['id'],s['sale_date'],s['company'],s['customer'],
                    s['from_loc'],s['to_loc'],s['via'],s['trip_type'],
                    s['buy_from'],s['tickets'],s['travel_date'],
                    s['net'],s['sell'],s['profit'],s['status'],
                    s['remarks'], s['agent_display']])
    for row in ws1.iter_rows(min_row=2, min_col=12, max_col=14):
        for cell in row: cell.number_format = currency_fmt

    tr = ws1.max_row + 1
    ws1.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    for col, attr in [(12,'net'),(13,'sell'),(14,'profit')]:
        c = ws1.cell(row=tr, column=col, value=sum(s[attr] for s in sales))
        c.font = Font(bold=True); c.fill = gold_fill; c.number_format = currency_fmt
    for i, w in enumerate([6,12,20,25,8,8,8,10,10,8,12,13,13,13,10,20,16], 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Payments")
    pay_hdrs = ["ID","Pay Date","Company","Amount (JOD)","Notes"]
    ws2.append(pay_hdrs)
    for col in range(1, len(pay_hdrs)+1):
        c = ws2.cell(row=1, column=col)
        c.font = header_font; c.fill = header_fill; c.alignment = center

    pmts = query_db('SELECT * FROM payments WHERE deleted=FALSE AND is_archived=FALSE ORDER BY pay_date DESC')
    for p in pmts:
        ws2.append([p['id'],p['pay_date'],p['company'],p['amount'],p['notes']])
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row: cell.number_format = currency_fmt

    tr2 = ws2.max_row + 1
    ws2.cell(row=tr2, column=3, value="TOTAL").font = Font(bold=True)
    c = ws2.cell(row=tr2, column=4, value=sum(p['amount'] for p in pmts))
    c.font = Font(bold=True); c.fill = gold_fill; c.number_format = currency_fmt
    for i, w in enumerate([6,12,22,14,30], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    log_action('EXPORT', 'sales', None, 'Excel export downloaded')
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    filename = f"alsondos_{date.today().strftime('%Y%m%d')}.xlsx"
    return Response(output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.route('/admin/reset-data/backup-excel')
@admin_required
def reset_data_backup():
    """
    C-03 FIX: Must download this backup BEFORE the reset form will activate.
    Sets a session flag so the reset page knows a backup was taken this session.
    """
    import io as _io
    import openpyxl as _xl
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align

    wb = _xl.Workbook()
    hdr_font = _Font(bold=True, color='FFFFFF')
    hdr_fill = _Fill('solid', fgColor='1B3A6B')
    center   = _Align(horizontal='center')

    # ── Sales sheet ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Sales'
    hdrs1 = ['ID','TXN#','Sale Date','Company','Customer','From','To',
             'Service','Buy From','Tickets','Travel Date','Net','Sell','Profit',
             'Status','Remarks','Agent']
    ws1.append(hdrs1)
    for c in range(1, len(hdrs1)+1):
        ws1.cell(1,c).font = hdr_font
        ws1.cell(1,c).fill = hdr_fill
        ws1.cell(1,c).alignment = center
    sales_all = query_db("""
        SELECT s.*, COALESCE(u.full_name, s.created_by_username, '—') AS agent_display
        FROM sales s LEFT JOIN users u ON s.created_by_user_id = u.id
        ORDER BY s.id
    """) or []
    for s in sales_all:
        ws1.append([s['id'], s.get('txn_number',''), s['sale_date'],
                    s['company'], s['customer'], s.get('from_loc',''), s.get('to_loc',''),
                    s.get('service_type','FLIGHT'), s.get('buy_from',''),
                    s.get('tickets',1), s.get('travel_date',''),
                    float(s.get('net',0) or 0), float(s.get('sell',0) or 0),
                    float(s.get('profit',0) or 0),
                    s.get('status',''), s.get('remarks',''), s['agent_display']])

    # ── Payments sheet ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Payments')
    hdrs2 = ['ID','Pay Date','Company','Amount','Method','Reference','Notes']
    ws2.append(hdrs2)
    for c in range(1, len(hdrs2)+1):
        ws2.cell(1,c).font = hdr_font
        ws2.cell(1,c).fill = hdr_fill
    for p in (query_db('SELECT * FROM payments ORDER BY id') or []):
        ws2.append([p['id'], p['pay_date'], p['company'], float(p['amount'] or 0),
                    p.get('payment_method',''), p.get('reference_number',''), p.get('notes','')])

    # ── Supplier payments sheet ──────────────────────────────────────────────
    ws3 = wb.create_sheet('Supplier Payments')
    hdrs3 = ['ID','Pay Date','Supplier','Amount','Service','Notes']
    ws3.append(hdrs3)
    for c in range(1, len(hdrs3)+1):
        ws3.cell(1,c).font = hdr_font
        ws3.cell(1,c).fill = hdr_fill
    for sp in (query_db('SELECT * FROM supplier_payments ORDER BY id') or []):
        ws3.append([sp['id'], sp['pay_date'], sp['supplier'], float(sp['amount'] or 0),
                    sp.get('service_type',''), sp.get('notes','')])

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)

    # Mark that a backup was taken in this session
    session['reset_backup_taken'] = True
    log_action('EXPORT', 'system', None,
               f"Pre-reset full backup downloaded by {session.get('username')}")

    filename = f"alsondos_BACKUP_{date.today().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/admin/reset-data', methods=['GET', 'POST'])
@admin_required
def reset_data():
    # C-03 FIX: backup must be downloaded before reset is allowed
    backup_taken = session.get('reset_backup_taken', False)

    if request.method == 'POST':
        if not backup_taken:
            flash('You must download the backup Excel first before resetting data.', 'danger')
            return redirect(url_for('reset_data'))

        confirm1 = request.form.get('confirm_text', '').strip()
        confirm2 = request.form.get('confirm_text2', '').strip()
        if confirm1 != 'DELETE ALL DATA':
            flash('First confirmation text incorrect. Nothing was deleted.', 'danger')
            return redirect(url_for('reset_data'))
        if confirm2 != session.get('username', '___'):
            flash('Second confirmation (your username) is incorrect. Nothing was deleted.', 'danger')
            return redirect(url_for('reset_data'))

        db = get_db()
        cur = db.cursor()

        # Get exact counts before deletion — these go into the audit log
        cur.execute('SELECT COUNT(*) AS n FROM sales')
        row = cur.fetchone()
        sales_count = row.get('n', 0) if hasattr(row, 'get') else row[0]
        cur.execute('SELECT COUNT(*) AS n FROM payments')
        row = cur.fetchone()
        pay_count = row.get('n', 0) if hasattr(row, 'get') else row[0]
        cur.execute('SELECT COUNT(*) AS n FROM supplier_payments')
        row = cur.fetchone()
        sp_count = row.get('n', 0) if hasattr(row, 'get') else row[0]
        cur.execute('SELECT COALESCE(SUM(sell),0) AS v FROM sales')
        row = cur.fetchone()
        total_sell = float(row.get('v',0) if hasattr(row,'get') else row[0])

        # Hard delete all transactional data
        for _t in ['accounting_entries','sale_items','refunds',
                   'invoices','vouchers','packages','balance_adjustments',
                   'sales','payments','supplier_payments','audit_logs']:
            try: cur.execute('DELETE FROM ' + _t)
            except Exception: db.rollback()

        # Reset ID sequences to 1
        for _seq in ['sales_id_seq','payments_id_seq','audit_logs_id_seq',
                     'invoices_id_seq','vouchers_id_seq','packages_id_seq',
                     'accounting_entries_id_seq','supplier_payments_id_seq']:
            try: cur.execute('ALTER SEQUENCE ' + _seq + ' RESTART WITH 1')
            except Exception: pass

        # Reset txn_seq so T-numbers restart from T-0001
        try: cur.execute('ALTER SEQUENCE txn_seq RESTART WITH 1')
        except Exception: pass

        # Reset document number sequences
        try: cur.execute("UPDATE doc_sequences SET last_num = 0")
        except Exception: pass

        db.commit()

        # Clear backup flag from session
        session.pop('reset_backup_taken', None)

        # Log the reset (first entry in fresh audit log)
        log_action('RESET', 'system', None,
                   f"Full data reset by {session.get('username')} — "
                   f"deleted {sales_count} sales, {pay_count} payments, "
                   f"{sp_count} supplier payments. Total sell was JOD {total_sell:.3f}. "
                   f"Backup was downloaded before reset.")

        flash(f'✅ All data cleared. {sales_count} sales and {pay_count} payments deleted. '
              f'System starts fresh from today.', 'success')
        return redirect(url_for('index'))

    # GET — show confirmation page with backup requirement
    stats = query_db('''
        SELECT
            (SELECT COUNT(*) FROM sales) as sales_count,
            (SELECT COUNT(*) FROM payments) as payments_count,
            (SELECT COUNT(*) FROM supplier_payments) as sp_count,
            (SELECT COALESCE(SUM(sell),0) FROM sales WHERE deleted=FALSE) as total_sell,
            (SELECT COALESCE(SUM(amount),0) FROM payments WHERE deleted=FALSE) as total_paid
    ''', one=True)
    return render_template('reset_data.html', stats=stats, backup_taken=backup_taken)

@app.route('/archive', methods=['GET'])
@admin_required
def archive():
    company   = request.args.get('company', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    table     = request.args.get('table', 'sales')
    page      = max(1, int(request.args.get('page', 1)))

    companies = [r['company'] for r in query_db(
        'SELECT DISTINCT company FROM sales WHERE is_archived=TRUE ORDER BY company'
    )]

    sales_data, payments_data = [], []
    total_pages = total_rows = 1
    total_payments = 0

    if table == 'payments':
        pq = 'SELECT * FROM payments WHERE deleted=FALSE AND is_archived=TRUE'
        pp = []
        if company:   pq += ' AND company=%s';   pp.append(company)
        if date_from: pq += ' AND pay_date>=%s'; pp.append(date_from)
        if date_to:   pq += ' AND pay_date<=%s'; pp.append(date_to)
        pq += ' ORDER BY pay_date DESC, id DESC'
        payments_data, total_rows, total_pages = paginate(pq, pp, page)
        total_payments = sum(r['amount'] for r in payments_data)
    else:
        sq = 'SELECT * FROM sales WHERE deleted=FALSE AND is_archived=TRUE'
        sp = []
        if company:   sq += ' AND company=%s';    sp.append(company)
        if date_from: sq += ' AND sale_date>=%s'; sp.append(date_from)
        if date_to:   sq += ' AND sale_date<=%s'; sp.append(date_to)
        sq += ' ORDER BY sale_date DESC, id DESC'
        sales_data, total_rows, total_pages = paginate(sq, sp, page)

    archive_stats = query_db('''
        SELECT
            (SELECT COUNT(*) FROM sales WHERE is_archived=TRUE AND deleted=FALSE) as sales_count,
            (SELECT COUNT(*) FROM payments WHERE is_archived=TRUE AND deleted=FALSE) as payments_count,
            (SELECT COALESCE(SUM(sell),0) FROM sales WHERE is_archived=TRUE AND deleted=FALSE) as total_sell,
            (SELECT COALESCE(SUM(profit),0) FROM sales WHERE is_archived=TRUE AND deleted=FALSE) as total_profit,
            (SELECT COALESCE(SUM(amount),0) FROM payments WHERE is_archived=TRUE AND deleted=FALSE) as total_paid
    ''', one=True)

    return render_template('archive.html',
        sales=sales_data, payments=payments_data,
        companies=companies, archive_stats=archive_stats,
        table=table, total_payments=total_payments,
        filters={'company':company,'date_from':date_from,'date_to':date_to},
        page=page, total_pages=total_pages, total_rows=total_rows,
        today=date.today().strftime('%d %B %Y')
    )

@app.route('/archive/do-archive', methods=['GET', 'POST'])
@admin_required
def do_archive():
    if request.method == 'POST':
        archive_date = request.form.get('archive_date', '').strip()
        confirm_text = request.form.get('confirm_text', '').strip()

        if not archive_date:
            flash('Please select an archive date.', 'danger')
            return redirect(url_for('do_archive'))

        if confirm_text != 'ARCHIVE':
            flash('Confirmation text incorrect. Nothing was archived.', 'danger')
            return redirect(url_for('do_archive'))

        # Count what will be archived
        sales_count = query_db(
            'SELECT COUNT(*) as cnt FROM sales WHERE sale_date < %s AND is_archived=FALSE AND deleted=FALSE',
            [archive_date], one=True
        )['cnt']
        pay_count = query_db(
            'SELECT COUNT(*) as cnt FROM payments WHERE pay_date < %s AND is_archived=FALSE AND deleted=FALSE',
            [archive_date], one=True
        )['cnt']

        # Archive sales before the date
        execute_db(
            'UPDATE sales SET is_archived=TRUE WHERE sale_date < %s AND deleted=FALSE',
            [archive_date]
        )
        # Archive payments before the date
        execute_db(
            'UPDATE payments SET is_archived=TRUE WHERE pay_date < %s AND deleted=FALSE',
            [archive_date]
        )

        log_action('ARCHIVE', 'system', None,
                   f"Archived {sales_count} sales and {pay_count} payments before {archive_date}")

        flash(f'✅ Successfully archived {sales_count} sales and {pay_count} payments before {archive_date}. Main system now shows only new data.', 'success')
        return redirect(url_for('index'))

    # GET — show archive form with preview
    preview_date = request.args.get('preview_date', '')
    preview = None
    if preview_date:
        preview = query_db('''
            SELECT
                (SELECT COUNT(*) FROM sales WHERE sale_date < %s AND is_archived=FALSE AND deleted=FALSE) as sales_count,
                (SELECT COUNT(*) FROM payments WHERE pay_date < %s AND is_archived=FALSE AND deleted=FALSE) as payments_count,
                (SELECT COALESCE(SUM(sell),0) FROM sales WHERE sale_date < %s AND is_archived=FALSE AND deleted=FALSE) as total_sell,
                (SELECT COALESCE(SUM(amount),0) FROM payments WHERE pay_date < %s AND is_archived=FALSE AND deleted=FALSE) as total_paid
        ''', [preview_date, preview_date, preview_date, preview_date], one=True)

    return render_template('do_archive.html',
        preview=preview, preview_date=preview_date)

@app.route('/archive/restore-all-sales', methods=['POST'])
@admin_required
def restore_all_sales():
    result = query_db('SELECT COUNT(*) as cnt FROM sales WHERE is_archived=TRUE AND deleted=FALSE', one=True)
    count = result['cnt']
    execute_db('UPDATE sales SET is_archived=FALSE WHERE is_archived=TRUE AND deleted=FALSE')
    log_action('RESTORE', 'sales', None, f"Restored ALL {count} archived sales to active system")
    flash(f'✅ {count} archived sales restored to active system.', 'success')
    return redirect(url_for('archive', table='sales'))

@app.route('/archive/restore-all-payments', methods=['POST'])
@admin_required
def restore_all_payments():
    result = query_db('SELECT COUNT(*) as cnt FROM payments WHERE is_archived=TRUE AND deleted=FALSE', one=True)
    count = result['cnt']
    execute_db('UPDATE payments SET is_archived=FALSE WHERE is_archived=TRUE AND deleted=FALSE')
    log_action('RESTORE', 'payments', None, f"Restored ALL {count} archived payments to active system")
    flash(f'✅ {count} archived payments restored to active system.', 'success')
    return redirect(url_for('archive', table='payments'))

@app.route('/archive/restore-sale/<int:sale_id>', methods=['POST'])
@admin_required
def restore_sale(sale_id):
    execute_db('UPDATE sales SET is_archived=FALSE WHERE id=%s', [sale_id])
    log_action('RESTORE', 'sales', sale_id, f"Sale #{sale_id} restored from archive")
    flash(f'Sale #{sale_id} restored to active system.', 'success')
    return redirect(url_for('archive', table='sales'))

@app.route('/archive/restore-payment/<int:pay_id>', methods=['POST'])
@admin_required
def restore_payment(pay_id):
    execute_db('UPDATE payments SET is_archived=FALSE WHERE id=%s', [pay_id])
    log_action('RESTORE', 'payments', pay_id, f"Payment #{pay_id} restored from archive")
    flash(f'Payment #{pay_id} restored to active system.', 'success')
    return redirect(url_for('archive', table='payments'))

@app.route('/archive/delete-sale/<int:sale_id>', methods=['POST'])
@admin_required
def archive_delete_sale(sale_id):
    s = query_db('SELECT customer, company FROM sales WHERE id=%s', [sale_id], one=True)
    execute_db('UPDATE sales SET deleted=TRUE WHERE id=%s', [sale_id])
    log_action('DELETE', 'sales', sale_id,
               f"Permanently deleted archived sale: {s['customer'] if s else ''}")
    flash(f'Sale #{sale_id} permanently deleted.', 'success')
    return redirect(url_for('archive', table='sales'))

@app.route('/archive/delete-payment/<int:pay_id>', methods=['POST'])
@admin_required
def archive_delete_payment(pay_id):
    p = query_db('SELECT company, amount FROM payments WHERE id=%s', [pay_id], one=True)
    execute_db('UPDATE payments SET deleted=TRUE WHERE id=%s', [pay_id])
    log_action('DELETE', 'payments', pay_id,
               f"Permanently deleted archived payment: {p['company'] if p else ''}")
    flash(f'Payment #{pay_id} permanently deleted.', 'success')
    return redirect(url_for('archive', table='payments'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

# ══════════════════════════════════════════════════════════════════════════════
#  EXTENSION v2 — User System · Invoices · Vouchers · Packages · Commission
# ══════════════════════════════════════════════════════════════════════════════

def init_extension_db():
    """Run once to add new tables and columns to existing schema."""
    db  = psycopg2.connect(os.environ.get("DATABASE_URL"))
    db.autocommit = True
    cur = db.cursor()

    # ── Add created_by_user to sales ─────────────────────────────────────────
    cur.execute("""
        DO $$ BEGIN
            ALTER TABLE sales ADD COLUMN IF NOT EXISTS created_by_user_id INTEGER
                REFERENCES users(id) ON DELETE SET NULL;
        EXCEPTION WHEN duplicate_column THEN NULL; END $$;
    """)
    # ── Add split cost columns ────────────────────────────────────────────────
    for _col in [
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS outbound_cost REAL DEFAULT 0",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS return_cost   REAL DEFAULT 0",
    ]:
        cur.execute(f"DO $$ BEGIN {_col}; EXCEPTION WHEN duplicate_column THEN NULL; END $$;")
    cur.execute("""
        DO $$ BEGIN
            ALTER TABLE sales ADD COLUMN IF NOT EXISTS created_by_username TEXT DEFAULT '';
        EXCEPTION WHEN duplicate_column THEN NULL; END $$;
    """)

    # ── Add ip_address to audit_logs (SEC-07) ────────────────────────────────
    cur.execute("""
        DO $$ BEGIN
            ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS ip_address TEXT DEFAULT '';
        EXCEPTION WHEN duplicate_column THEN NULL; END $$;
    """)

    # ── Ticket numbers on sales (Phase 3) ────────────────────────────────────
    for _tn_col in [
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS ticket_number TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS return_ticket_number TEXT DEFAULT ''",
    ]:
        cur.execute(f"DO $$ BEGIN {_tn_col}; EXCEPTION WHEN duplicate_column THEN NULL; END $$;")

    # ── Return PNR — separate PNR for return leg when supplier differs ────────
    cur.execute("""
        DO $$ BEGIN
            ALTER TABLE sales ADD COLUMN IF NOT EXISTS return_pnr TEXT DEFAULT '';
        EXCEPTION WHEN duplicate_column THEN NULL; END $$;
    """)

    # ── Extend users table ────────────────────────────────────────────────────
    extra_user_cols = [
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS full_name TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS email TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS phone TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS commission_rate REAL DEFAULT 20.0",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE",
        # Employee Management upgrade — extends the existing users table only,
        # no separate employees/HR table. Single source of truth stays `users`.
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS position TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS department TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS manager_id INTEGER REFERENCES users(id) ON DELETE SET NULL",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS join_date TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS employment_status TEXT DEFAULT 'Active'",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS notes TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS last_login TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS must_change_password BOOLEAN DEFAULT FALSE",
    ]
    for c in extra_user_cols:
        cur.execute(f"DO $$ BEGIN {c}; EXCEPTION WHEN duplicate_column THEN NULL; END $$;")

    # ── Invoices table ────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoices (
            id              SERIAL PRIMARY KEY,
            invoice_number  TEXT NOT NULL UNIQUE,
            sale_id         INTEGER REFERENCES sales(id) ON DELETE SET NULL,
            created_by_id   INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_by      TEXT NOT NULL DEFAULT '',
            company         TEXT NOT NULL DEFAULT '',
            customer        TEXT NOT NULL DEFAULT '',
            service_desc    TEXT DEFAULT '',
            amount          REAL NOT NULL DEFAULT 0,
            discount        REAL NOT NULL DEFAULT 0,
            total           REAL NOT NULL DEFAULT 0,
            status          TEXT NOT NULL DEFAULT 'UNPAID',
            invoice_date    TEXT NOT NULL,
            due_date        TEXT DEFAULT '',
            notes           TEXT DEFAULT '',
            deleted         BOOLEAN DEFAULT FALSE,
            created_at      TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    """)

    # ── Hotel Vouchers table ──────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS vouchers (
            id              SERIAL PRIMARY KEY,
            voucher_number  TEXT NOT NULL UNIQUE,
            sale_id         INTEGER REFERENCES sales(id) ON DELETE SET NULL,
            created_by_id   INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_by      TEXT NOT NULL DEFAULT '',
            guest_name      TEXT NOT NULL DEFAULT '',
            num_guests      INTEGER DEFAULT 1,
            hotel_name      TEXT NOT NULL DEFAULT '',
            hotel_address   TEXT DEFAULT '',
            hotel_phone     TEXT DEFAULT '',
            hotel_contact   TEXT DEFAULT '',
            room_type       TEXT DEFAULT '',
            meal_plan       TEXT DEFAULT '',
            checkin_date    TEXT DEFAULT '',
            checkout_date   TEXT DEFAULT '',
            nights          INTEGER DEFAULT 0,
            include_transfer BOOLEAN DEFAULT FALSE,
            include_tours   BOOLEAN DEFAULT FALSE,
            include_insurance BOOLEAN DEFAULT FALSE,
            arrival_flight  TEXT DEFAULT '',
            pickup_sign     TEXT DEFAULT '',
            driver_contact  TEXT DEFAULT '',
            pickup_time     TEXT DEFAULT '',
            vehicle_type    TEXT DEFAULT '',
            emergency_contact TEXT DEFAULT '',
            cancellation_policy TEXT DEFAULT '',
            remarks         TEXT DEFAULT '',
            deleted         BOOLEAN DEFAULT FALSE,
            created_at      TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    """)

    # ── Packages table ────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS packages (
            id              SERIAL PRIMARY KEY,
            package_number  TEXT NOT NULL UNIQUE,
            sale_id         INTEGER REFERENCES sales(id) ON DELETE SET NULL,
            created_by_id   INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_by      TEXT NOT NULL DEFAULT '',
            company         TEXT DEFAULT '',
            customer        TEXT DEFAULT '',
            package_date    TEXT DEFAULT '',
            -- Flight
            airline         TEXT DEFAULT '',
            flight_route    TEXT DEFAULT '',
            departure_date  TEXT DEFAULT '',
            return_date     TEXT DEFAULT '',
            baggage         TEXT DEFAULT '',
            pnr             TEXT DEFAULT '',
            -- Hotel
            hotel_name      TEXT DEFAULT '',
            room_type       TEXT DEFAULT '',
            meal_plan       TEXT DEFAULT '',
            checkin_date    TEXT DEFAULT '',
            checkout_date   TEXT DEFAULT '',
            nights          INTEGER DEFAULT 0,
            -- Transfer
            transfer_type   TEXT DEFAULT '',
            pickup_time     TEXT DEFAULT '',
            vehicle_type    TEXT DEFAULT '',
            -- Tours
            tour_name       TEXT DEFAULT '',
            tour_date       TEXT DEFAULT '',
            tour_included   BOOLEAN DEFAULT FALSE,
            -- Financials
            net_cost        REAL DEFAULT 0,
            sell_price      REAL DEFAULT 0,
            profit          REAL DEFAULT 0,
            status          TEXT DEFAULT 'ACTIVE',
            notes           TEXT DEFAULT '',
            deleted         BOOLEAN DEFAULT FALSE,
            created_at      TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    """)

    # ── Sequence helpers ──────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS doc_sequences (
            doc_type TEXT PRIMARY KEY,
            last_num INTEGER DEFAULT 0
        )
    """)
    for dt in ('INV', 'VCH', 'PKG'):
        cur.execute("""
            INSERT INTO doc_sequences (doc_type, last_num)
            VALUES (%s, 0) ON CONFLICT (doc_type) DO NOTHING
        """, [dt])

    # ── Service type + extra service columns ─────────────────────────────────────
    svc_cols = [
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS service_type      TEXT DEFAULT 'FLIGHT'",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS hotel_supplier     TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS hotel_name         TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS hotel_room         TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS hotel_meal         TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS hotel_checkin      TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS hotel_checkout     TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS hotel_nights       INTEGER DEFAULT 0",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS hotel_net          REAL DEFAULT 0",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS transfer_supplier  TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS transfer_type      TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS transfer_pickup    TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS transfer_vehicle   TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS transfer_net       REAL DEFAULT 0",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS tours_json         TEXT DEFAULT '[]'",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS visa_supplier      TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS visa_type          TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS passport_number    TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS visa_status        TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS insurance_supplier TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS insurance_type     TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS airline            TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS pnr                TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS baggage            TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS passengers_json    TEXT DEFAULT '[]'",
    ]
    for c in svc_cols:
        cur.execute(f"DO $$ BEGIN {c}; EXCEPTION WHEN duplicate_column THEN NULL; END $$;")
    # Add passengers_json to vouchers too
    cur.execute("""DO $$ BEGIN
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS passengers_json TEXT DEFAULT '[]';
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS flight_details  TEXT DEFAULT '';
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS tours_json      TEXT DEFAULT '[]';
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS airline         TEXT DEFAULT '';
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS pnr             TEXT DEFAULT '';
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS baggage         TEXT DEFAULT '';
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS from_loc        TEXT DEFAULT '';
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS to_loc          TEXT DEFAULT '';
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS departure_date  TEXT DEFAULT '';
        ALTER TABLE vouchers ADD COLUMN IF NOT EXISTS return_date     TEXT DEFAULT '';
    EXCEPTION WHEN duplicate_column THEN NULL; END $$;""")

    # ── Master data tables (companies, suppliers, customers) ─────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS master_companies (
            id         SERIAL PRIMARY KEY,
            name       TEXT NOT NULL UNIQUE,
            phone      TEXT DEFAULT '',
            email      TEXT DEFAULT '',
            address    TEXT DEFAULT '',
            notes      TEXT DEFAULT '',
            is_active  BOOLEAN DEFAULT TRUE,
            created_at TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS master_suppliers (
            id           SERIAL PRIMARY KEY,
            name         TEXT NOT NULL UNIQUE,
            service_type TEXT DEFAULT 'FLIGHT',
            phone        TEXT DEFAULT '',
            email        TEXT DEFAULT '',
            notes        TEXT DEFAULT '',
            is_active    BOOLEAN DEFAULT TRUE,
            created_at   TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS master_customers (
            id             SERIAL PRIMARY KEY,
            full_name      TEXT NOT NULL,
            passport       TEXT DEFAULT '',
            nationality    TEXT DEFAULT '',
            phone          TEXT DEFAULT '',
            email          TEXT DEFAULT '',
            date_of_birth  TEXT DEFAULT '',
            notes          TEXT DEFAULT '',
            is_active      BOOLEAN DEFAULT TRUE,
            created_at     TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    """)
    # Auto-seed master_companies from existing sales data (safe — ignore errors)
    try:
        cur.execute("""
            INSERT INTO master_companies (name)
            SELECT DISTINCT UPPER(TRIM(company))
            FROM sales
            WHERE deleted=FALSE AND TRIM(COALESCE(company,'')) <> ''
            ON CONFLICT (name) DO NOTHING
        """)
    except Exception:
        db.rollback()

    # Auto-seed master_suppliers (each supplier type separately so one failure
    # doesn't block the others — some columns may not exist on first deploy)
    for _supp_sql in [
        "INSERT INTO master_suppliers (name,service_type) SELECT DISTINCT UPPER(TRIM(buy_from)),'FLIGHT' FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(buy_from,''))<>'' ON CONFLICT(name) DO NOTHING",
        "INSERT INTO master_suppliers (name,service_type) SELECT DISTINCT UPPER(TRIM(hotel_supplier)),'HOTEL' FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(hotel_supplier,''))<>'' ON CONFLICT(name) DO NOTHING",
        "INSERT INTO master_suppliers (name,service_type) SELECT DISTINCT UPPER(TRIM(transfer_supplier)),'TRANSFER' FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(transfer_supplier,''))<>'' ON CONFLICT(name) DO NOTHING",
        "INSERT INTO master_suppliers (name,service_type) SELECT DISTINCT UPPER(TRIM(visa_supplier)),'VISA' FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(visa_supplier,''))<>'' ON CONFLICT(name) DO NOTHING",
        "INSERT INTO master_suppliers (name,service_type) SELECT DISTINCT UPPER(TRIM(insurance_supplier)),'INSURANCE' FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(insurance_supplier,''))<>'' ON CONFLICT(name) DO NOTHING",
    ]:
        try:
            cur.execute(_supp_sql)
        except Exception:
            db.rollback()

    # ── Supplier payments table ──────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS supplier_payments (
            id           SERIAL PRIMARY KEY,
            supplier     TEXT NOT NULL,
            amount       REAL NOT NULL DEFAULT 0,
            pay_date     TEXT NOT NULL,
            service_type TEXT DEFAULT 'FLIGHT',
            notes        TEXT DEFAULT '',
            deleted      BOOLEAN DEFAULT FALSE,
            is_archived  BOOLEAN DEFAULT FALSE,
            created_at   TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'))
        )
    """)

    # ── Indexes ───────────────────────────────────────────────────────────────
    for idx in [
        "CREATE INDEX IF NOT EXISTS idx_sales_user ON sales(created_by_user_id)",
        "CREATE INDEX IF NOT EXISTS idx_sales_svctype ON sales(service_type)",
        "CREATE INDEX IF NOT EXISTS idx_invoices_user ON invoices(created_by_id)",
        "CREATE INDEX IF NOT EXISTS idx_invoices_sale ON invoices(sale_id)",
        "CREATE INDEX IF NOT EXISTS idx_vouchers_user ON vouchers(created_by_id)",
        "CREATE INDEX IF NOT EXISTS idx_vouchers_sale ON vouchers(sale_id)",
        "CREATE INDEX IF NOT EXISTS idx_packages_user ON packages(created_by_id)",
        "CREATE INDEX IF NOT EXISTS idx_supp_pay ON supplier_payments(supplier)",
    ]:
        cur.execute(idx)

    # ── Money: REAL → NUMERIC(12,3) — exact decimal, no float rounding ─────
    for _sql in [
        "ALTER TABLE sales ALTER COLUMN net          TYPE NUMERIC(12,3)",
        "ALTER TABLE sales ALTER COLUMN sell         TYPE NUMERIC(12,3)",
        "ALTER TABLE sales ALTER COLUMN profit       TYPE NUMERIC(12,3)",
        "ALTER TABLE sales ALTER COLUMN outbound_cost TYPE NUMERIC(12,3)",
        "ALTER TABLE sales ALTER COLUMN return_cost  TYPE NUMERIC(12,3)",
        "ALTER TABLE sales ALTER COLUMN hotel_net    TYPE NUMERIC(12,3)",
        "ALTER TABLE sales ALTER COLUMN transfer_net TYPE NUMERIC(12,3)",
        "ALTER TABLE payments ALTER COLUMN amount    TYPE NUMERIC(12,3)",
        "ALTER TABLE supplier_payments ALTER COLUMN amount TYPE NUMERIC(12,3)",
    ]:
        try: cur.execute(_sql)
        except Exception: pass

    # ── Performance indexes ──────────────────────────────────────────────────
    for _idx in [
        "CREATE INDEX IF NOT EXISTS idx_sales_company_date ON sales (company, sale_date) WHERE deleted=FALSE",
        "CREATE INDEX IF NOT EXISTS idx_sales_svc_type ON sales (service_type) WHERE deleted=FALSE",
        "CREATE INDEX IF NOT EXISTS idx_sales_buy_from ON sales (buy_from) WHERE deleted=FALSE AND buy_from IS NOT NULL AND buy_from<>''",
        "CREATE INDEX IF NOT EXISTS idx_payments_company ON payments (company, pay_date) WHERE deleted=FALSE",
        "CREATE INDEX IF NOT EXISTS idx_sup_pay_supplier ON supplier_payments (supplier, pay_date) WHERE deleted=FALSE",
        "CREATE INDEX IF NOT EXISTS idx_sales_travel_date ON sales (travel_date) WHERE deleted=FALSE",
        "CREATE INDEX IF NOT EXISTS idx_sales_status ON sales (status) WHERE deleted=FALSE",
        # Phase 4 — UPPER() functional indexes for case-insensitive company/supplier lookups
        # Turns O(n) seqscans in statement() into O(log n) index scans
        "CREATE INDEX IF NOT EXISTS idx_sales_company_upper    ON sales (UPPER(TRIM(company))) WHERE deleted=FALSE",
        "CREATE INDEX IF NOT EXISTS idx_sales_buy_from_upper   ON sales (UPPER(TRIM(buy_from))) WHERE deleted=FALSE AND buy_from IS NOT NULL AND buy_from<>''",
        "CREATE INDEX IF NOT EXISTS idx_payments_company_upper ON payments (UPPER(TRIM(company))) WHERE deleted=FALSE",
        "CREATE INDEX IF NOT EXISTS idx_supp_pay_upper         ON supplier_payments (UPPER(TRIM(supplier))) WHERE deleted=FALSE",
        "CREATE INDEX IF NOT EXISTS idx_accounting_company_upper ON accounting_entries (UPPER(company)) WHERE is_reversed=FALSE",
        # Delivery page performance
        "CREATE INDEX IF NOT EXISTS idx_sales_outbound_delivery ON sales (outbound_delivery) WHERE deleted=FALSE AND outbound_delivery IS NOT NULL AND outbound_delivery<>''",
        "CREATE INDEX IF NOT EXISTS idx_sales_return_delivery   ON sales (return_delivery)   WHERE deleted=FALSE AND return_delivery   IS NOT NULL AND return_delivery  <>''",
    ]:
        try: cur.execute(_idx)
        except Exception: pass

    # ════════════════════════════════════════════════════════════════════════
    # PHASE A — sale_items: one booking → multiple service line items
    # ════════════════════════════════════════════════════════════════════════
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sale_items (
            id           SERIAL PRIMARY KEY,
            sale_id      INTEGER NOT NULL REFERENCES sales(id) ON DELETE CASCADE,
            service_type TEXT NOT NULL DEFAULT 'FLIGHT',
            supplier     TEXT DEFAULT '',
            description  TEXT DEFAULT '',
            net_cost     NUMERIC(12,3) DEFAULT 0,
            sell_price   NUMERIC(12,3) DEFAULT 0,
            quantity     INTEGER DEFAULT 1,
            notes        TEXT DEFAULT '',
            created_at   TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sale_items_sale_id ON sale_items(sale_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sale_items_svc     ON sale_items(service_type)")

    # ════════════════════════════════════════════════════════════════════════
    # PHASE B — accounting_entries: immutable shadow ledger
    # ════════════════════════════════════════════════════════════════════════
    cur.execute("""
        CREATE TABLE IF NOT EXISTS accounting_entries (
            id           SERIAL PRIMARY KEY,
            entry_date   DATE NOT NULL,
            ref_type     TEXT NOT NULL,
            ref_id       INTEGER,
            ref_table    TEXT NOT NULL,
            company      TEXT NOT NULL,
            description  TEXT DEFAULT '',
            debit        NUMERIC(12,3) DEFAULT 0,
            credit       NUMERIC(12,3) DEFAULT 0,
            created_by   INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at   TIMESTAMP DEFAULT NOW(),
            is_reversed  BOOLEAN DEFAULT FALSE,
            reversed_by  INTEGER REFERENCES accounting_entries(id) ON DELETE SET NULL
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_acc_company ON accounting_entries(company, entry_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_acc_ref     ON accounting_entries(ref_table, ref_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_acc_date    ON accounting_entries(entry_date)")

    # ════════════════════════════════════════════════════════════════════════
    # PHASE C — refunds: cancellations and partial refunds
    # ════════════════════════════════════════════════════════════════════════
    cur.execute("""
        CREATE TABLE IF NOT EXISTS refunds (
            id                SERIAL PRIMARY KEY,
            sale_id           INTEGER REFERENCES sales(id) ON DELETE SET NULL,
            refund_date       DATE NOT NULL DEFAULT CURRENT_DATE,
            refund_type       TEXT NOT NULL DEFAULT 'FULL',
            company           TEXT NOT NULL,
            customer          TEXT DEFAULT '',
            gross_amount      NUMERIC(12,3) DEFAULT 0,
            cancellation_fee  NUMERIC(12,3) DEFAULT 0,
            refund_amount     NUMERIC(12,3) DEFAULT 0,
            supplier_penalty  NUMERIC(12,3) DEFAULT 0,
            supplier_refund   NUMERIC(12,3) DEFAULT 0,
            reason            TEXT DEFAULT '',
            status            TEXT DEFAULT 'PENDING',
            processed_by      INTEGER REFERENCES users(id) ON DELETE SET NULL,
            processed_at      TIMESTAMP,
            created_by        INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at        TIMESTAMP DEFAULT NOW(),
            notes             TEXT DEFAULT ''
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_refunds_sale    ON refunds(sale_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_refunds_company ON refunds(company, refund_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_refunds_status  ON refunds(status)")

    # ════════════════════════════════════════════════════════════════════════
    # PHASE D — company_id / supplier_id alongside existing TEXT columns
    # TEXT columns stay — these are additive FK references only
    # ════════════════════════════════════════════════════════════════════════
    for _col_sql in [
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS company_id  INTEGER REFERENCES master_companies(id) ON DELETE SET NULL",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS supplier_id INTEGER REFERENCES master_suppliers(id)  ON DELETE SET NULL",
        "ALTER TABLE payments ADD COLUMN IF NOT EXISTS company_id       INTEGER REFERENCES master_companies(id) ON DELETE SET NULL",
        "ALTER TABLE payments ADD COLUMN IF NOT EXISTS payment_method   TEXT DEFAULT ''",
        "ALTER TABLE payments ADD COLUMN IF NOT EXISTS reference_number TEXT DEFAULT ''",
        "ALTER TABLE payments ADD COLUMN IF NOT EXISTS created_by       INTEGER",
        "ALTER TABLE payments ADD COLUMN IF NOT EXISTS voided           BOOLEAN DEFAULT FALSE",
        "ALTER TABLE supplier_payments ADD COLUMN IF NOT EXISTS supplier_id       INTEGER REFERENCES master_suppliers(id) ON DELETE SET NULL",
        "ALTER TABLE supplier_payments ADD COLUMN IF NOT EXISTS payment_method   TEXT DEFAULT ''",
        "ALTER TABLE supplier_payments ADD COLUMN IF NOT EXISTS reference_number TEXT DEFAULT ''",
    ]:
        try: cur.execute(_col_sql)
        except Exception: pass

    # Backfill company_id from master_companies where name matches
    try:
        cur.execute("""
            UPDATE sales s SET company_id = mc.id
            FROM master_companies mc
            WHERE UPPER(TRIM(s.company)) = UPPER(TRIM(mc.name))
              AND s.company_id IS NULL
        """)
        cur.execute("""
            UPDATE payments p SET company_id = mc.id
            FROM master_companies mc
            WHERE UPPER(TRIM(p.company)) = UPPER(TRIM(mc.name))
              AND p.company_id IS NULL
        """)
        cur.execute("""
            UPDATE sales s SET supplier_id = ms.id
            FROM master_suppliers ms
            WHERE UPPER(TRIM(s.buy_from)) = UPPER(TRIM(ms.name))
              AND s.supplier_id IS NULL
        """)
    except Exception: pass

    cur.execute("CREATE INDEX IF NOT EXISTS idx_sales_company_id  ON sales(company_id)  WHERE company_id IS NOT NULL")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sales_supplier_id ON sales(supplier_id) WHERE supplier_id IS NOT NULL")

    # ════════════════════════════════════════════════════════════════════════
    # PHASE E-1 — txn_number: human-readable transaction number (T-0001...)
    # ════════════════════════════════════════════════════════════════════════
    cur.execute("""
        DO $$ BEGIN
            ALTER TABLE sales ADD COLUMN IF NOT EXISTS txn_number TEXT DEFAULT '';
        EXCEPTION WHEN duplicate_column THEN NULL; END $$;
    """)
    # Backfill existing active records that have no txn_number yet
    cur.execute("""
        UPDATE sales s
        SET txn_number = sub.new_num
        FROM (
            SELECT id,
                   'T-' || LPAD(ROW_NUMBER() OVER (ORDER BY id)::TEXT, 4, '0') AS new_num
            FROM sales
            WHERE deleted=FALSE
        ) sub
        WHERE s.id = sub.id AND (s.txn_number IS NULL OR s.txn_number = '')
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sales_txn_number ON sales(txn_number) WHERE deleted=FALSE")

    # ── C-01 FIX: txn_seq PostgreSQL sequence (race-condition-free numbering) ─
    # Create sequence if it doesn't exist, then sync its value to the current
    # highest T-number so the next generated number is always one higher.
    cur.execute("CREATE SEQUENCE IF NOT EXISTS txn_seq START 1")
    cur.execute("""
        SELECT setval(
            'txn_seq',
            GREATEST(
                (SELECT COALESCE(
                    MAX(CAST(REGEXP_REPLACE(txn_number, '[^0-9]', '', 'g') AS INTEGER)), 0
                ) FROM sales
                WHERE deleted=FALSE
                  AND txn_number ~ '^T-[0-9]+$'),
                currval('txn_seq')
            )
        )
        WHERE EXISTS (SELECT 1 FROM sales WHERE deleted=FALSE AND txn_number ~ '^T-[0-9]+$')
    """)
    # Unique index: prevents duplicate T-numbers at the DB level even if app logic fails
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_sales_txn_number
        ON sales(txn_number)
        WHERE deleted=FALSE AND txn_number IS NOT NULL AND txn_number <> ''
    """)

    # ════════════════════════════════════════════════════════════════════════
    # PHASE E — balance_adjustments: opening balance / historical corrections
    # Never modifies existing transactions. Appears as a normal ledger line.
    # ════════════════════════════════════════════════════════════════════════
    cur.execute("""
        CREATE TABLE IF NOT EXISTS balance_adjustments (
            id           SERIAL PRIMARY KEY,
            account      TEXT NOT NULL,
            account_type TEXT NOT NULL DEFAULT 'COMPANY',
            adj_date     TEXT NOT NULL,
            amount       NUMERIC(12,3) NOT NULL,
            entry_type   TEXT NOT NULL DEFAULT 'CREDIT',
            reason       TEXT NOT NULL DEFAULT '',
            notes        TEXT DEFAULT '',
            created_by   INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at   TEXT DEFAULT (to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS')),
            deleted      BOOLEAN DEFAULT FALSE
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_adj_account ON balance_adjustments(UPPER(account)) WHERE deleted=FALSE")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_adj_date    ON balance_adjustments(adj_date) WHERE deleted=FALSE")

    # ── DB-06 FIX: payments.created_by FK ────────────────────────────────────
    # The column was added as plain INTEGER with no FK. Add the constraint now.
    cur.execute("""
        DO $$ BEGIN
            ALTER TABLE payments
                ADD CONSTRAINT fk_payments_created_by
                FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE SET NULL;
        EXCEPTION WHEN duplicate_object THEN NULL; END $$;
    """)

    # ── DB-03 FIX: unique partial index on txn_number ────────────────────────
    # Belt-and-suspenders: even if app logic fails, the DB rejects duplicates.
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_sales_txn_number
        ON sales(txn_number)
        WHERE deleted=FALSE AND txn_number IS NOT NULL AND txn_number <> ''
    """)

    # ── M-03 FIX: supplier_payments.created_by column ────────────────────────
    # H-04 from audit: supplier_payments had no created_by column.
    cur.execute("""
        DO $$ BEGIN
            ALTER TABLE supplier_payments
                ADD COLUMN IF NOT EXISTS created_by INTEGER REFERENCES users(id) ON DELETE SET NULL;
        EXCEPTION WHEN duplicate_column THEN NULL; END $$;
    """)

    # ── M-03 FIX: supplier_payments voided flag ───────────────────────────────
    # DB-08 from audit: supplier_payments had no voided flag unlike payments.
    cur.execute("""
        DO $$ BEGIN
            ALTER TABLE supplier_payments
                ADD COLUMN IF NOT EXISTS voided BOOLEAN DEFAULT FALSE;
        EXCEPTION WHEN duplicate_column THEN NULL; END $$;
    """)

    # ── Direct Customer columns ────────────────────────────────────────────────
    for _dc_sql in [
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS customer_type     TEXT DEFAULT 'COMPANY'",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS direct_phone      TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS direct_nationality TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS direct_email      TEXT DEFAULT ''",
    ]:
        cur.execute(f"DO $$ BEGIN {_dc_sql}; EXCEPTION WHEN duplicate_column THEN NULL; END $$;")

    # ── Supplier email contact columns ────────────────────────────────────────
    for _se_sql in [
        "ALTER TABLE master_suppliers ADD COLUMN IF NOT EXISTS cc_email       TEXT DEFAULT ''",
        "ALTER TABLE master_suppliers ADD COLUMN IF NOT EXISTS contact_person TEXT DEFAULT ''",
        "ALTER TABLE master_suppliers ADD COLUMN IF NOT EXISTS email_template TEXT DEFAULT ''",
    ]:
        try:
            cur.execute(_se_sql)
        except Exception:
            pass

    db.commit()
    cur.close()
    db.close()

try:
    init_extension_db()
    logger.info("Extension DB initialised OK")
except Exception as _ext_err:
    logger.error(f"Extension DB init error: {_ext_err}")

# ── Direct-customer columns: standalone migration (runs even if init_db failed)
try:
    _dc_conn = psycopg2.connect(os.environ.get("DATABASE_URL"))
    _dc_cur  = _dc_conn.cursor()
    for _dc_sql in [
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS customer_type     TEXT DEFAULT 'COMPANY'",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS direct_phone      TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS direct_nationality TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN IF NOT EXISTS direct_email      TEXT DEFAULT ''",
    ]:
        _dc_cur.execute(_dc_sql)
    _dc_conn.commit()
    _dc_cur.close()
    _dc_conn.close()
    logger.info("Direct-customer columns OK")
except Exception as _dc_err:
    logger.error(f"Direct-customer column migration error: {_dc_err}")


# ── Sequence helper ───────────────────────────────────────────────────────────
def next_txn_number() -> str:
    """
    Generate the next sequential transaction number T-XXXX.
    Uses a PostgreSQL sequence (txn_seq) — atomic, race-condition-free.
    Two concurrent inserts will always get different numbers.
    """
    try:
        row = query_db("SELECT nextval('txn_seq') AS n", one=True)
        num = int(row['n']) if row else 1
        return f"T-{num:04d}"
    except Exception as _e:
        logger.error(f"next_txn_number sequence error: {_e}")
        # Fallback: timestamp-based to avoid collision (never duplicate)
        import time as _time
        return f"T-{int(_time.time() * 100) % 9999999:07d}"


def next_doc_number(doc_type: str) -> str:
    """
    Race-condition-free document numbering for INV/VCH/PKG.

    Uses SELECT FOR UPDATE to lock the doc_sequences row before any sync
    logic, ensuring two concurrent calls can never get the same number.
    Dedicated connection so it never interferes with the request connection.
    """
    table_map = {'INV': 'invoices', 'VCH': 'vouchers', 'PKG': 'packages'}
    table = table_map.get(doc_type)
    conn = None
    try:
        conn = psycopg2.connect(
            os.environ.get("DATABASE_URL"),
            cursor_factory=psycopg2.extras.RealDictCursor
        )
        conn.autocommit = False
        cur = conn.cursor()

        # Lock this doc_type row exclusively before any reads or writes.
        # Any concurrent call blocks here until we commit — no race possible.
        cur.execute(
            "SELECT last_num FROM doc_sequences WHERE doc_type=%s FOR UPDATE",
            [doc_type]
        )
        locked_row = cur.fetchone()
        current_last = int(locked_row['last_num']) if locked_row else 0

        if table:
            # Count ALL rows (including soft-deleted) to detect true empty state
            cur.execute(f"SELECT COUNT(*) as total, COALESCE(MAX(id),0) as max_id FROM {table}")
            row    = cur.fetchone()
            total  = int(row['total'])  if row else 0
            max_id = int(row['max_id']) if row else 0

            if total == 0:
                # Table is truly empty — reset to 0 so next = 0001
                cur.execute("UPDATE doc_sequences SET last_num=0 WHERE doc_type=%s", [doc_type])
                current_last = 0
            elif max_id > current_last:
                # Existing records have higher IDs than sequence — sync up
                cur.execute(
                    "UPDATE doc_sequences SET last_num=%s WHERE doc_type=%s",
                    [max_id, doc_type]
                )
                current_last = max_id

        # Atomically increment and return
        cur.execute(
            "UPDATE doc_sequences SET last_num = last_num + 1 WHERE doc_type=%s RETURNING last_num",
            [doc_type]
        )
        row = cur.fetchone()
        conn.commit()
        num = int(row['last_num']) if row else (current_last + 1)
        return f"{doc_type}-{str(num).zfill(4)}"
    except Exception as _e:
        if conn:
            try: conn.rollback()
            except: pass
        logger.error(f"next_doc_number error: {_e}")
        import uuid
        return f"{doc_type}-{uuid.uuid4().hex[:6].upper()}"
    finally:
        if conn:
            try: conn.close()
            except: pass


# ── Package sync helper ───────────────────────────────────────────────────────
def sync_package_from_sale(sale_id):
    """
    Create or update a packages table row from a PACKAGE-type sale.
    Safe to call multiple times — uses sale_id as the dedup key.
    Does nothing and returns None if sale is not found or not PACKAGE type.
    """
    try:
        sale = query_db(
            "SELECT * FROM sales WHERE id=%s AND deleted=FALSE", [sale_id], one=True
        )
        if not sale or (sale.get('service_type') or 'FLIGHT').upper() != 'PACKAGE':
            return None

        # Pull first tour from tours_json for the simple tour fields
        tours = []
        try:
            raw = sale.get('tours_json') or '[]'
            tours = json.loads(raw) if isinstance(raw, str) else (raw or [])
        except Exception:
            tours = []
        first_tour = tours[0] if tours else {}

        flight_route = ''
        fl = (sale.get('from_loc') or '').strip()
        tl = (sale.get('to_loc')   or '').strip()
        if fl and fl != '-':
            flight_route = f"{fl} → {tl}" if tl and tl != '-' else fl

        existing = query_db(
            "SELECT id, package_number FROM packages WHERE sale_id=%s AND deleted=FALSE",
            [sale_id], one=True
        )

        fields = dict(
            company       = (sale.get('company')       or '').upper().strip(),
            customer      = (sale.get('customer')      or '').upper().strip(),
            package_date  = sale.get('sale_date')      or str(date.today()),
            airline       = (sale.get('airline')       or '').upper().strip(),
            flight_route  = flight_route,
            departure_date= sale.get('travel_date')    or '',
            return_date   = sale.get('return_date')    or '',
            baggage       = sale.get('baggage')        or '',
            pnr           = (sale.get('pnr')           or '').upper().strip(),
            hotel_name    = sale.get('hotel_name')     or '',
            room_type     = sale.get('hotel_room')     or '',
            meal_plan     = sale.get('hotel_meal')     or '',
            checkin_date  = sale.get('hotel_checkin')  or '',
            checkout_date = sale.get('hotel_checkout') or '',
            nights        = int(sale.get('hotel_nights') or 0),
            transfer_type = sale.get('transfer_type')  or '',
            pickup_time   = sale.get('transfer_pickup') or '',
            vehicle_type  = sale.get('transfer_vehicle')or '',
            tour_name     = first_tour.get('name', ''),
            tour_date     = first_tour.get('date', ''),
            tour_included = first_tour.get('status', '') == 'INCLUDED',
            net_cost      = float(sale.get('net')  or 0),
            sell_price    = float(sale.get('sell') or 0),
            profit        = round(float(sale.get('sell') or 0) - float(sale.get('net') or 0), 3),
            notes         = sale.get('remarks') or '',
        )

        if existing:
            execute_db("""
                UPDATE packages SET
                    company=%s, customer=%s, package_date=%s,
                    airline=%s, flight_route=%s, departure_date=%s, return_date=%s,
                    baggage=%s, pnr=%s,
                    hotel_name=%s, room_type=%s, meal_plan=%s,
                    checkin_date=%s, checkout_date=%s, nights=%s,
                    transfer_type=%s, pickup_time=%s, vehicle_type=%s,
                    tour_name=%s, tour_date=%s, tour_included=%s,
                    net_cost=%s, sell_price=%s, profit=%s, notes=%s
                WHERE sale_id=%s AND deleted=FALSE
            """, (
                fields['company'], fields['customer'], fields['package_date'],
                fields['airline'], fields['flight_route'], fields['departure_date'], fields['return_date'],
                fields['baggage'], fields['pnr'],
                fields['hotel_name'], fields['room_type'], fields['meal_plan'],
                fields['checkin_date'], fields['checkout_date'], fields['nights'],
                fields['transfer_type'], fields['pickup_time'], fields['vehicle_type'],
                fields['tour_name'], fields['tour_date'], fields['tour_included'],
                fields['net_cost'], fields['sell_price'], fields['profit'], fields['notes'],
                sale_id,
            ))
            return existing['id']
        else:
            pkg_num = next_doc_number('PKG')
            new_pkg_id = execute_db("""
                INSERT INTO packages
                (package_number, sale_id, created_by_id, created_by,
                 company, customer, package_date,
                 airline, flight_route, departure_date, return_date, baggage, pnr,
                 hotel_name, room_type, meal_plan, checkin_date, checkout_date, nights,
                 transfer_type, pickup_time, vehicle_type,
                 tour_name, tour_date, tour_included,
                 net_cost, sell_price, profit, status, notes)
                VALUES (%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,%s,%s,%s,
                        %s,%s,%s,%s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s,%s,%s)
                RETURNING id
            """, (
                pkg_num, sale_id,
                sale.get('created_by_user_id'), sale.get('created_by_username') or '',
                fields['company'], fields['customer'], fields['package_date'],
                fields['airline'], fields['flight_route'], fields['departure_date'],
                fields['return_date'], fields['baggage'], fields['pnr'],
                fields['hotel_name'], fields['room_type'], fields['meal_plan'],
                fields['checkin_date'], fields['checkout_date'], fields['nights'],
                fields['transfer_type'], fields['pickup_time'], fields['vehicle_type'],
                fields['tour_name'], fields['tour_date'], fields['tour_included'],
                fields['net_cost'], fields['sell_price'], fields['profit'],
                'ACTIVE', fields['notes'],
            ))
            return new_pkg_id
    except Exception as _e:
        logger.error(f"sync_package_from_sale({sale_id}) error: {_e}")
        return None


# ── Ownership guard — employees see only their own sales ─────────────────────
def own_sale_required(f):
    """Allow admins full access; users only access their own sales."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in.', 'warning')
            return redirect(url_for('login'))
        sale_id = kwargs.get('sale_id')
        if sale_id and session.get('user_role') != 'admin':
            s = query_db('SELECT created_by_user_id FROM sales WHERE id=%s AND deleted=FALSE',
                         [sale_id], one=True)
            if not s or s['created_by_user_id'] != session['user_id']:
                flash('Access denied — you can only manage your own transactions.', 'danger')
                return redirect(url_for('my_sales'))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════════════════════
#  EMPLOYEE DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/my')
@login_required
def employee_dashboard():
    if session.get('user_role') == 'admin':
        return redirect(url_for('index'))

    uid = session['user_id']
    today_str = date.today().strftime('%d %B %Y')

    date_from, date_to, quick = resolve_dashboard_dates(request.args)
    drange_u = [uid, date_from, date_to]

    # ── My stats ──────────────────────────────────────────────────────────────
    my_stats = query_db("""
        SELECT COUNT(*) as total_txns,
               COALESCE(SUM(sell),0) as total_sell,
               COALESCE(SUM(profit),0) as total_profit
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND created_by_user_id=%s
          AND sale_date BETWEEN %s AND %s
    """, drange_u, one=True)
    my_stats = dict(my_stats)
    commission_rate = 20.0
    user_row = query_db('SELECT commission_rate FROM users WHERE id=%s', [uid], one=True)
    if user_row and user_row['commission_rate']:
        commission_rate = float(user_row['commission_rate'])
    my_stats['commission'] = round(float(my_stats['total_profit'] or 0) * commission_rate / 100, 3)
    my_stats['commission_rate'] = commission_rate

    # ── Monthly breakdown ─────────────────────────────────────────────────────
    my_monthly = query_db("""
        SELECT to_char(to_date(sale_date,'YYYY-MM-DD'),'MM') as month,
               COALESCE(SUM(sell),0) as total_sell,
               COALESCE(SUM(profit),0) as total_profit,
               COUNT(*) as count
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND created_by_user_id=%s
          AND sale_date BETWEEN %s AND %s
        GROUP BY month ORDER BY month
    """, drange_u)

    # ── Status counts ─────────────────────────────────────────────────────────
    sc = query_db("""
        SELECT
          SUM(CASE WHEN status='DONE'  THEN 1 ELSE 0 END) as done_count,
          SUM(CASE WHEN status='STILL' THEN 1 ELSE 0 END) as still_count,
          SUM(CASE WHEN status NOT IN ('DONE','STILL') THEN 1 ELSE 0 END) as other_count
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE AND created_by_user_id=%s
          AND sale_date BETWEEN %s AND %s
    """, drange_u, one=True)
    if sc:
        my_stats['done_count']  = int(sc['done_count']  or 0)
        my_stats['still_count'] = int(sc['still_count'] or 0)
        my_stats['other_count'] = int(sc['other_count'] or 0)

    # ── Recent transactions ───────────────────────────────────────────────────
    recent = query_db("""
        SELECT * FROM sales
        WHERE deleted=FALSE AND created_by_user_id=%s
          AND sale_date BETWEEN %s AND %s
        ORDER BY created_at DESC LIMIT 10
    """, drange_u)

    # ── Upcoming departures — forward-looking operational list, intentionally
    #    NOT scoped to the historical report date filter (filtering it by a
    #    past range would always empty it out and defeat its purpose). ──────
    upcoming = query_db("""
        SELECT * FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE
          AND created_by_user_id=%s
          AND travel_date >= %s
        ORDER BY travel_date ASC LIMIT 8
    """, [uid, str(date.today())])

    # ── My invoices ───────────────────────────────────────────────────────────
    my_invoices = query_db("""
        SELECT * FROM invoices WHERE deleted=FALSE AND created_by_id=%s
        ORDER BY created_at DESC LIMIT 5
    """, [uid]) or []

    # ── My vouchers ───────────────────────────────────────────────────────────
    my_vouchers = query_db("""
        SELECT * FROM vouchers WHERE deleted=FALSE AND created_by_id=%s
        ORDER BY created_at DESC LIMIT 5
    """, [uid]) or []

    # Chart data
    chart_labels  = [str(r['month']) for r in my_monthly]
    chart_sells   = [float(r['total_sell'] or 0) for r in my_monthly]
    chart_profits = [float(r['total_profit'] or 0) for r in my_monthly]
    has_chart     = any(v > 0 for v in chart_sells)

    # Service type breakdown for this employee
    emp_svc = query_db("""
        SELECT COALESCE(service_type,'FLIGHT') as svc,
               COUNT(*) as cnt,
               COALESCE(SUM(sell),0) as total_sell
        FROM sales
        WHERE deleted=FALSE AND created_by_user_id=%s
          AND sale_date BETWEEN %s AND %s
        GROUP BY svc ORDER BY cnt DESC
    """, drange_u) or []

    return render_template('employee_dashboard.html',
        my_stats=my_stats, my_monthly=my_monthly, recent=recent,
        upcoming=upcoming, my_invoices=my_invoices, my_vouchers=my_vouchers,
        chart_labels=chart_labels, chart_sells=chart_sells,
        chart_profits=chart_profits, has_chart=has_chart,
        today=today_str, commission_rate=commission_rate,
        emp_svc=emp_svc,
        date_from=date_from, date_to=date_to, quick=quick,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  MY SALES (employee view of own transactions)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/my/sales')
@login_required
def my_sales():
    if session.get('user_role') == 'admin':
        return redirect(url_for('sales_report'))
    uid       = session['user_id']
    company   = request.args.get('company', '')
    status    = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    customer  = request.args.get('customer', '').strip().upper()
    page      = max(1, int(request.args.get('page', 1)))

    q = 'SELECT * FROM sales WHERE deleted=FALSE AND created_by_user_id=%s'
    cq = ('SELECT COALESCE(SUM(sell),0) as sell, COALESCE(SUM(profit),0) as profit,'
          ' COUNT(*) as cnt FROM sales WHERE deleted=FALSE AND created_by_user_id=%s')
    params = [uid]
    if company:
        q += ' AND company=%s'; cq += ' AND company=%s'; params.append(company)
    if status:
        q += ' AND status=%s';  cq += ' AND status=%s';  params.append(status)
    if date_from:
        q += ' AND sale_date>=%s'; cq += ' AND sale_date>=%s'; params.append(date_from)
    if date_to:
        q += ' AND sale_date<=%s'; cq += ' AND sale_date<=%s'; params.append(date_to)
    if customer:
        q += ' AND UPPER(customer) LIKE %s'; cq += ' AND UPPER(customer) LIKE %s'
        params.append('%' + customer + '%')
    q += ' ORDER BY sale_date DESC, id DESC'

    agg = query_db(cq, params, one=True)
    totals = {'sell': float(agg['sell'] or 0), 'profit': float(agg['profit'] or 0), 'count': int(agg['cnt'] or 0)}
    sales, total_rows, total_pages = paginate(q, params, page)

    companies_q = query_db("""
        SELECT DISTINCT company FROM sales
        WHERE deleted=FALSE AND created_by_user_id=%s ORDER BY company
    """, [uid])
    companies = [r['company'] for r in companies_q]

    return render_template('my_sales.html',
        sales=sales, totals=totals, companies=companies,
        filters={'company':company,'status':status,'date_from':date_from,'date_to':date_to,'customer':customer},
        page=page, total_pages=total_pages, total_rows=total_rows,
    )


# ── Override add_sale to capture user ─────────────────────────────────────────
# Monkey-patch the existing add_sale to store user info
_orig_add_sale_func = app.view_functions.get('add_sale')

@app.route('/add-v2', methods=['GET', 'POST'])
@login_required
def add_sale_v2():
    """Extended add sale that captures which user created it."""
    try:
        companies = get_companies_list()
    except Exception:
        companies = []
    if request.method == 'POST':
        errors = validate_sale_form(request.form)
        if errors:
            for e in errors: flash(e, 'danger')
            return render_template('add.html', companies=companies,
                                   today=str(date.today()), form=request.form)
        net  = float(request.form.get('net', 0))
        sell = float(request.form.get('sell', 0))
        outbound_delivery = request.form.get('outbound_delivery','').strip()
        return_delivery   = request.form.get('return_delivery','').strip()
        return_date       = request.form.get('return_date','').strip()
        return_supplier   = request.form.get('return_supplier','').upper().strip()
        outbound_status, return_status, overall = compute_ticket_status(
            outbound_delivery, return_delivery)

        service_type = request.form.get('service_type','FLIGHT').upper().strip()
        _json = json
        tour_names    = request.form.getlist('tour_name[]')
        tour_dates    = request.form.getlist('tour_date[]')
        tour_pickups  = request.form.getlist('tour_pickup[]')
        tour_statuses = request.form.getlist('tour_status[]')
        tour_suppliers= request.form.getlist('tour_supplier[]')
        tour_costs    = request.form.getlist('tour_cost[]')
        tour_notes    = request.form.getlist('tour_notes[]')
        tours_list = []
        for i in range(len(tour_names)):
            if tour_names[i].strip():
                tours_list.append({'name':tour_names[i].strip(),'date':tour_dates[i].strip() if i<len(tour_dates) else '','pickup':tour_pickups[i].strip() if i<len(tour_pickups) else '','status':tour_statuses[i].strip() if i<len(tour_statuses) else 'INCLUDED','supplier':tour_suppliers[i].strip() if i<len(tour_suppliers) else '','cost':float(tour_costs[i]) if i<len(tour_costs) and tour_costs[i] else 0,'notes':tour_notes[i].strip() if i<len(tour_notes) else ''})
        _pax_n2=request.form.getlist('pax_name[]'); _pax_p2=request.form.getlist('pax_passport[]'); _pax_nat2=request.form.getlist('pax_nationality[]'); _pax_d2=request.form.getlist('pax_dob[]')
        ev = dict(service_type=service_type,hotel_supplier=request.form.get('hotel_supplier','').strip(),hotel_name=request.form.get('hotel_name','').strip(),hotel_room=request.form.get('hotel_room','').strip(),hotel_meal=request.form.get('hotel_meal','').strip(),hotel_checkin=request.form.get('hotel_checkin','').strip(),hotel_checkout=request.form.get('hotel_checkout','').strip(),hotel_nights=int(request.form.get('hotel_nights',0) or 0),hotel_net=float(request.form.get('hotel_net',0) or 0),transfer_supplier=request.form.get('transfer_supplier','').strip(),transfer_type=request.form.get('transfer_type','').strip(),transfer_pickup=request.form.get('transfer_pickup','').strip(),transfer_vehicle=request.form.get('transfer_vehicle','').strip(),transfer_net=float(request.form.get('transfer_net',0) or 0),tours_json=_json.dumps(tours_list),visa_supplier=request.form.get('visa_supplier','').strip(),visa_type=request.form.get('visa_type','').strip(),passport_number=request.form.get('passport_number','').upper().strip(),visa_status=request.form.get('visa_status','').strip(),insurance_supplier=request.form.get('insurance_supplier','').strip(),insurance_type=request.form.get('insurance_type','').strip(),airline=request.form.get('airline','').upper().strip(),pnr=request.form.get('pnr','').upper().strip(),baggage=request.form.get('baggage','').strip(),passengers_json=_json.dumps([{'name':_pax_n2[i].strip(),'passport':_pax_p2[i].strip() if i<len(_pax_p2) else '','nationality':_pax_nat2[i].strip() if i<len(_pax_nat2) else '','dob':_pax_d2[i].strip() if i<len(_pax_d2) else ''} for i in range(len(_pax_n2)) if _pax_n2[i].strip()]))
        _oc   = float(request.form.get('outbound_cost',0) or 0)
        _rc   = float(request.form.get('return_cost',0) or 0)
        _tkn  = request.form.get('ticket_number','').strip()
        _rtn  = request.form.get('return_ticket_number','').strip()
        _rpnr = request.form.get('return_pnr','').upper().strip()
        _ctype_v2 = request.form.get('customer_type', 'COMPANY').upper().strip() or 'COMPANY'
        _dphone   = request.form.get('direct_phone', '').strip()
        _dnat     = request.form.get('direct_nationality', '').strip()
        _demail   = request.form.get('direct_email', '').strip()
        new_id = execute_db("""
            INSERT INTO sales
            (from_loc,to_loc,via,trip_type,buy_from,company,tickets,
             customer,sale_date,travel_date,return_date,return_supplier,
             outbound_delivery,return_delivery,outbound_status,return_status,
             net,sell,profit,status,remarks,created_by_user_id,created_by_username,
             service_type,hotel_supplier,hotel_name,hotel_room,hotel_meal,
             hotel_checkin,hotel_checkout,hotel_nights,hotel_net,
             transfer_supplier,transfer_type,transfer_pickup,transfer_vehicle,transfer_net,
             tours_json,visa_supplier,visa_type,passport_number,visa_status,
             insurance_supplier,insurance_type,airline,pnr,baggage,passengers_json,
             outbound_cost,return_cost,ticket_number,return_ticket_number,return_pnr,
             customer_type,direct_phone,direct_nationality,direct_email)
            VALUES (
                %s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s
            ) RETURNING id
        """, (
            (request.form.get('from_loc','').upper().strip() or '-'),  # 1  from_loc
            (request.form.get('to_loc','').upper().strip() or '-'),    # 2  to_loc
            request.form.get('via','').upper().strip(),                  # 3  via
            request.form.get('trip_type',''),                           # 4  trip_type
            request.form.get('buy_from','').upper().strip(),             # 5  buy_from
            request.form.get('company','').upper().strip(),              # 6  company
            int(request.form.get('tickets',1) or 1),                     # 7  tickets
            request.form.get('customer','').upper().strip(),             # 8  customer
            request.form.get('sale_date', str(date.today())),             # 9  sale_date
            request.form.get('travel_date','').strip(),                  # 10 travel_date
            return_date,                                                   # 11 return_date
            return_supplier,                                               # 12 return_supplier
            outbound_delivery,                                             # 13 outbound_delivery
            return_delivery,                                               # 14 return_delivery
            outbound_status,                                               # 15 outbound_status
            return_status,                                                 # 16 return_status
            net,                                                           # 17 net
            sell,                                                          # 18 sell
            sell - net,                                                    # 19 profit
            overall,                                                       # 20 status
            request.form.get('remarks','').strip(),                      # 21 remarks
            session.get('user_id'),                                       # 22 created_by_user_id
            session.get('username',''),                                  # 23 created_by_username
            ev['service_type'],                                           # 24 service_type
            ev['hotel_supplier'],                                         # 25 hotel_supplier
            ev['hotel_name'],                                             # 26 hotel_name
            ev['hotel_room'],                                             # 27 hotel_room
            ev['hotel_meal'],                                             # 28 hotel_meal
            ev['hotel_checkin'],                                          # 29 hotel_checkin
            ev['hotel_checkout'],                                         # 30 hotel_checkout
            ev['hotel_nights'],                                           # 31 hotel_nights
            ev['hotel_net'],                                              # 32 hotel_net
            ev['transfer_supplier'],                                      # 33 transfer_supplier
            ev['transfer_type'],                                          # 34 transfer_type
            ev['transfer_pickup'],                                        # 35 transfer_pickup
            ev['transfer_vehicle'],                                       # 36 transfer_vehicle
            ev['transfer_net'],                                           # 37 transfer_net
            ev['tours_json'],                                             # 38 tours_json
            ev['visa_supplier'],                                          # 39 visa_supplier
            ev['visa_type'],                                              # 40 visa_type
            ev['passport_number'],                                        # 41 passport_number
            ev['visa_status'],                                            # 42 visa_status
            ev['insurance_supplier'],                                     # 43 insurance_supplier
            ev['insurance_type'],                                         # 44 insurance_type
            ev['airline'],                                                # 45 airline
            ev['pnr'],                                                    # 46 pnr
            ev['baggage'],                                                # 47 baggage
            ev['passengers_json'],                                         # 48 passengers_json
            _oc,                                                           # 49 outbound_cost
            _rc,                                                           # 50 return_cost
            _tkn,                                                          # 51 ticket_number
            _rtn,                                                          # 52 return_ticket_number
            _rpnr,                                                         # 53 return_pnr
            _ctype_v2,                                                     # 54 customer_type
            _dphone,                                                       # 55 direct_phone
            _dnat,                                                         # 56 direct_nationality
            _demail,                                                       # 57 direct_email
        ))
        # Assign transaction number
        if new_id:
            _txn = next_txn_number()
            try:
                execute_db('UPDATE sales SET txn_number=%s WHERE id=%s', (_txn, new_id))
            except Exception:
                pass
        try:
            log_action('CREATE','sales',new_id,
                       f"{request.form.get('customer','').upper()} | Sell:{sell}")
        except Exception:
            pass
        # ── Shadow ledger + sale_items ────────────────────────────────────────
        if new_id:
            _co  = request.form.get('company','').upper().strip()
            _svc = service_type
            _fl  = request.form.get('from_loc','') or ''
            _tl  = request.form.get('to_loc','')   or ''
            _cust= request.form.get('customer','')
            _dt  = request.form.get('sale_date', str(date.today()))
            _ledger_entity = _cust.upper().strip() if _ctype_v2 == 'DIRECT' else _co
            post_ledger('SALE', new_id, 'sales', _ledger_entity,
                        f"{_svc} {_fl}{'→'+_tl if _tl else ''} | {_cust}",
                        debit=sell, credit=0, entry_date=_dt)
            post_sale_items(new_id, _svc,
                            request.form.get('buy_from',''),
                            f"{_svc} {_fl}{'→'+_tl if _tl else ''}",
                            _oc, sell,
                            request.form.get('tickets',1))
        # Sync PACKAGE transactions to the packages table
        if new_id and service_type == 'PACKAGE':
            try:
                sync_package_from_sale(new_id)
            except Exception:
                pass
        flash('Transaction saved successfully.', 'success')
        # POST → REDIRECT → GET: prevents duplicate submission on back/refresh
        if session.get('user_role') != 'admin':
            return redirect(url_for('my_sales'))
        return redirect(url_for('sales_report'))
    return render_template('add.html', companies=companies, today=str(date.today()), form={})


@app.route('/my/edit/<int:sale_id>', methods=['GET', 'POST'])
@login_required
@own_sale_required
def my_edit_sale(sale_id):
    sale = query_db('SELECT * FROM sales WHERE id=%s AND deleted=FALSE', [sale_id], one=True)
    if not sale:
        flash('Transaction not found.', 'danger')
        return redirect(url_for('my_sales'))
    sale = safe_sale(sale)
    companies = get_companies_list()
    if request.method == 'POST':
        errors = validate_sale_form(request.form)
        if errors:
            for e in errors: flash(e, 'danger')
            return render_template('add.html', sale=sale, companies=companies, edit=True)
        net  = float(request.form.get('net', 0))
        sell = float(request.form.get('sell', 0))
        outbound_delivery = request.form.get('outbound_delivery','').strip()
        return_delivery   = request.form.get('return_delivery','').strip()
        return_date       = request.form.get('return_date','').strip()
        return_supplier   = request.form.get('return_supplier','').upper().strip()
        outbound_status, return_status, overall = compute_ticket_status(
            outbound_delivery, return_delivery)

        service_type = request.form.get('service_type','FLIGHT').upper().strip()
        tour_names=request.form.getlist('tour_name[]'); tour_dates=request.form.getlist('tour_date[]'); tour_pickups=request.form.getlist('tour_pickup[]'); tour_statuses=request.form.getlist('tour_status[]'); tour_suppliers=request.form.getlist('tour_supplier[]'); tour_costs=request.form.getlist('tour_cost[]'); tour_notes=request.form.getlist('tour_notes[]')
        _json = json
        tours_list=[{'name':tour_names[i].strip(),'date':tour_dates[i].strip() if i<len(tour_dates) else '','pickup':tour_pickups[i].strip() if i<len(tour_pickups) else '','status':tour_statuses[i].strip() if i<len(tour_statuses) else 'INCLUDED','supplier':tour_suppliers[i].strip() if i<len(tour_suppliers) else '','cost':float(tour_costs[i]) if i<len(tour_costs) and tour_costs[i] else 0,'notes':tour_notes[i].strip() if i<len(tour_notes) else ''} for i in range(len(tour_names)) if tour_names[i].strip()]
        _pax_n2=request.form.getlist('pax_name[]'); _pax_p2=request.form.getlist('pax_passport[]'); _pax_nat2=request.form.getlist('pax_nationality[]'); _pax_d2=request.form.getlist('pax_dob[]')
        ev=dict(service_type=service_type,hotel_supplier=request.form.get('hotel_supplier','').strip(),hotel_name=request.form.get('hotel_name','').strip(),hotel_room=request.form.get('hotel_room','').strip(),hotel_meal=request.form.get('hotel_meal','').strip(),hotel_checkin=request.form.get('hotel_checkin','').strip(),hotel_checkout=request.form.get('hotel_checkout','').strip(),hotel_nights=int(request.form.get('hotel_nights',0) or 0),hotel_net=float(request.form.get('hotel_net',0) or 0),transfer_supplier=request.form.get('transfer_supplier','').strip(),transfer_type=request.form.get('transfer_type','').strip(),transfer_pickup=request.form.get('transfer_pickup','').strip(),transfer_vehicle=request.form.get('transfer_vehicle','').strip(),transfer_net=float(request.form.get('transfer_net',0) or 0),tours_json=_json.dumps(tours_list),visa_supplier=request.form.get('visa_supplier','').strip(),visa_type=request.form.get('visa_type','').strip(),passport_number=request.form.get('passport_number','').upper().strip(),visa_status=request.form.get('visa_status','').strip(),insurance_supplier=request.form.get('insurance_supplier','').strip(),insurance_type=request.form.get('insurance_type','').strip(),airline=request.form.get('airline','').upper().strip(),pnr=request.form.get('pnr','').upper().strip(),baggage=request.form.get('baggage','').strip(),passengers_json=_json.dumps([{'name':_pax_n2[i].strip(),'passport':_pax_p2[i].strip() if i<len(_pax_p2) else '','nationality':_pax_nat2[i].strip() if i<len(_pax_nat2) else '','dob':_pax_d2[i].strip() if i<len(_pax_d2) else ''} for i in range(len(_pax_n2)) if _pax_n2[i].strip()]),return_pnr=request.form.get('return_pnr','').upper().strip())
        # ACC-01 / ACC-04 — Snapshot old values BEFORE update for ledger reversal
        _old2 = query_db(
            'SELECT sell, company, sale_date FROM sales WHERE id=%s AND deleted=FALSE',
            [sale_id], one=True
        )
        _old2_sell    = float(_old2['sell']    if _old2 else 0)
        _old2_company = str(_old2['company']   if _old2 else '').upper().strip()
        _old2_date    = str(_old2['sale_date'] if _old2 else str(date.today()))

        _ctype_my = request.form.get('customer_type', 'COMPANY').upper().strip() or 'COMPANY'
        execute_db('''
            UPDATE sales SET
                from_loc=%s,to_loc=%s,via=%s,trip_type=%s,buy_from=%s,
                company=%s,tickets=%s,customer=%s,sale_date=%s,travel_date=%s,
                return_date=%s,return_supplier=%s,
                outbound_delivery=%s,return_delivery=%s,
                outbound_status=%s,return_status=%s,
                net=%s,sell=%s,profit=%s,status=%s,remarks=%s,
                service_type=%s,hotel_supplier=%s,hotel_name=%s,hotel_room=%s,hotel_meal=%s,
                hotel_checkin=%s,hotel_checkout=%s,hotel_nights=%s,hotel_net=%s,
                transfer_supplier=%s,transfer_type=%s,transfer_pickup=%s,transfer_vehicle=%s,transfer_net=%s,
                tours_json=%s,visa_supplier=%s,visa_type=%s,passport_number=%s,visa_status=%s,
                insurance_supplier=%s,insurance_type=%s,airline=%s,pnr=%s,baggage=%s,passengers_json=%s,
                ticket_number=%s,return_ticket_number=%s,return_pnr=%s,
                customer_type=%s,direct_phone=%s,direct_nationality=%s,direct_email=%s
            WHERE id=%s
        ''', (
            request.form.get('from_loc','').upper().strip(),request.form.get('to_loc','').upper().strip(),
            request.form.get('via','').upper().strip(),request.form.get('trip_type',''),
            request.form.get('buy_from','').upper().strip(),request.form.get('company','').upper().strip(),
            int(request.form.get('tickets',1)),request.form.get('customer','').upper().strip(),
            request.form.get('sale_date',''),request.form.get('travel_date','').strip(),
            return_date,return_supplier,outbound_delivery,return_delivery,
            outbound_status,return_status,net,sell,sell-net,overall,
            request.form.get('remarks','').strip(),
            ev['service_type'],ev['hotel_supplier'],ev['hotel_name'],ev['hotel_room'],ev['hotel_meal'],
            ev['hotel_checkin'],ev['hotel_checkout'],ev['hotel_nights'],ev['hotel_net'],
            ev['transfer_supplier'],ev['transfer_type'],ev['transfer_pickup'],ev['transfer_vehicle'],ev['transfer_net'],
            ev['tours_json'],ev['visa_supplier'],ev['visa_type'],ev['passport_number'],ev['visa_status'],
            ev['insurance_supplier'],ev['insurance_type'],ev['airline'],ev['pnr'],ev['baggage'],ev['passengers_json'],
            request.form.get('ticket_number','').strip(),
            request.form.get('return_ticket_number','').strip(),
            ev['return_pnr'],
            _ctype_my,
            request.form.get('direct_phone','').strip(),
            request.form.get('direct_nationality','').strip(),
            request.form.get('direct_email','').strip(),
            sale_id
        ))

        # ACC-01 / ACC-04 — Reverse original ledger entry, post corrective entry
        _orig2 = query_db(
            """SELECT id FROM accounting_entries
               WHERE ref_type='SALE' AND ref_id=%s AND ref_table='sales'
                 AND is_reversed=FALSE
               ORDER BY id DESC LIMIT 1""",
            [sale_id], one=True
        )
        _new2_company  = request.form.get('company', _old2_company).upper().strip()
        _new2_date     = request.form.get('sale_date', _old2_date)
        _new2_customer = request.form.get('customer', '').upper().strip()
        _fl2 = request.form.get('from_loc','') or ''
        _tl2 = request.form.get('to_loc','')   or ''
        _ledger_ent_my = _new2_customer if _ctype_my == 'DIRECT' else _new2_company
        if _orig2:
            execute_db(
                "UPDATE accounting_entries SET is_reversed=TRUE WHERE id=%s",
                [_orig2['id']]
            )
            post_ledger(
                'SALE_REVERSAL', sale_id, 'sales', _old2_company or _new2_customer,
                f"Reversal of SALE-{sale_id:04d} (edit by {session.get('username','')})",
                debit=0, credit=_old2_sell, entry_date=_old2_date
            )
        post_ledger(
            'SALE', sale_id, 'sales', _ledger_ent_my,
            f"{ev['service_type']} {_fl2}{'→'+_tl2 if _tl2 else ''} | {_new2_customer} (corrected)",
            debit=sell, credit=0, entry_date=_new2_date
        )

        log_action('UPDATE','sales',sale_id,
                   f"{_new2_customer} | {ev['service_type']} | OldSell:{_old2_sell} NewSell:{sell}")
        flash('Transaction updated.','success')
        return redirect(url_for('my_sales'))
    return render_template('add.html', sale=sale, companies=companies, edit=True)


# ══════════════════════════════════════════════════════════════════════════════
#  INVOICES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/invoices')
@login_required
def invoice_list():
    uid = session['user_id']
    is_admin = session.get('user_role') == 'admin'
    status    = request.args.get('status', '')
    company   = request.args.get('company', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    page      = max(1, int(request.args.get('page', 1)))

    q = 'SELECT * FROM invoices WHERE deleted=FALSE'
    if not is_admin:
        q += ' AND created_by_id=%s'
        params = [uid]
    else:
        params = []
    if status:  q += ' AND status=%s';   params.append(status)
    if company: q += ' AND company ILIKE %s'; params.append(f'%{company}%')
    if date_from: q += ' AND invoice_date>=%s'; params.append(date_from)
    if date_to:   q += ' AND invoice_date<=%s'; params.append(date_to)
    q += ' ORDER BY created_at DESC'

    invoices, total_rows, total_pages = paginate(q, params, page)
    total_amount = sum(float(i['total'] or 0) for i in invoices)

    return render_template('invoices.html',
        invoices=invoices, total_amount=total_amount,
        filters={'status':status,'company':company,'date_from':date_from,'date_to':date_to},
        page=page, total_pages=total_pages, total_rows=total_rows,
    )


@app.route('/invoices/new', methods=['GET', 'POST'])
@login_required
def new_invoice():
    sale_id   = request.args.get('sale_id', '')
    sale      = query_db('SELECT * FROM sales WHERE id=%s AND deleted=FALSE', [sale_id], one=True) if sale_id else None
    companies = get_companies_list()
    if request.method == 'POST':
        try:
            amount   = float(request.form.get('amount') or 0)
            discount = float(request.form.get('discount') or 0)
            total    = round(amount - discount, 3)
            inv_num  = next_doc_number('INV')
            # sale_id: convert empty string to None for FK constraint
            raw_sid  = request.form.get('sale_id','').strip()
            sid      = int(raw_sid) if raw_sid and raw_sid.isdigit() else None
            # due_date: convert empty string to None
            due_date_val = request.form.get('due_date','').strip() or None

            new_id = execute_db("""
                INSERT INTO invoices
                (invoice_number,sale_id,created_by_id,created_by,company,customer,
                 service_desc,amount,discount,total,status,invoice_date,due_date,notes)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (
                inv_num, sid,
                session['user_id'], session.get('username',''),
                request.form.get('company','').upper().strip(),
                request.form.get('customer','').upper().strip(),
                request.form.get('service_desc','').strip(),
                amount, discount, total,
                request.form.get('status','UNPAID'),
                request.form.get('invoice_date', str(date.today())),
                due_date_val,
                request.form.get('notes','').strip()
            ))
            log_action('CREATE', 'invoices', new_id or 0, f"Invoice {inv_num} | {total} JOD")
            flash(f'Invoice {inv_num} created successfully.', 'success')
            if new_id:
                return redirect(url_for('view_invoice', inv_id=new_id))
            else:
                return redirect(url_for('invoice_list'))
        except Exception as ex:
            logger.error(f"new_invoice POST error: {ex}", exc_info=True)
            flash(f'Error creating invoice: {ex}', 'danger')
    # Load transactions for pre-selected company (from sale or URL param)
    company_filter = request.args.get('company', '')
    if sale and sale['company']:
        company_filter = sale['company']
    company_transactions = []
    if company_filter:
        company_transactions = query_db("""
            SELECT id, sale_date, customer, from_loc, to_loc, sell, status
            FROM sales
            WHERE deleted=FALSE AND is_archived=FALSE
              AND company=%s
            ORDER BY sale_date DESC
            LIMIT 50
        """, [company_filter]) or []

    # Auto-build service description from sale details
    auto_desc = ''
    if sale:
        svc = (sale.get('service_type') or 'FLIGHT').upper()
        parts = []
        if svc in ('FLIGHT', 'PACKAGE'):
            fl = f"✈ Flight: {sale.get('from_loc','')} → {sale.get('to_loc','')} | {sale.get('airline','') or ''}"
            if sale.get('pnr'):    fl += f" | PNR: {sale['pnr']}"
            if sale.get('travel_date'): fl += f" | Dep: {sale['travel_date']}"
            if sale.get('return_date'): fl += f" | Ret: {sale['return_date']}"
            if sale.get('baggage'):     fl += f" | Bag: {sale['baggage']}"
            parts.append(fl.strip(' |'))
        if svc in ('HOTEL', 'PACKAGE') and sale.get('hotel_name'):
            ht = f"🏨 Hotel: {sale['hotel_name']}"
            if sale.get('hotel_room'):     ht += f" | {sale['hotel_room']}"
            if sale.get('hotel_meal'):     ht += f" | {sale['hotel_meal']}"
            if sale.get('hotel_checkin'):  ht += f" | CI: {sale['hotel_checkin']}"
            if sale.get('hotel_checkout'): ht += f" | CO: {sale['hotel_checkout']}"
            if sale.get('hotel_nights'):   ht += f" | {sale['hotel_nights']} nights"
            parts.append(ht)
        if svc in ('TRANSFER', 'PACKAGE') and sale.get('transfer_type'):
            tr = f"🚗 Transfer: {sale['transfer_type']}"
            if sale.get('transfer_vehicle'): tr += f" | {sale['transfer_vehicle']}"
            if sale.get('transfer_pickup'):  tr += f" | {sale['transfer_pickup']}"
            parts.append(tr)
        if svc == 'VISA':
            parts.append(f"📋 Visa: {sale.get('visa_type','')} | Passport: {sale.get('passport_number','')}")
        if svc == 'INSURANCE':
            parts.append(f"🛡 Insurance: {sale.get('insurance_type','')}")
        # Passengers
        import json as _j
        pax = []
        try: pax = _j.loads(sale.get('passengers_json') or '[]')
        except: pass
        if pax:
            names = ', '.join(p.get('name','') for p in pax if p.get('name'))
            if names: parts.append(f"👥 Passengers: {names}")
        # Tours
        tours = []
        try: tours = _j.loads(sale.get('tours_json') or '[]')
        except: pass
        if tours:
            tnames = ', '.join(t.get('name','') for t in tours if t.get('name'))
            if tnames: parts.append(f"🗺 Tours: {tnames}")
        if svc == 'FLIGHT' and not parts:
            parts.append(f"Flight {sale.get('from_loc','')} → {sale.get('to_loc','')} | {sale.get('customer','')}")
        auto_desc = '\n'.join(parts) if parts else f"{svc} — {sale.get('customer','')}"

    return render_template('invoice_form.html',
        sale=sale, companies=companies, today=str(date.today()),
        company_filter=company_filter,
        company_transactions=company_transactions,
        auto_desc=auto_desc)


@app.route('/invoices/<int:inv_id>')
@login_required
def view_invoice(inv_id):
    inv = query_db('SELECT * FROM invoices WHERE id=%s AND deleted=FALSE', [inv_id], one=True)
    if not inv:
        flash('Invoice not found.', 'danger')
        return redirect(url_for('invoice_list'))
    if session.get('user_role') != 'admin' and inv['created_by_id'] != session['user_id']:
        flash('Access denied.', 'danger')
        return redirect(url_for('invoice_list'))
    sale = query_db('SELECT * FROM sales WHERE id=%s', [inv['sale_id']], one=True) if inv['sale_id'] else None
    return render_template('invoice_view.html', inv=inv, sale=sale)


@app.route('/invoices/<int:inv_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_invoice(inv_id):
    inv = query_db('SELECT * FROM invoices WHERE id=%s AND deleted=FALSE', [inv_id], one=True)
    if not inv:
        flash('Invoice not found.', 'danger')
        return redirect(url_for('invoice_list'))
    if session.get('user_role') != 'admin' and inv['created_by_id'] != session['user_id']:
        flash('Access denied.', 'danger')
        return redirect(url_for('invoice_list'))
    companies = get_companies_list()
    if request.method == 'POST':
        amount   = float(request.form.get('amount', 0))
        discount = float(request.form.get('discount', 0))
        total    = round(amount - discount, 3)
        execute_db("""
            UPDATE invoices SET
                company=%s,customer=%s,service_desc=%s,amount=%s,discount=%s,total=%s,
                status=%s,invoice_date=%s,due_date=%s,notes=%s
            WHERE id=%s
        """, (
            request.form.get('company','').upper().strip(),
            request.form.get('customer','').upper().strip(),
            request.form.get('service_desc','').strip(),
            amount, discount, total,
            request.form.get('status','UNPAID'),
            request.form.get('invoice_date', str(date.today())),
            request.form.get('due_date',''),
            request.form.get('notes','').strip(), inv_id
        ))
        log_action('UPDATE', 'invoices', inv_id, f"Total:{total}")
        flash('Invoice updated.', 'success')
        return redirect(url_for('view_invoice', inv_id=inv_id))
    return render_template('invoice_form.html',
        inv=inv, companies=companies, today=str(date.today()))


@app.route('/invoices/<int:inv_id>/delete', methods=['POST'])
@login_required
def delete_invoice(inv_id):
    inv = query_db('SELECT * FROM invoices WHERE id=%s', [inv_id], one=True)
    if inv and (session.get('user_role') == 'admin' or inv['created_by_id'] == session['user_id']):
        execute_db('UPDATE invoices SET deleted=TRUE WHERE id=%s', [inv_id])
        log_action('DELETE', 'invoices', inv_id, f"Invoice {inv['invoice_number']}")
        flash('Invoice deleted.', 'success')
    return redirect(url_for('invoice_list'))





# ══════════════════════════════════════════════════════════════════════════════
#  AUTO-GENERATE VOUCHER FROM PACKAGE TRANSACTION
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/vouchers/from-sale/<int:sale_id>', methods=['GET', 'POST'])
@login_required
def auto_generate_voucher(sale_id):
    """One-click: create a full voucher from a PACKAGE transaction — no re-entry."""
    sale = query_db('SELECT * FROM sales WHERE id=%s AND deleted=FALSE', [sale_id], one=True)
    if not sale:
        flash('Transaction not found.', 'danger')
        return redirect(url_for('voucher_list'))

    if session.get('user_role') != 'admin' and sale['created_by_user_id'] != session['user_id']:
        flash('Access denied.', 'danger')
        return redirect(url_for('voucher_list'))

    vcn = next_doc_number('VCH')

    # Try with extended columns first; fall back to base columns if migration pending
    try:
        new_id = execute_db("""
            INSERT INTO vouchers
            (voucher_number,sale_id,created_by_id,created_by,
             guest_name,num_guests,hotel_name,hotel_address,hotel_phone,hotel_contact,
             room_type,meal_plan,checkin_date,checkout_date,nights,
             include_transfer,include_tours,include_insurance,
             arrival_flight,pickup_sign,driver_contact,pickup_time,vehicle_type,
             emergency_contact,cancellation_policy,remarks,
             passengers_json,tours_json,airline,pnr,baggage,from_loc,to_loc,departure_date,return_date)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            vcn, sale_id,
            session['user_id'], session.get('username',''),
            sale['customer'] or '',
            sale['tickets'] or 1,
            sale.get('hotel_name','') or '',
            '',
            '',
            sale.get('hotel_supplier','') or '',
            sale.get('hotel_room','') or '',
            sale.get('hotel_meal','') or '',
            sale.get('hotel_checkin','') or '',
            sale.get('hotel_checkout','') or '',
            int(sale.get('hotel_nights') or 0),
            bool(sale.get('transfer_type')),
            bool(sale.get('tours_json') and sale.get('tours_json') != '[]'),
            False,
            sale.get('airline','') or '',
            sale['customer'] or '',
            sale.get('transfer_supplier','') or '',
            sale.get('transfer_pickup','') or '',
            sale.get('transfer_vehicle','') or '',
            '',
            '',
            sale.get('remarks','') or '',
            sale.get('passengers_json','[]') or '[]',
            sale.get('tours_json','[]') or '[]',
            sale.get('airline','') or '',
            sale.get('pnr','') or '',
            sale.get('baggage','') or '',
            sale.get('from_loc','') or '',
            sale.get('to_loc','') or '',
            sale.get('travel_date','') or '',
            sale.get('return_date','') or '',
        ))
    except Exception as _ve:
        logger.warning(f"Extended voucher INSERT failed ({_ve}), using base columns")
        # Rollback and retry with only the original base columns
        try: get_db().rollback()
        except: pass
        new_id = execute_db("""
            INSERT INTO vouchers
            (voucher_number,sale_id,created_by_id,created_by,
             guest_name,num_guests,hotel_name,hotel_address,hotel_phone,hotel_contact,
             room_type,meal_plan,checkin_date,checkout_date,nights,
             include_transfer,include_tours,include_insurance,
             arrival_flight,pickup_sign,driver_contact,pickup_time,vehicle_type,
             emergency_contact,cancellation_policy,remarks)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            vcn, sale_id,
            session['user_id'], session.get('username',''),
            sale['customer'] or '',
            sale['tickets'] or 1,
            sale.get('hotel_name','') or '',
            '',
            '',
            sale.get('hotel_supplier','') or '',
            sale.get('hotel_room','') or '',
            sale.get('hotel_meal','') or '',
            sale.get('hotel_checkin','') or '',
            sale.get('hotel_checkout','') or '',
            int(sale.get('hotel_nights') or 0),
            bool(sale.get('transfer_type')),
            bool(sale.get('tours_json') and sale.get('tours_json') != '[]'),
            False,
            sale.get('airline','') or '',
            sale['customer'] or '',
            sale.get('transfer_supplier','') or '',
            sale.get('transfer_pickup','') or '',
            sale.get('transfer_vehicle','') or '',
            '',
            '',
            sale.get('remarks','') or '',
        ))

    log_action('CREATE', 'vouchers', new_id or 0, f"Auto-voucher {vcn} from sale #{sale_id}")
    flash(f'Voucher {vcn} generated automatically! ✓', 'success')
    return redirect(url_for('view_voucher', vch_id=new_id))


# ══════════════════════════════════════════════════════════════════════════════
#  PAYMENT RECEIPT PRINT PAGE
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/payments/<int:pay_id>/receipt')
@login_required
def payment_receipt(pay_id):
    pay = query_db("""
        SELECT p.*, u.username AS created_by_name, u.full_name AS created_by_full
        FROM payments p
        LEFT JOIN users u ON u.id = p.created_by
        WHERE p.id=%s AND p.deleted=FALSE
    """, [pay_id], one=True)
    if not pay:
        flash('Payment not found.', 'danger')
        return redirect(url_for('payments'))
    # Get company balance
    total_sell = query_db(
        'SELECT COALESCE(SUM(sell),0) as s FROM sales WHERE deleted=FALSE AND is_archived=FALSE AND company=%s',
        [pay['company']], one=True
    )
    total_paid = query_db(
        'SELECT COALESCE(SUM(amount),0) as s FROM payments WHERE deleted=FALSE AND is_archived=FALSE AND COALESCE(voided,FALSE)=FALSE AND company=%s',
        [pay['company']], one=True
    )
    balance = float(total_sell['s'] or 0) - float(total_paid['s'] or 0)
    return render_template('payment_receipt.html',
        pay=pay,
        balance=balance,
        total_invoiced=float(total_sell['s'] or 0),
        total_paid_all=float(total_paid['s'] or 0),
        today=date.today().strftime('%d %B %Y'),
    )


# ══════════════════════════════════════════════════════════════════════════════
#  SUPPLIER STATEMENTS
# ══════════════════════════════════════════════════════════════════════════════

def _get_supplier_list():
    """Return a sorted list of unique supplier name strings."""
    try:
        rows = query_db("""
            SELECT DISTINCT supplier FROM (
                SELECT NULLIF(TRIM(buy_from),'')         AS supplier FROM sales WHERE deleted=FALSE AND buy_from IS NOT NULL AND buy_from <> ''
                UNION
                SELECT NULLIF(TRIM(return_supplier),'')               FROM sales WHERE deleted=FALSE AND return_supplier IS NOT NULL AND return_supplier <> ''
                UNION
                SELECT NULLIF(TRIM(hotel_supplier),'')                FROM sales WHERE deleted=FALSE AND hotel_supplier IS NOT NULL AND hotel_supplier <> ''
                UNION
                SELECT NULLIF(TRIM(transfer_supplier),'')             FROM sales WHERE deleted=FALSE AND transfer_supplier IS NOT NULL AND transfer_supplier <> ''
                UNION
                SELECT NULLIF(TRIM(visa_supplier),'')                 FROM sales WHERE deleted=FALSE AND visa_supplier IS NOT NULL AND visa_supplier <> ''
                UNION
                SELECT NULLIF(TRIM(insurance_supplier),'')            FROM sales WHERE deleted=FALSE AND insurance_supplier IS NOT NULL AND insurance_supplier <> ''
                UNION
                SELECT NULLIF(TRIM(company),'')                       FROM sales WHERE deleted=FALSE AND company IS NOT NULL AND company <> ''
                UNION
                SELECT NULLIF(TRIM(supplier),'')                      FROM supplier_payments WHERE deleted=FALSE AND supplier IS NOT NULL AND supplier <> ''
            ) t WHERE supplier IS NOT NULL
            ORDER BY supplier
        """) or []
        return [r['supplier'] for r in rows if r.get('supplier')]
    except Exception:
        return []



# ── Refunds ───────────────────────────────────────────────────────────────────
@app.route('/refunds', methods=['GET','POST'])
@admin_required
def refunds():
    if request.method == 'POST':
        action = request.form.get('action','create')

        if action == 'create':
            sale_id     = request.form.get('sale_id') or None
            company     = request.form.get('company','').upper().strip()
            customer    = request.form.get('customer','').upper().strip()
            refund_type = request.form.get('refund_type','FULL')
            gross_amt   = round(float(request.form.get('gross_amount',0) or 0), 3)
            cancel_fee  = round(float(request.form.get('cancellation_fee',0) or 0), 3)
            refund_amt  = round(float(request.form.get('refund_amount',0) or 0), 3)
            sup_penalty = round(float(request.form.get('supplier_penalty',0) or 0), 3)
            sup_refund  = round(float(request.form.get('supplier_refund',0) or 0), 3)
            reason      = request.form.get('reason','').strip()
            refund_date = request.form.get('refund_date', str(date.today()))

            if not company:
                flash('Company is required.', 'danger')
                return redirect(url_for('refunds'))
            if gross_amt <= 0:
                flash('Refund gross amount must be greater than zero.', 'danger')
                return redirect(url_for('refunds'))
            if refund_amt < 0 or cancel_fee < 0:
                flash('Refund amount and cancellation fee cannot be negative.', 'danger')
                return redirect(url_for('refunds'))
            # Validate refund does not exceed original sale amount
            if sale_id:
                _orig_sale = query_db(
                    'SELECT sell FROM sales WHERE id=%s AND deleted=FALSE', [sale_id], one=True
                )
                if _orig_sale and gross_amt > float(_orig_sale['sell']) * 1.001:
                    flash(
                        f"Refund gross amount ({gross_amt:.3f} JOD) exceeds the "
                        f"original sale amount ({float(_orig_sale['sell']):.3f} JOD).",
                        'danger'
                    )
                    return redirect(url_for('refunds'))

            rid = execute_db("""
                INSERT INTO refunds
                (sale_id, refund_date, refund_type, company, customer,
                 gross_amount, cancellation_fee, refund_amount,
                 supplier_penalty, supplier_refund, reason, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (sale_id, refund_date, refund_type, company, customer,
                  gross_amt, cancel_fee, refund_amt,
                  sup_penalty, sup_refund, reason, session.get('user_id')))

            # Post to ledger: reverse the original sale credit
            if rid:
                post_ledger('REFUND', rid, 'refunds', company,
                            f"Refund {refund_type} — {reason or customer}",
                            debit=0, credit=gross_amt, entry_date=refund_date)
                if cancel_fee > 0:
                    post_ledger('CANCEL_FEE', rid, 'refunds', company,
                                f"Cancellation fee retained",
                                debit=cancel_fee, credit=0, entry_date=refund_date)

            # ACC-05: mark the original sale as REFUNDED / PARTIAL so the report
            # reflects the cancellation and avoids double-counting on statements.
            if rid and sale_id:
                try:
                    _new_status = 'REFUNDED' if refund_type == 'FULL' else 'PARTIAL'
                    execute_db(
                        "UPDATE sales SET status=%s WHERE id=%s AND deleted=FALSE",
                        [_new_status, sale_id]
                    )
                    log_action('UPDATE', 'sales', int(sale_id),
                               f"Status→{_new_status} due to refund #{rid}")
                except Exception as _rfe:
                    logger.error(f"ACC-05 sale status update failed: {_rfe}")

            log_action('CREATE','refunds', rid, f"{company} | {refund_type} | JOD {refund_amt}")
            flash(f'Refund created — JOD {refund_amt:.3f} for {company}.', 'success')
            return redirect(url_for('refunds'))

        elif action == 'process':
            rid    = int(request.form.get('refund_id'))
            status = request.form.get('status','PROCESSED')
            execute_db("""
                UPDATE refunds SET status=%s, processed_by=%s, processed_at=NOW()
                WHERE id=%s
            """, (status, session.get('user_id'), rid))
            log_action('UPDATE','refunds', rid, f"Status→{status}")
            flash(f'Refund #{rid:04d} marked as {status}.', 'success')
            return redirect(url_for('refunds'))

    # GET
    status_f = request.args.get('status','')
    company_f = request.args.get('company','').strip()

    q = """
        SELECT r.*, s.from_loc, s.to_loc, s.travel_date,
               s.service_type, u.username as processed_by_name
        FROM refunds r
        LEFT JOIN sales s ON r.sale_id = s.id
        LEFT JOIN users u ON r.processed_by = u.id
        WHERE 1=1
    """
    params = []
    if status_f:
        q += " AND r.status=%s"; params.append(status_f)
    if company_f:
        q += " AND UPPER(r.company)=UPPER(%s)"; params.append(company_f)
    q += " ORDER BY r.created_at DESC LIMIT 100"
    refund_list = query_db(q, params) or []

    # Summary totals
    totals = query_db("""
        SELECT
            COUNT(*)                                         as total_count,
            COALESCE(SUM(gross_amount),0)                   as total_gross,
            COALESCE(SUM(refund_amount),0)                  as total_refunded,
            COALESCE(SUM(cancellation_fee),0)               as total_fees,
            COALESCE(SUM(CASE WHEN status='PENDING' THEN 1 END),0) as pending_count
        FROM refunds
    """, one=True)

    try:    companies = get_companies_list()
    except: companies = []

    return render_template('refunds.html',
        refunds=refund_list, totals=totals,
        companies=companies,
        filters={'status': status_f, 'company': company_f},
        today=str(date.today()),
    )


@app.route('/supplier-statement')
@finance_required
def supplier_statement():
    """
    Mutual supplier ledger.

    We track BOTH directions with the same company:
      PAYABLE  = we bought from them (we owe them)
      RECEIVABLE = they bought from us (they owe us)

    Running balance = DEBIT − CREDIT
    Positive = they owe us (net receivable)
    Negative = we owe them (net payable)
    """
    supplier  = request.args.get('supplier', '').strip()
    svc_type  = request.args.get('svc_type', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()

    ledger_entries = []

    if supplier:
        sup_upper = supplier.upper().strip()

        def _q(q, p):
            try:    return query_db(q, p) or []
            except:
                try: get_db().rollback()
                except: pass
                return []

        ds = ''   # date filter for sale_date
        dp = ''   # date filter for pay_date
        ps = []   # params for sales queries
        pp = []   # params for payment queries

        if date_from:
            ds += ' AND s.sale_date>=%s';  ps.append(date_from)
            dp += ' AND pay_date>=%s';     pp.append(date_from)
        if date_to:
            ds += ' AND s.sale_date<=%s';  ps.append(date_to)
            dp += ' AND pay_date<=%s';     pp.append(date_to)
        if svc_type:
            ds += " AND UPPER(TRIM(s.service_type))=%s"; ps.append(svc_type.upper())

        # ── CREDIT: we bought outbound ticket from this supplier ─────────────
        for s in _q(
            "SELECT * FROM sales WHERE deleted=FALSE AND is_archived=FALSE "
            "AND UPPER(TRIM(buy_from))=UPPER(%s)" + ds + " ORDER BY sale_date,id",
            [sup_upper] + ps
        ):
            cost = float(s.get('outbound_cost') or s.get('net') or 0)
            if cost > 0:
                fl = s.get('from_loc','') or ''; tl = s.get('to_loc','') or ''
                ledger_entries.append({
                    'date':          s['sale_date'],
                    'ref':           "PUR-%04d" % s['id'],
                    'description':   "Flight " + fl + ("→"+tl if tl else ''),
                    'customer':      s.get('customer','') or '',
                    'svc':           'FLIGHT',
                    'debit':         0.0,
                    'credit':        cost,
                    'type':          'purchase',
                    'id':            s['id'],
                    'from_loc':      fl,
                    'to_loc':        tl,
                    'travel_date':   s.get('travel_date','') or '',
                    'pnr':           s.get('pnr','') or '',
                    'ticket_number': s.get('ticket_number','') or '',
                })

        # ── CREDIT: we bought return ticket from this supplier ───────────────
        # NOTE: No buy_from exclusion — outbound_cost and return_cost are separate fields.
        # A round-trip where buy_from == return_supplier correctly generates two distinct
        # ledger lines: PUR-XXXX (outbound) and RTN-XXXX (return).
        for s in _q(
            "SELECT * FROM sales WHERE deleted=FALSE AND is_archived=FALSE "
            "AND UPPER(TRIM(return_supplier))=UPPER(%s)"
            " AND return_cost > 0" + ds + " ORDER BY sale_date,id",
            [sup_upper] + ps
        ):
            cost  = float(s.get('return_cost') or 0)
            # Return route is reversed: destination→origin
            _fl_orig = s.get('from_loc','') or ''; _tl_orig = s.get('to_loc','') or ''
            _rpnr    = s.get('return_pnr','') or ''
            _rtn     = s.get('return_ticket_number','') or ''
            _desc    = "Return Flight " + _tl_orig + ("→"+_fl_orig if _fl_orig else '')
            if _rpnr: _desc += " [" + _rpnr + "]"
            ledger_entries.append({
                'date':                 s['sale_date'],
                'ref':                  "RTN-%04d" % s['id'],
                'description':          _desc,
                'customer':             s.get('customer','') or '',
                'svc':                  'FLIGHT',
                'debit':                0.0,
                'credit':               cost,
                'type':                 'purchase',
                'id':                   s['id'],
                'from_loc':             _tl_orig,   # return: depart from destination
                'to_loc':               _fl_orig,   # return: arrive at origin
                'travel_date':          s.get('return_date','') or '',
                'pnr':                  _rpnr,
                'ticket_number':        _rtn,
                'return_ticket_number': _rtn,
                'return_pnr':           _rpnr,
            })

        # ── CREDIT: hotel/transfer/visa/insurance from this supplier ─────────
        for supp_field, cost_field, label, svctag in [
            ('hotel_supplier',    'hotel_net',    'Hotel',     'HOTEL'),
            ('transfer_supplier', 'transfer_net', 'Transfer',  'TRANSFER'),
            ('visa_supplier',     'net',          'Visa',      'VISA'),
            ('insurance_supplier','net',          'Insurance', 'INSURANCE'),
        ]:
            for s in _q(
                "SELECT * FROM sales WHERE deleted=FALSE AND is_archived=FALSE "
                "AND UPPER(TRIM(" + supp_field + "))=UPPER(%s)" + ds + " ORDER BY sale_date,id",
                [sup_upper] + ps
            ):
                cost = float(s.get(cost_field) or 0)
                if cost > 0:
                    detail = ''
                    if label == 'Hotel'  and s.get('hotel_name'): detail = ' - ' + s['hotel_name']
                    elif label == 'Visa' and s.get('visa_type'):  detail = ' - ' + s['visa_type']
                    ledger_entries.append({
                        'date':        s['sale_date'],
                        'ref':         label[:3].upper() + "-%04d" % s['id'],
                        'description': label + detail,
                        'customer':    s.get('customer','') or '',
                        'svc':         svctag,
                        'debit':       0.0,
                        'credit':      cost,
                        'type':        'purchase',
                        'id':          s['id'],
                        'from_loc':    '',
                        'to_loc':      s.get('hotel_name','') or '' if label == 'Hotel' else '',
                        'travel_date': s.get('hotel_checkin','') or s.get('travel_date','') or '',
                        'pnr':         '',
                        'ticket_number': '',
                    })

        # ── DEBIT: supplier bought from us (they owe us) ─────────────────────
        for s in _q(
            "SELECT * FROM sales WHERE deleted=FALSE AND is_archived=FALSE "
            "AND UPPER(TRIM(company))=UPPER(%s)" + ds + " ORDER BY sale_date,id",
            [sup_upper] + ps
        ):
            ledger_entries.append({
                'date':          s['sale_date'],
                'ref':           "TXN-%04d" % s['id'],
                'description':   (s.get('service_type') or 'FLIGHT') + " - " + (s.get('customer','') or ''),
                'customer':      s.get('customer','') or '',
                'svc':           s.get('service_type') or 'FLIGHT',
                'debit':         float(s['sell'] or 0),
                'credit':        0.0,
                'type':          'sale',
                'id':            s['id'],
                'from_loc':      s.get('from_loc','') or '',
                'to_loc':        s.get('to_loc','') or '',
                'travel_date':   s.get('travel_date','') or '',
                'pnr':           s.get('pnr','') or '',
                'ticket_number': s.get('ticket_number','') or '',
            })

        # ── CREDIT: payment received from supplier ───────────────────────────
        for p in _q(
            "SELECT * FROM payments WHERE deleted=FALSE AND is_archived=FALSE "
            "AND UPPER(TRIM(company))=UPPER(%s)" + dp + " ORDER BY pay_date,id",
            [sup_upper] + pp
        ):
            ledger_entries.append({
                'date':        p['pay_date'],
                'ref':         "PAY-%04d" % p['id'],
                'description': "Payment received" + (" - " + p['notes'] if p.get('notes') else ''),
                'customer':    '',
                'svc':         '',
                'debit':       0.0,
                'credit':      float(p['amount'] or 0),
                'type':        'payment_in',
                'id':          p['id'],
            })

        # ── DEBIT: we paid this supplier ─────────────────────────────────────
        # ACC-03: filter is_archived=FALSE; move date range into SQL (not Python)
        for sp in _q(
            "SELECT * FROM supplier_payments WHERE deleted=FALSE AND is_archived=FALSE "
            "AND UPPER(TRIM(supplier))=UPPER(%s)" + dp + " ORDER BY pay_date,id",
            [sup_upper] + pp
        ):
            _pm  = sp.get('payment_method','') or ''
            _ref = sp.get('reference_number','') or ''
            _spy_desc = "Payment made"
            if _pm:  _spy_desc += " via " + _pm.replace('_',' ').title()
            if _ref: _spy_desc += " [" + _ref + "]"
            if sp.get('notes'): _spy_desc += " — " + sp['notes']
            ledger_entries.append({
                'date':        sp['pay_date'],
                'ref':         "SPY-%04d" % sp['id'],
                'description': _spy_desc,
                'customer':    '',
                'svc':         sp.get('service_type',''),
                'debit':       float(sp['amount'] or 0),
                'credit':      0.0,
                'type':        'payment_out',
                'id':          sp['id'],
            })

        # ── Normalize: ensure every entry has all display fields ─────────────
        for _e in ledger_entries:
            _e.setdefault('from_loc',      '')
            _e.setdefault('to_loc',        '')
            _e.setdefault('travel_date',   '')
            _e.setdefault('pnr',           '')
            _e.setdefault('ticket_number', '')

        # Opening balance adjustments for this supplier
        for adj in (query_db(
            "SELECT * FROM balance_adjustments WHERE deleted=FALSE AND account_type='SUPPLIER' "
            "AND UPPER(TRIM(account))=UPPER(%s) ORDER BY adj_date,id",
            [supplier]
        ) or []):
            ledger_entries.append({
                'date':        adj['adj_date'],
                'ref':         f"ADJ-{adj['id']:04d}",
                'description': adj['reason'] or 'Balance Adjustment',
                'customer':    adj['notes'] or '',
                'svc':         '',
                'debit':       float(adj['amount']) if adj['entry_type'] == 'DEBIT'  else 0.0,
                'credit':      float(adj['amount']) if adj['entry_type'] == 'CREDIT' else 0.0,
                'type':        'adjustment',
                'travel_date': '',
                'from_loc':    '',
                'to_loc':      '',
                'pnr':         '',
                'ticket_number': '',
                'id':          adj['id'],
            })

        # ── Sort: date → type priority → ref ─────────────────────────────────
        _ORD = {'purchase':0, 'sale':1, 'payment_in':2, 'payment_out':3, 'adjustment':4}
        ledger_entries.sort(key=lambda x:(str(x['date']), _ORD.get(x['type'],9), x['ref']))

        # Renumber payment refs sequentially within this supplier's statement
        _spy_n = 0
        for e in ledger_entries:
            if e['type'] == 'payment_out':
                _spy_n += 1
                e['ref'] = f"PAY-{_spy_n:02d}"

        # ── Running balance ───────────────────────────────────────────────────
        # Positive = they owe us (net receivable)
        # Negative = we owe them (net payable)
        running = 0.0
        for e in ledger_entries:
            running      += e['debit'] - e['credit']
            e['balance']  = round(running, 3)

    # Summary
    summary = {}
    if ledger_entries:
        purchased = sum(e['credit'] for e in ledger_entries if e['type']=='purchase')
        sold_to   = sum(e['debit']  for e in ledger_entries if e['type']=='sale')
        rcvd      = sum(e['credit'] for e in ledger_entries if e['type']=='payment_in')
        paid_out  = sum(e['debit']  for e in ledger_entries if e['type']=='payment_out')
        net_bal   = round(running, 3)
        summary = {
            'total_purchased': purchased,
            'total_sold_to':   sold_to,
            'total_received':  rcvd,
            'total_paid_out':  paid_out,
            'net_payable':     round(purchased  - paid_out, 3),
            'net_receivable':  round(sold_to    - rcvd,     3),
            'net_balance':     net_bal,
            'status': 'RECEIVABLE' if net_bal >  0.005
                 else ('PAYABLE'   if net_bal < -0.005 else 'SETTLED'),
            'entry_count': len(ledger_entries),
        }

    all_suppliers = _get_supplier_list()

    return render_template('supplier_statement.html',
        ledger=ledger_entries,
        summary=summary,
        supplier=supplier,
        all_suppliers=all_suppliers,
        filters={'supplier':supplier,'svc_type':svc_type,
                 'date_from':date_from,'date_to':date_to},
        today=date.today().strftime('%d %B %Y'),
    )



@app.route('/supplier-payment/add', methods=['POST'])
@admin_required
def add_supplier_payment():
    supplier       = request.form.get('supplier','').strip().upper()
    amount_raw     = request.form.get('amount','').strip()
    pay_date       = request.form.get('pay_date', str(date.today())).strip()
    svc_type       = request.form.get('svc_type','FLIGHT').strip().upper()
    notes          = request.form.get('notes','').strip()
    payment_method = request.form.get('payment_method','').strip()
    reference_num  = request.form.get('reference_number','').strip().upper()
    redirect_to    = request.form.get('redirect_to', 'supplier_statement').strip()

    if not supplier or not amount_raw:
        flash('Supplier and amount are required.', 'danger')
        return redirect(url_for('supplier_payments_page'))
    try:
        amount = float(amount_raw)
        if amount <= 0:
            raise ValueError
    except ValueError:
        flash('Invalid payment amount — must be a positive number.', 'danger')
        return redirect(url_for('supplier_payments_page'))

    # ACC-07: RETURNING id eliminates the separate SELECT MAX race condition
    sp_new_id = execute_db("""
        INSERT INTO supplier_payments
               (supplier, amount, pay_date, service_type, notes, payment_method, reference_number)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (supplier, amount, pay_date, svc_type, notes, payment_method, reference_num))

    try:
        _ldesc = f"Payment to {supplier} — {svc_type}"
        if payment_method: _ldesc += f" via {payment_method}"
        if reference_num:  _ldesc += f" [Ref:{reference_num}]"
        post_ledger('SUPPLIER_PAYMENT', sp_new_id or 0, 'supplier_payments',
                    supplier, _ldesc, debit=amount, credit=0, entry_date=pay_date)
    except Exception: pass

    log_action('CREATE', 'supplier_payments', sp_new_id or 0,
               f"Supplier payment: {supplier} JOD {amount} via {payment_method or 'N/A'} ref:{reference_num or '-'}")
    flash(f'Payment of JOD {amount:.3f} recorded for {supplier}.', 'success')

    # Redirect back to the page that submitted the form
    if redirect_to == 'supplier_payments_page':
        return redirect(url_for('supplier_payments_page'))
    return redirect(url_for('supplier_statement', supplier=supplier, svc_type=svc_type))


@app.route('/supplier-payment/<int:pay_id>/delete', methods=['POST'])
@admin_required
def delete_supplier_payment(pay_id):
    sp = query_db('SELECT supplier, amount, pay_date FROM supplier_payments WHERE id=%s', [pay_id], one=True)
    if sp:
        _orig = query_db(
            """SELECT id FROM accounting_entries
               WHERE ref_type='SUPPLIER_PAYMENT' AND ref_id=%s AND ref_table='supplier_payments'
                 AND is_reversed=FALSE
               ORDER BY id DESC LIMIT 1""",
            [pay_id], one=True
        )
        if _orig:
            execute_db("UPDATE accounting_entries SET is_reversed=TRUE WHERE id=%s", [_orig['id']])
        post_ledger('SUPPLIER_PAYMENT_REVERSAL', pay_id, 'supplier_payments', sp['supplier'],
                    f"Reversal of supplier payment #{pay_id} (deleted)",
                    debit=0, credit=sp['amount'], entry_date=sp['pay_date'])
    execute_db('UPDATE supplier_payments SET deleted=TRUE WHERE id=%s', [pay_id])
    log_action('DELETE', 'supplier_payments', pay_id,
               f"Supplier payment deleted: {sp['supplier'] if sp else ''} JOD {sp['amount'] if sp else ''}")
    flash('Payment deleted.', 'success')
    supplier = request.form.get('supplier', sp['supplier'] if sp else '')
    return redirect(url_for('supplier_statement', supplier=supplier))


@app.route('/supplier-payments')
@login_required
def supplier_payments_page():
    """Dedicated Supplier Payment form + recent payments list."""
    prefill_supplier = request.args.get('supplier', '').strip().upper()

    # Recent payments (last 200, not deleted)
    payments = query_db("""
        SELECT sp.id, sp.supplier, sp.amount, sp.pay_date,
               sp.service_type, sp.notes,
               COALESCE(sp.payment_method,'')   AS payment_method,
               COALESCE(sp.reference_number,'') AS reference_number,
               sp.created_at
        FROM supplier_payments sp
        WHERE sp.deleted = FALSE
        ORDER BY sp.pay_date DESC, sp.id DESC
        LIMIT 200
    """) or []

    suppliers = get_suppliers_list()

    return render_template(
        'supplier_payments_page.html',
        payments=payments,
        suppliers=suppliers,
        prefill_supplier=prefill_supplier,
        today=str(date.today()),
    )


# ══════════════════════════════════════════════════════════════════════════════
#  MASTER DATA — Companies, Suppliers, Customers
# ══════════════════════════════════════════════════════════════════════════════

_companies_cache = {'data': None, 'ts': 0.0}

def get_companies_list():
    """All companies — cached for 30 minutes to avoid full sales table scan on every form load."""
    import time
    now = time.time()
    if _companies_cache['data'] is not None and now - _companies_cache['ts'] < 1800:
        return _companies_cache['data']
    try:
        rows = query_db("""
            SELECT name FROM (
                SELECT name FROM master_companies WHERE is_active=TRUE
                UNION
                SELECT UPPER(TRIM(company)) FROM sales
                WHERE deleted=FALSE AND TRIM(COALESCE(company,''))<>''
            ) t ORDER BY name
        """) or []
        result = [r['name'] for r in rows]
    except Exception:
        try: get_db().rollback()
        except: pass
        try:
            rows = query_db("SELECT DISTINCT UPPER(TRIM(company)) as name FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(company,''))<>'' ORDER BY name") or []
            result = [r['name'] for r in rows]
        except Exception:
            result = _companies_cache['data'] or []
    _companies_cache['data'] = result
    _companies_cache['ts']   = now
    return result

def get_suppliers_list(svc_type=None):
    """All active suppliers — from master table + any in sales not yet seeded."""
    q = """
        SELECT name, service_type FROM (
            SELECT name, service_type FROM master_suppliers WHERE is_active=TRUE
            UNION
            SELECT supplier, svc FROM (
                SELECT UPPER(TRIM(buy_from)) AS supplier, 'FLIGHT' AS svc
                  FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(buy_from,''))<>''
                UNION
                -- Return ticket supplier — distinct supplier for the return leg
                SELECT UPPER(TRIM(return_supplier)), 'FLIGHT'
                  FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(return_supplier,''))<>''
                UNION
                SELECT UPPER(TRIM(hotel_supplier)), 'HOTEL'
                  FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(hotel_supplier,''))<>''
                UNION
                SELECT UPPER(TRIM(transfer_supplier)), 'TRANSFER'
                  FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(transfer_supplier,''))<>''
                UNION
                SELECT UPPER(TRIM(visa_supplier)), 'VISA'
                  FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(visa_supplier,''))<>''
                UNION
                SELECT UPPER(TRIM(insurance_supplier)), 'INSURANCE'
                  FROM sales WHERE deleted=FALSE AND TRIM(COALESCE(insurance_supplier,''))<>''
            ) s2 WHERE supplier<>''
        ) t
    """
    params = []
    if svc_type:
        q += " WHERE service_type=%s"
        params.append(svc_type.upper())
    q += " ORDER BY name"
    rows = query_db(q, params) or []
    return rows


@app.route('/api/companies')
@login_required
def api_companies():
    """AJAX: return company list as JSON."""
    from flask import jsonify
    q = request.args.get('q','').strip().upper()
    companies = get_companies_list()
    if q:
        companies = [c for c in companies if q in c.upper()]
    return jsonify({'companies': companies[:50]})


@app.route('/api/suppliers')
@login_required
def api_suppliers():
    """AJAX: return supplier list as JSON, optionally filtered by service type."""
    from flask import jsonify
    svc = request.args.get('svc','').strip()
    q   = request.args.get('q','').strip().upper()
    rows = get_suppliers_list(svc or None)
    data = [{'name': r['name'], 'svc': r['service_type']} for r in rows]
    if q:
        data = [d for d in data if q in d['name'].upper()]
    return jsonify({'suppliers': data[:60]})


@app.route('/api/supplier-email-data/<int:sale_id>')
@login_required
def api_supplier_email_data(sale_id):
    """Return pre-filled mailto data for emailing a supplier about a sale."""
    from flask import jsonify
    sale = query_db('SELECT * FROM sales WHERE id=%s AND deleted=FALSE', [sale_id], one=True)
    if not sale:
        return jsonify({'error': 'Not found'}), 404

    supplier_name = (sale.get('buy_from') or '').strip()
    if not supplier_name and sale.get('hotel_supplier'):
        supplier_name = sale['hotel_supplier'].strip()

    sup = None
    if supplier_name:
        sup = query_db(
            'SELECT * FROM master_suppliers WHERE UPPER(TRIM(name))=UPPER(TRIM(%s)) LIMIT 1',
            [supplier_name], one=True
        )

    txn_number = sale.get('txn_number') or ('T-%04d' % sale_id)
    route = ''
    if sale.get('from_loc') and sale.get('to_loc'):
        route = '%s to %s' % (sale['from_loc'], sale['to_loc'])
    elif sale.get('hotel_name'):
        route = sale['hotel_name']

    contact = (sup['contact_person'] if sup and sup.get('contact_person') else 'Team') if sup else 'Team'

    custom_tpl = (sup.get('email_template') or '').strip() if sup else ''
    if custom_tpl:
        body = custom_tpl
        body = body.replace('{txn_number}', txn_number)
        body = body.replace('{customer}', sale.get('customer') or '')
        body = body.replace('{route}', route)
        body = body.replace('{travel_date}', sale.get('travel_date') or '')
        body = body.replace('{supplier}', supplier_name)
        body = body.replace('{contact}', contact)
        body = body.replace('{agent}', session.get('username') or '')
    else:
        body = (
            'Dear %s,\r\n\r\n'
            'Please process the following booking:\r\n\r\n'
            'Transaction: %s\r\n'
            'Passenger(s): %s\r\n'
            'Route: %s\r\n'
            'Travel Date: %s\r\n'
            'Supplier: %s\r\n\r\n'
            'Regards,\r\n%s'
        ) % (
            contact,
            txn_number,
            sale.get('customer') or '',
            route,
            sale.get('travel_date') or '',
            supplier_name,
            session.get('username') or ''
        )

    return jsonify({
        'to':       (sup['email']         if sup and sup.get('email')         else ''),
        'cc':       (sup['cc_email']       if sup and sup.get('cc_email')       else ''),
        'subject':  'Booking Request - %s - %s' % (txn_number, route),
        'body':     body,
        'supplier': supplier_name,
    })


@app.route('/api/supplier-email-data-by-name')
@login_required
def api_supplier_email_data_by_name():
    """Return pre-filled email fields for a supplier looked up by name."""
    from flask import jsonify
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'error': 'name required'}), 400

    sup = query_db(
        'SELECT * FROM master_suppliers WHERE UPPER(TRIM(name))=UPPER(TRIM(%s)) LIMIT 1',
        [name], one=True
    )

    contact = (sup['contact_person'] if sup and sup.get('contact_person') else 'Team')

    custom_tpl = (sup.get('email_template') or '').strip() if sup else ''
    if custom_tpl:
        body = custom_tpl
        body = body.replace('{contact}', contact)
        body = body.replace('{supplier}', name)
        body = body.replace('{agent}', session.get('username') or '')
    else:
        body = (
            'Dear %s,\r\n\r\n'
            'Please process the following booking request.\r\n\r\n'
            'Regards,\r\nAlsondos Travel'
        ) % contact

    return jsonify({
        'email':   (sup['email']         if sup and sup.get('email')    else ''),
        'cc':      (sup['cc_email']       if sup and sup.get('cc_email') else ''),
        'subject': 'Booking Request',
        'body':    body,
    })


# ── Master Data Management Pages ─────────────────────────────────────────────


def ensure_master_tables():
    """Create master_companies / master_suppliers if they don't exist yet."""
    try:
        db  = get_db()
        cur = db.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS master_companies (
                id         SERIAL PRIMARY KEY,
                name       TEXT NOT NULL UNIQUE,
                phone      TEXT DEFAULT '',
                email      TEXT DEFAULT '',
                address    TEXT DEFAULT '',
                notes      TEXT DEFAULT '',
                is_active  BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS master_suppliers (
                id           SERIAL PRIMARY KEY,
                name         TEXT NOT NULL,
                service_type TEXT DEFAULT 'FLIGHT',
                phone        TEXT DEFAULT '',
                email        TEXT DEFAULT '',
                address      TEXT DEFAULT '',
                notes        TEXT DEFAULT '',
                is_active    BOOLEAN DEFAULT TRUE,
                created_at   TIMESTAMP DEFAULT NOW()
            )
        """)
        db.commit()
    except Exception as _e:
        try: get_db().rollback()
        except: pass
        logger.warning(f"ensure_master_tables: {_e}")

@app.route('/master/companies', methods=['GET', 'POST'])
@admin_required
def master_companies():
    ensure_master_tables()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name','').strip().upper()
            if name:
                try:
                    execute_db(
                        "INSERT INTO master_companies (name,phone,email,address,notes) VALUES (%s,%s,%s,%s,%s)",
                        (name,
                         request.form.get('phone','').strip(),
                         request.form.get('email','').strip(),
                         request.form.get('address','').strip(),
                         request.form.get('notes','').strip())
                    )
                    flash(f'Company {name} added.', 'success')
                except Exception as ex:
                    flash(f'Error: {ex}', 'danger')
        elif action == 'toggle':
            cid = request.form.get('id')
            execute_db('UPDATE master_companies SET is_active = NOT is_active WHERE id=%s', [cid])
            flash('Company status updated.', 'success')
        elif action == 'delete':
            cid = request.form.get('id')
            execute_db('DELETE FROM master_companies WHERE id=%s', [cid])
            flash('Company deleted.', 'success')
        return redirect(url_for('master_companies'))
    try:
        companies = query_db('SELECT * FROM master_companies ORDER BY name') or []
    except Exception:
        try: get_db().rollback()
        except: pass
        companies = []
    return render_template('master_companies.html', companies=companies)


@app.route('/master/suppliers', methods=['GET', 'POST'])
@admin_required
def master_suppliers():
    ensure_master_tables()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name','').strip().upper()
            svc  = request.form.get('service_type','FLIGHT').strip().upper()
            if name:
                try:
                    execute_db(
                        "INSERT INTO master_suppliers (name,service_type,phone,email,cc_email,contact_person,notes) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                        (name, svc,
                         request.form.get('phone','').strip(),
                         request.form.get('email','').strip(),
                         request.form.get('cc_email','').strip(),
                         request.form.get('contact_person','').strip(),
                         request.form.get('notes','').strip())
                    )
                    flash(f'Supplier {name} added.', 'success')
                except Exception as ex:
                    flash(f'Error: {ex}', 'danger')
        elif action == 'edit':
            sid = request.form.get('id')
            if sid:
                try:
                    execute_db(
                        """UPDATE master_suppliers SET phone=%s, email=%s, cc_email=%s,
                           contact_person=%s, email_template=%s, notes=%s WHERE id=%s""",
                        (request.form.get('phone','').strip(),
                         request.form.get('email','').strip(),
                         request.form.get('cc_email','').strip(),
                         request.form.get('contact_person','').strip(),
                         request.form.get('email_template','').strip(),
                         request.form.get('notes','').strip(),
                         sid)
                    )
                    flash('Supplier updated.', 'success')
                except Exception as ex:
                    flash(f'Error: {ex}', 'danger')
        elif action == 'toggle':
            sid = request.form.get('id')
            execute_db('UPDATE master_suppliers SET is_active = NOT is_active WHERE id=%s', [sid])
            flash('Supplier status updated.', 'success')
        elif action == 'delete':
            sid = request.form.get('id')
            execute_db('DELETE FROM master_suppliers WHERE id=%s', [sid])
            flash('Supplier deleted.', 'success')
        return redirect(url_for('master_suppliers'))
    suppliers = query_db('SELECT * FROM master_suppliers ORDER BY service_type, name') or []
    return render_template('master_suppliers.html', suppliers=suppliers)

@app.route('/api/company-transactions')
@login_required
def api_company_transactions():
    """AJAX: return full transaction details for a company (invoice form auto-load)."""
    company = request.args.get('company', '').strip()
    if not company:
        return {'transactions': []}
    rows = query_db("""
        SELECT id, sale_date, customer, from_loc, to_loc, via, sell, status, tickets,
               service_type, airline, pnr, baggage, trip_type,
               travel_date, return_date, buy_from,
               hotel_name, hotel_room, hotel_meal, hotel_checkin, hotel_checkout, hotel_nights,
               transfer_type, transfer_vehicle, transfer_pickup, transfer_supplier,
               visa_type, passport_number, visa_status, visa_supplier,
               insurance_type, insurance_supplier,
               passengers_json, tours_json, remarks
        FROM sales
        WHERE deleted=FALSE AND is_archived=FALSE AND company=%s
        ORDER BY sale_date DESC
        LIMIT 100
    """, [company]) or []
    data = []
    for r in rows:
        import json as _j
        # Parse JSON arrays safely
        pax = []
        tours = []
        try: pax   = _j.loads(r['passengers_json'] or '[]')
        except: pass
        try: tours = _j.loads(r['tours_json'] or '[]')
        except: pass
        data.append({
            'id':           r['id'],
            'date':         r['sale_date'],
            'customer':     r['customer'],
            'route':        f"{r['from_loc'] or ''} → {r['to_loc'] or ''}",
            'sell':         float(r['sell'] or 0),
            'status':       r['status'],
            'tickets':      r['tickets'],
            # All fields for description building
            'svc':          (r['service_type'] or 'FLIGHT').upper(),
            'airline':      r['airline'] or '',
            'pnr':          r['pnr'] or '',
            'baggage':      r['baggage'] or '',
            'trip_type':    r['trip_type'] or '',
            'travel_date':  r['travel_date'] or '',
            'return_date':  r['return_date'] or '',
            'buy_from':     r['buy_from'] or '',
            'hotel_name':   r['hotel_name'] or '',
            'hotel_room':   r['hotel_room'] or '',
            'hotel_meal':   r['hotel_meal'] or '',
            'hotel_checkin':  r['hotel_checkin'] or '',
            'hotel_checkout': r['hotel_checkout'] or '',
            'hotel_nights':   r['hotel_nights'] or 0,
            'transfer_type':    r['transfer_type'] or '',
            'transfer_vehicle': r['transfer_vehicle'] or '',
            'transfer_pickup':  r['transfer_pickup'] or '',
            'visa_type':      r['visa_type'] or '',
            'passport_number':r['passport_number'] or '',
            'insurance_type': r['insurance_type'] or '',
            'passengers':  pax,
            'tours':       tours,
            'remarks':     r['remarks'] or '',
        })
    from flask import jsonify
    return jsonify({'transactions': data, 'company': company})

# ══════════════════════════════════════════════════════════════════════════════
#  HOTEL VOUCHERS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/vouchers')
@login_required
def voucher_list():
    uid      = session['user_id']
    is_admin = session.get('user_role') == 'admin'
    page     = max(1, int(request.args.get('page', 1)))
    q = 'SELECT * FROM vouchers WHERE deleted=FALSE'
    params = []
    if not is_admin:
        q += ' AND created_by_id=%s'; params.append(uid)
    q += ' ORDER BY created_at DESC'
    vouchers, total_rows, total_pages = paginate(q, params, page)
    return render_template('vouchers.html',
        vouchers=vouchers, page=page,
        total_pages=total_pages, total_rows=total_rows)


@app.route('/vouchers/new', methods=['GET', 'POST'])
@login_required
def new_voucher():
    sale_id = request.args.get('sale_id', '')
    sale    = query_db('SELECT * FROM sales WHERE id=%s AND deleted=FALSE', [sale_id], one=True) if sale_id else None
    if request.method == 'POST':
        vcn     = next_doc_number('VCH')
        raw_sid = request.form.get('sale_id','').strip()
        sid     = int(raw_sid) if raw_sid and raw_sid.isdigit() else None
        new_id = execute_db("""
            INSERT INTO vouchers
            (voucher_number,sale_id,created_by_id,created_by,
             guest_name,num_guests,hotel_name,hotel_address,hotel_phone,hotel_contact,
             room_type,meal_plan,checkin_date,checkout_date,nights,
             include_transfer,include_tours,include_insurance,
             arrival_flight,pickup_sign,driver_contact,pickup_time,vehicle_type,
             emergency_contact,cancellation_policy,remarks)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            vcn, sid,
            session['user_id'], session.get('username',''),
            request.form.get('guest_name','').strip(),
            int(request.form.get('num_guests', 1)),
            request.form.get('hotel_name','').strip(),
            request.form.get('hotel_address','').strip(),
            request.form.get('hotel_phone','').strip(),
            request.form.get('hotel_contact','').strip(),
            request.form.get('room_type','').strip(),
            request.form.get('meal_plan','').strip(),
            request.form.get('checkin_date','').strip(),
            request.form.get('checkout_date','').strip(),
            int(request.form.get('nights', 0) or 0),
            request.form.get('include_transfer') == 'on',
            request.form.get('include_tours') == 'on',
            request.form.get('include_insurance') == 'on',
            request.form.get('arrival_flight','').strip(),
            request.form.get('pickup_sign','').strip(),
            request.form.get('driver_contact','').strip(),
            request.form.get('pickup_time','').strip(),
            request.form.get('vehicle_type','').strip(),
            request.form.get('emergency_contact','').strip(),
            request.form.get('cancellation_policy','').strip(),
            request.form.get('remarks','').strip()
        ))
        log_action('CREATE', 'vouchers', new_id, f"Voucher {vcn}")
        flash(f'Voucher {vcn} created.', 'success')
        return redirect(url_for('view_voucher', vch_id=new_id))
    return render_template('voucher_form.html', sale=sale, today=str(date.today()))


@app.route('/vouchers/<int:vch_id>')
@login_required
def view_voucher(vch_id):
    vch = query_db('SELECT * FROM vouchers WHERE id=%s AND deleted=FALSE', [vch_id], one=True)
    if not vch:
        flash('Voucher not found.', 'danger')
        return redirect(url_for('voucher_list'))
    if session.get('user_role') != 'admin' and vch['created_by_id'] != session['user_id']:
        flash('Access denied.', 'danger')
        return redirect(url_for('voucher_list'))
    sale = query_db('SELECT * FROM sales WHERE id=%s', [vch['sale_id']], one=True) if vch['sale_id'] else None
    return render_template('voucher_view.html', vch=vch, sale=sale)


@app.route('/vouchers/<int:vch_id>/delete', methods=['POST'])
@login_required
def delete_voucher(vch_id):
    vch = query_db('SELECT * FROM vouchers WHERE id=%s', [vch_id], one=True)
    if vch and (session.get('user_role') == 'admin' or vch['created_by_id'] == session['user_id']):
        execute_db('UPDATE vouchers SET deleted=TRUE WHERE id=%s', [vch_id])
        log_action('DELETE', 'vouchers', vch_id, f"Voucher {vch['voucher_number']}")
        flash('Voucher deleted.', 'success')
    return redirect(url_for('voucher_list'))


# ══════════════════════════════════════════════════════════════════════════════
#  PACKAGES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin/sync-packages', methods=['POST'])
@admin_required
def admin_sync_packages():
    """One-time backfill: create packages table rows for all PACKAGE-type sales."""
    sales = query_db(
        "SELECT id FROM sales WHERE UPPER(COALESCE(service_type,'FLIGHT'))='PACKAGE' AND deleted=FALSE"
    ) or []
    synced = skipped = errors = 0
    for s in sales:
        existing = query_db(
            "SELECT id FROM packages WHERE sale_id=%s AND deleted=FALSE", [s['id']], one=True
        )
        if existing:
            skipped += 1
            continue
        result = sync_package_from_sale(s['id'])
        if result:
            synced += 1
        else:
            errors += 1
    flash(f'Sync complete: {synced} created, {skipped} already existed, {errors} errors.', 'success')
    return redirect(url_for('package_list'))


@app.route('/packages')
@login_required
def package_list():
    uid      = session['user_id']
    is_admin = session.get('user_role') == 'admin'
    page     = max(1, int(request.args.get('page', 1)))
    q = 'SELECT * FROM packages WHERE deleted=FALSE'
    params = []
    if not is_admin:
        q += ' AND created_by_id=%s'; params.append(uid)
    q += ' ORDER BY created_at DESC'
    pkgs, total_rows, total_pages = paginate(q, params, page)
    return render_template('packages.html', pkgs=pkgs, page=page,
                           total_pages=total_pages, total_rows=total_rows,
                           is_admin=is_admin)


@app.route('/packages/new', methods=['GET', 'POST'])
@login_required
def new_package():
    sale_id  = request.args.get('sale_id', '')
    sale     = query_db('SELECT * FROM sales WHERE id=%s AND deleted=FALSE', [sale_id], one=True) if sale_id else None
    companies= [r['company'] for r in query_db(
        'SELECT DISTINCT company FROM sales WHERE deleted=FALSE AND is_archived=FALSE ORDER BY company'
    )]
    if request.method == 'POST':
        net  = float(request.form.get('net_cost', 0) or 0)
        sell = float(request.form.get('sell_price', 0) or 0)
        pkg_num = next_doc_number('PKG')
        raw_sid = request.form.get('sale_id','').strip()
        sid     = int(raw_sid) if raw_sid and raw_sid.isdigit() else None
        new_id  = execute_db("""
            INSERT INTO packages
            (package_number,sale_id,created_by_id,created_by,company,customer,package_date,
             airline,flight_route,departure_date,return_date,baggage,pnr,
             hotel_name,room_type,meal_plan,checkin_date,checkout_date,nights,
             transfer_type,pickup_time,vehicle_type,
             tour_name,tour_date,tour_included,
             net_cost,sell_price,profit,status,notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            pkg_num, sid,
            session['user_id'], session.get('username',''),
            request.form.get('company','').upper().strip(),
            request.form.get('customer','').upper().strip(),
            request.form.get('package_date', str(date.today())),
            request.form.get('airline','').strip(),
            request.form.get('flight_route','').strip(),
            request.form.get('departure_date','').strip(),
            request.form.get('return_date','').strip(),
            request.form.get('baggage','').strip(),
            request.form.get('pnr','').upper().strip(),
            request.form.get('hotel_name','').strip(),
            request.form.get('room_type','').strip(),
            request.form.get('meal_plan','').strip(),
            request.form.get('checkin_date','').strip(),
            request.form.get('checkout_date','').strip(),
            int(request.form.get('nights', 0) or 0),
            request.form.get('transfer_type','').strip(),
            request.form.get('pickup_time','').strip(),
            request.form.get('vehicle_type','').strip(),
            request.form.get('tour_name','').strip(),
            request.form.get('tour_date','').strip(),
            request.form.get('tour_included') == 'on',
            net, sell, round(sell - net, 3),
            'ACTIVE',
            request.form.get('notes','').strip()
        ))
        log_action('CREATE', 'packages', new_id, f"Package {pkg_num}")
        flash(f'Package {pkg_num} created.', 'success')
        return redirect(url_for('view_package', pkg_id=new_id))
    return render_template('package_form.html',
        sale=sale, companies=companies, today=str(date.today()))


@app.route('/packages/<int:pkg_id>')
@login_required
def view_package(pkg_id):
    pkg = query_db('SELECT * FROM packages WHERE id=%s AND deleted=FALSE', [pkg_id], one=True)
    if not pkg:
        flash('Package not found.', 'danger')
        return redirect(url_for('package_list'))
    if session.get('user_role') != 'admin' and pkg['created_by_id'] != session['user_id']:
        flash('Access denied.', 'danger')
        return redirect(url_for('package_list'))
    return render_template('package_view.html', pkg=pkg)


@app.route('/packages/<int:pkg_id>/delete', methods=['POST'])
@login_required
def delete_package(pkg_id):
    pkg = query_db('SELECT * FROM packages WHERE id=%s', [pkg_id], one=True)
    if pkg and (session.get('user_role') == 'admin' or pkg['created_by_id'] == session['user_id']):
        execute_db('UPDATE packages SET deleted=TRUE WHERE id=%s', [pkg_id])
        log_action('DELETE', 'packages', pkg_id, f"Package {pkg['package_number']}")
        flash('Package deleted.', 'success')
    return redirect(url_for('package_list'))


# ══════════════════════════════════════════════════════════════════════════════
#  COMMISSION REPORT
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/my/commission')
@login_required
def my_commission():
    uid  = session['user_id']
    year = int(request.args.get('year', date.today().year))

    user_row = query_db('SELECT * FROM users WHERE id=%s', [uid], one=True)
    rate = float(user_row['commission_rate'] or 20) if user_row else 20.0

    monthly = query_db("""
        SELECT to_char(to_date(sale_date,'YYYY-MM-DD'),'MM') as month,
               to_char(to_date(sale_date,'YYYY-MM-DD'),'Month') as month_name,
               COALESCE(SUM(sell),0)   as total_sell,
               COALESCE(SUM(profit),0) as total_profit,
               COUNT(*) as count
        FROM sales
        WHERE deleted=FALSE AND created_by_user_id=%s
          AND to_char(to_date(sale_date,'YYYY-MM-DD'),'YYYY') = %s
        GROUP BY month, month_name ORDER BY month
    """, [uid, str(year)])

    monthly = [dict(r) for r in monthly]
    for m in monthly:
        m['commission'] = round(float(m['total_profit'] or 0) * rate / 100, 3)

    totals = {
        'sell':       sum(float(m['total_sell'] or 0)   for m in monthly),
        'profit':     sum(float(m['total_profit'] or 0) for m in monthly),
        'commission': sum(m['commission'] for m in monthly),
        'count':      sum(int(m['count'] or 0) for m in monthly),
    }
    return render_template('commission_report.html',
        monthly=monthly, totals=totals, rate=rate, year=year)


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN — EMPLOYEE OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin/employees')
@admin_required
def admin_employees():
    employees = query_db("""
        SELECT u.id, u.username, u.full_name, u.email, u.role,
               u.commission_rate, u.is_active, u.created_at,
               COALESCE(s.txn_count, 0) as txn_count,
               COALESCE(s.total_sell, 0) as total_sell,
               COALESCE(s.total_profit, 0) as total_profit
        FROM users u
        LEFT JOIN (
            SELECT created_by_user_id,
                   COUNT(*) as txn_count,
                   SUM(sell) as total_sell,
                   SUM(profit) as total_profit
            FROM sales WHERE deleted=FALSE AND is_archived=FALSE
            GROUP BY created_by_user_id
        ) s ON u.id = s.created_by_user_id
        ORDER BY total_sell DESC NULLS LAST
    """)
    employees = [dict(e) for e in employees]
    rate = 20.0
    for e in employees:
        r = float(e['commission_rate'] or 20)
        e['commission'] = round(float(e['total_profit'] or 0) * r / 100, 3)

    total_sell   = sum(float(e['total_sell']   or 0) for e in employees)
    total_profit = sum(float(e['total_profit'] or 0) for e in employees)
    total_comm   = sum(e['commission'] for e in employees)

    return render_template('admin_employees.html',
        employees=employees,
        total_sell=total_sell, total_profit=total_profit, total_comm=total_comm,
    )


@app.route('/admin/employees/<int:uid>/sales')
@admin_required
def admin_employee_sales(uid):
    emp  = query_db('SELECT * FROM users WHERE id=%s', [uid], one=True)
    if not emp:
        flash('User not found.', 'danger')
        return redirect(url_for('admin_employees'))
    page = max(1, int(request.args.get('page', 1)))
    q    = 'SELECT * FROM sales WHERE deleted=FALSE AND created_by_user_id=%s ORDER BY sale_date DESC, id DESC'
    sales, total_rows, total_pages = paginate(q, [uid], page)
    totals = {
        'sell':   sum(float(s['sell']   or 0) for s in sales),
        'profit': sum(float(s['profit'] or 0) for s in sales),
        'count':  total_rows,
    }
    rate = float(emp['commission_rate'] or 20)
    totals['commission'] = round(totals['profit'] * rate / 100, 3)
    return render_template('admin_employee_sales.html',
        emp=emp, sales=sales, totals=totals,
        page=page, total_pages=total_pages, total_rows=total_rows, rate=rate,
    )


@app.route('/admin/employees/<int:uid>/commission-rate', methods=['POST'])
@admin_required
def set_commission_rate(uid):
    try:
        rate = float(request.form.get('commission_rate', 20))
        rate = max(0, min(100, rate))
        execute_db('UPDATE users SET commission_rate=%s WHERE id=%s', (rate, uid))
        flash(f'Commission rate updated to {rate}%.', 'success')
    except ValueError:
        flash('Invalid commission rate.', 'danger')
    return redirect(url_for('admin_employees'))


@app.route('/admin/employees/<int:uid>/update', methods=['POST'])
@admin_required
def admin_update_employee(uid):
    full_name = request.form.get('full_name','').strip()
    email     = request.form.get('email','').strip()
    phone     = request.form.get('phone','').strip()
    execute_db('UPDATE users SET full_name=%s, email=%s, phone=%s WHERE id=%s',
               (full_name, email, phone, uid))
    flash('Employee profile updated.', 'success')
    return redirect(url_for('admin_employees'))


# ── Balance Adjustments ───────────────────────────────────────────────────────
@app.route('/admin/adjustments', methods=['GET', 'POST'])
@admin_required
def balance_adjustments():
    if request.method == 'POST':
        action = request.form.get('action', 'create')

        if action == 'create':
            account      = request.form.get('account', '').upper().strip()
            account_type = request.form.get('account_type', 'COMPANY')
            adj_date     = request.form.get('adj_date', str(date.today()))
            amount       = round(float(request.form.get('amount', 0) or 0), 3)
            entry_type   = request.form.get('entry_type', 'CREDIT')
            reason       = request.form.get('reason', '').strip()
            notes        = request.form.get('notes', '').strip()

            if not account:
                flash('Account name is required.', 'danger')
                return redirect(url_for('balance_adjustments'))
            if amount <= 0:
                flash('Amount must be greater than zero.', 'danger')
                return redirect(url_for('balance_adjustments'))
            if not reason:
                flash('Reason is required.', 'danger')
                return redirect(url_for('balance_adjustments'))

            adj_id = execute_db("""
                INSERT INTO balance_adjustments
                (account, account_type, adj_date, amount, entry_type, reason, notes, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (account, account_type, adj_date, amount, entry_type,
                  reason, notes, session.get('user_id')))

            # ACC-05 FIX: post to shadow ledger so validate_ledger_balance includes adjustments
            if adj_id:
                _adj_dr = amount if entry_type == 'DEBIT'  else 0
                _adj_cr = amount if entry_type == 'CREDIT' else 0
                post_ledger(
                    'ADJUSTMENT', adj_id, 'balance_adjustments', account.upper().strip(),
                    f"{account_type} Adjustment: {reason}",
                    debit=_adj_dr, credit=_adj_cr, entry_date=adj_date
                )

            log_action('CREATE', 'balance_adjustments', adj_id,
                       f"{account_type} {account} | {entry_type} JOD {amount:.3f} | {reason}")
            flash(f'Adjustment created — {entry_type} JOD {amount:.3f} for {account}.', 'success')
            return redirect(url_for('balance_adjustments'))

        elif action == 'delete':
            adj_id = int(request.form.get('adj_id', 0))
            adj = query_db('SELECT account, account_type, amount, entry_type, adj_date FROM balance_adjustments WHERE id=%s', [adj_id], one=True)
            execute_db('UPDATE balance_adjustments SET deleted=TRUE WHERE id=%s', [adj_id])
            # ACC-05 FIX: reverse the ledger entry when adjustment is deleted
            if adj:
                _orig_entry = query_db(
                    "SELECT id FROM accounting_entries WHERE ref_type='ADJUSTMENT' AND ref_id=%s AND is_reversed=FALSE ORDER BY id DESC LIMIT 1",
                    [adj_id], one=True
                )
                if _orig_entry:
                    execute_db("UPDATE accounting_entries SET is_reversed=TRUE WHERE id=%s", [_orig_entry['id']])
                # Post reversal
                _rev_dr = float(adj['amount']) if adj['entry_type'] == 'CREDIT' else 0
                _rev_cr = float(adj['amount']) if adj['entry_type'] == 'DEBIT'  else 0
                post_ledger(
                    'ADJUSTMENT_REVERSAL', adj_id, 'balance_adjustments',
                    str(adj['account']).upper().strip(),
                    f"Reversal of {adj['account_type']} Adjustment #{adj_id} (deleted)",
                    debit=_rev_dr, credit=_rev_cr, entry_date=adj['adj_date']
                )
            log_action('DELETE', 'balance_adjustments', adj_id,
                       f"Deleted adjustment: {adj['account'] if adj else ''} {adj['entry_type'] if adj else ''} JOD {adj['amount'] if adj else ''}")
            flash('Adjustment deleted.', 'success')
            return redirect(url_for('balance_adjustments'))

    # GET — list all adjustments
    adjustments = query_db("""
        SELECT a.*, u.username as created_by_name
        FROM balance_adjustments a
        LEFT JOIN users u ON a.created_by = u.id
        WHERE a.deleted=FALSE
        ORDER BY a.adj_date DESC, a.id DESC
    """) or []

    companies = get_companies_list()
    suppliers = get_suppliers_list()

    return render_template('balance_adjustments.html',
        adjustments=adjustments,
        companies=companies,
        suppliers=suppliers,
        today=str(date.today()),
    )


# ── Contract generation ────────────────────────────────────────────────────────
@app.route('/contract/<int:sale_id>')
@login_required
def view_contract(sale_id):
    sale = query_db('SELECT * FROM sales WHERE id=%s AND deleted=FALSE', [sale_id], one=True)
    if not sale:
        flash('Transaction not found.', 'danger')
        return redirect(url_for('sales_report'))
    sale = safe_sale(sale)
    passengers = []
    try:
        passengers = json.loads(sale.get('passengers_json') or '[]')
    except Exception:
        pass
    tours = []
    try:
        tours = json.loads(sale.get('tours_json') or '[]')
    except Exception:
        pass
    return render_template('contract_view.html', sale=sale, passengers=passengers, tours=tours,
                           today=str(date.today()))




# ── Health check ──────────────────────────────────────────────────────────────
@app.route('/health')
@csrf.exempt
def health():
    """Lightweight liveness probe for Nginx / uptime monitors."""
    try:
        query_db('SELECT 1', one=True)
        return jsonify(status='ok', db='connected'), 200
    except Exception as e:
        return jsonify(status='error', db=str(e)), 500
